"""설계도서 자료분류 GUI 도구

검토위원 배정 엑셀에 따라 접수 폴더의 설계도서를 검토위원별로 분류합니다.
실행: python classify_gui.py
"""

import os
import re
import shutil
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from pathlib import Path

from openpyxl import load_workbook

MGMT_PATTERN = re.compile(r"^(\d{4}-\d{4})")


class ClassifyApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("설계도서 자료분류")
        self.root.geometry("700x650")
        self.root.resizable(True, True)

        # 변수
        self.source_var = tk.StringVar()
        self.target_var = tk.StringVar()
        self.assign_var = tk.StringVar()

        self._build_ui()

    def _build_ui(self):
        # 제목
        title = tk.Label(self.root, text="설계도서 자료분류", font=("", 16, "bold"))
        title.pack(pady=(15, 5))

        desc = tk.Label(self.root, text="접수 폴더의 설계도서를 검토위원별 폴더로 분류합니다", fg="gray")
        desc.pack(pady=(0, 15))

        # 입력 영역
        input_frame = tk.Frame(self.root)
        input_frame.pack(fill="x", padx=20)

        # 배정 엑셀
        self._add_path_row(input_frame, "배정 엑셀:", self.assign_var, 0, file_mode=True)

        # 소스 폴더
        self._add_path_row(input_frame, "접수 폴더:", self.source_var, 1)

        # 대상 폴더
        self._add_path_row(input_frame, "배포 폴더:", self.target_var, 2)

        # 실행 버튼
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=15)

        self.run_btn = tk.Button(
            btn_frame, text="분류 실행", command=self._run_classify,
            font=("", 12, "bold"), bg="#2563eb", fg="white",
            padx=30, pady=8, cursor="hand2"
        )
        self.run_btn.pack(side="left", padx=5)

        self.copy_btn = tk.Button(
            btn_frame, text="관리번호 복사", command=self._copy_mgmt_nos,
            font=("", 11), padx=15, pady=8, cursor="hand2", state="disabled"
        )
        self.copy_btn.pack(side="left", padx=5)

        # 로그 영역
        log_label = tk.Label(self.root, text="처리 결과:", anchor="w")
        log_label.pack(fill="x", padx=20)

        self.log_text = scrolledtext.ScrolledText(
            self.root, height=18, font=("Courier", 11), wrap="word"
        )
        self.log_text.pack(fill="both", expand=True, padx=20, pady=(5, 20))

        # 분류된 관리번호 저장
        self.classified_mgmt_nos: list[str] = []

    def _add_path_row(self, parent: tk.Frame, label: str, var: tk.StringVar, row: int, file_mode: bool = False):
        tk.Label(parent, text=label, width=10, anchor="e").grid(row=row, column=0, padx=(0, 5), pady=5)
        entry = tk.Entry(parent, textvariable=var, width=50)
        entry.grid(row=row, column=1, padx=5, pady=5, sticky="ew")
        parent.columnconfigure(1, weight=1)

        if file_mode:
            cmd = lambda: var.set(filedialog.askopenfilename(
                filetypes=[("Excel", "*.xlsx *.xls")],
                title="배정 엑셀 선택"
            ))
        else:
            cmd = lambda: var.set(filedialog.askdirectory(title=label.replace(":", "")))

        btn = tk.Button(parent, text="찾기", command=cmd, cursor="hand2")
        btn.grid(row=row, column=2, padx=5, pady=5)

    def _log(self, msg: str):
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.root.update_idletasks()

    def _run_classify(self):
        source = self.source_var.get().strip()
        target = self.target_var.get().strip()
        assign_path = self.assign_var.get().strip()

        if not assign_path:
            messagebox.showwarning("경고", "배정 엑셀을 선택해주세요")
            return
        if not source:
            messagebox.showwarning("경고", "접수 폴더를 선택해주세요")
            return
        if not target:
            messagebox.showwarning("경고", "배포 폴더를 선택해주세요")
            return

        self.run_btn.config(state="disabled", text="처리 중...")
        self.log_text.delete("1.0", "end")
        self.classified_mgmt_nos.clear()

        # 별도 스레드에서 실행 (UI 멈춤 방지)
        thread = threading.Thread(target=self._do_classify, args=(source, target, assign_path))
        thread.start()

    def _do_classify(self, source: str, target: str, assign_path: str):
        try:
            self._log("=" * 50)
            self._log("설계도서 자료분류 시작")
            self._log("=" * 50)
            self._log(f"배정 엑셀: {assign_path}")
            self._log(f"접수 폴더: {source}")
            self._log(f"배포 폴더: {target}")
            self._log("")

            # 배정 로드
            assignment = self._load_assignment(assign_path)
            self._log(f"검토위원 배정: {len(assignment)}건 로드")
            self._log("")

            # 분류 실행
            source_path = Path(source)
            target_path = Path(target)
            target_path.mkdir(parents=True, exist_ok=True)

            classified = 0
            skipped = 0

            items = sorted(source_path.iterdir())
            self._log(f"소스 폴더 항목: {len(items)}개")
            self._log("")

            for item in items:
                mgmt_no = self._extract_mgmt_no(item.name)
                if not mgmt_no:
                    self._log(f"  ⊘ 스킵: {item.name} (관리번호 인식 불가)")
                    skipped += 1
                    continue

                reviewer = assignment.get(mgmt_no)
                if not reviewer:
                    self._log(f"  ⊘ 스킵: {item.name} (검토위원 미배정)")
                    skipped += 1
                    continue

                # 검토위원 폴더 생성 + 복사
                reviewer_dir = target_path / reviewer
                reviewer_dir.mkdir(exist_ok=True)
                dest = reviewer_dir / item.name

                if item.is_dir():
                    if dest.exists():
                        shutil.rmtree(dest)
                    shutil.copytree(item, dest)
                else:
                    shutil.copy2(item, dest)

                self._log(f"  ✓ {item.name} → {reviewer}/")
                self.classified_mgmt_nos.append(mgmt_no)
                classified += 1

            # 결과 요약
            self._log("")
            self._log("=" * 50)
            self._log(f"분류 완료: {classified}건 / 스킵: {skipped}건")
            self._log("")

            # 검토위원별 건수
            self._log("--- 검토위원별 건수 ---")
            for reviewer_dir in sorted(target_path.iterdir()):
                if reviewer_dir.is_dir():
                    count = len(list(reviewer_dir.iterdir()))
                    self._log(f"  {reviewer_dir.name}: {count}건")

            # 관리번호 목록
            self._log("")
            self._log("--- 접수 관리번호 (웹에 붙여넣기용) ---")
            self._log("\n".join(self.classified_mgmt_nos))

            self.root.after(0, lambda: self.copy_btn.config(state="normal"))
            messagebox.showinfo("완료", f"분류 완료: {classified}건")

        except Exception as e:
            self._log(f"\n오류 발생: {e}")
            messagebox.showerror("오류", str(e))

        finally:
            self.root.after(0, lambda: self.run_btn.config(state="normal", text="분류 실행"))

    def _load_assignment(self, path: str) -> dict[str, str]:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        mapping = {}
        for row in ws.iter_rows(min_row=2):
            mgmt = str(row[0].value).strip() if row[0].value else None
            name = str(row[1].value).strip() if len(row) > 1 and row[1].value else None
            if mgmt and name:
                mapping[mgmt] = name
        wb.close()
        return mapping

    def _extract_mgmt_no(self, name: str) -> str | None:
        match = MGMT_PATTERN.match(name)
        return match.group(1) if match else None

    def _copy_mgmt_nos(self):
        if not self.classified_mgmt_nos:
            return
        text = "\n".join(self.classified_mgmt_nos)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("복사 완료", f"{len(self.classified_mgmt_nos)}건의 관리번호가 클립보드에 복사되었습니다.\n웹의 '도서접수/배포' 페이지에 붙여넣기 하세요.")


if __name__ == "__main__":
    root = tk.Tk()
    app = ClassifyApp(root)
    root.mainloop()
