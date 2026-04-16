"""설계도서 자료분류 CLI 도구

접수 폴더의 설계도서를 검토위원별 폴더로 분류합니다.

사용법:
    python classify.py --source "D:/2025모니터링/01.접수자료/예비검토/20250501" \
                       --target "D:/2025모니터링/02.배포자료" \
                       --assignment "검토위원배정.xlsx"

엑셀 형식: A열=관리번호, B열=검토위원 이름
소스 폴더 내 하위 폴더/zip 형식: 2025-xxxx_... (관리번호로 시작)
"""

import argparse
import os
import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import load_workbook

MGMT_PATTERN = re.compile(r"^(\d{4}-\d{4})")


def load_assignment(excel_path: str) -> dict[str, str]:
    """배정 엑셀에서 관리번호 → 검토위원 매핑 로드"""
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=2):
        mgmt_no = str(row[0].value).strip() if row[0].value else None
        reviewer = str(row[1].value).strip() if len(row) > 1 and row[1].value else None
        if mgmt_no and reviewer:
            mapping[mgmt_no] = reviewer
    wb.close()
    return mapping


def extract_mgmt_no(name: str) -> str | None:
    """폴더/파일명에서 관리번호 추출"""
    match = MGMT_PATTERN.match(name)
    return match.group(1) if match else None


def classify(source_dir: str, target_dir: str, assignment: dict[str, str]) -> dict:
    """소스 폴더의 파일을 검토위원별로 분류

    Returns:
        {"classified": int, "skipped": int, "details": [...]}
    """
    source = Path(source_dir)
    target = Path(target_dir)

    if not source.exists():
        print(f"소스 폴더가 존재하지 않습니다: {source}")
        return {"classified": 0, "skipped": 0, "details": []}

    target.mkdir(parents=True, exist_ok=True)

    classified = 0
    skipped = 0
    details = []

    # 소스 폴더 내 항목 순회
    for item in sorted(source.iterdir()):
        mgmt_no = extract_mgmt_no(item.name)
        if not mgmt_no:
            skipped += 1
            details.append(f"  스킵: {item.name} (관리번호 인식 불가)")
            continue

        reviewer = assignment.get(mgmt_no)
        if not reviewer:
            skipped += 1
            details.append(f"  스킵: {item.name} (검토위원 미배정)")
            continue

        # 검토위원 폴더 생성
        reviewer_dir = target / reviewer
        reviewer_dir.mkdir(exist_ok=True)

        dest = reviewer_dir / item.name

        if item.is_dir():
            # 폴더 → 복사
            if dest.exists():
                details.append(f"  덮어쓰기: {item.name} → {reviewer}/")
                shutil.rmtree(dest)
            shutil.copytree(item, dest)
            classified += 1
            details.append(f"  분류: {item.name} → {reviewer}/")

        elif item.suffix.lower() == ".zip":
            # ZIP 파일 → 검토위원 폴더에 복사
            shutil.copy2(item, dest)
            classified += 1
            details.append(f"  분류: {item.name} → {reviewer}/")

        else:
            # 기타 파일
            shutil.copy2(item, dest)
            classified += 1
            details.append(f"  분류: {item.name} → {reviewer}/")

    return {"classified": classified, "skipped": skipped, "details": details}


def main():
    parser = argparse.ArgumentParser(
        description="설계도서 자료분류 - 검토위원별 폴더 분류"
    )
    parser.add_argument(
        "--source", required=True,
        help="접수 자료 폴더 경로 (예: D:/2025모니터링/01.접수자료/예비검토/20250501)"
    )
    parser.add_argument(
        "--target", required=True,
        help="배포 대상 폴더 경로 (예: D:/2025모니터링/02.배포자료)"
    )
    parser.add_argument(
        "--assignment", required=True,
        help="검토위원 배정 엑셀 경로 (A열: 관리번호, B열: 검토위원)"
    )

    args = parser.parse_args()

    print("=" * 60)
    print("설계도서 자료분류")
    print("=" * 60)
    print(f"소스 폴더: {args.source}")
    print(f"대상 폴더: {args.target}")
    print(f"배정 엑셀: {args.assignment}")
    print()

    # 배정 로드
    assignment = load_assignment(args.assignment)
    print(f"검토위원 배정: {len(assignment)}건 로드")
    print()

    # 분류 실행
    result = classify(args.source, args.target, assignment)

    print("--- 분류 결과 ---")
    for detail in result["details"]:
        print(detail)
    print()
    print(f"분류 완료: {result['classified']}건")
    print(f"스킵: {result['skipped']}건")

    # 검토위원별 건수 출력
    target = Path(args.target)
    if target.exists():
        print()
        print("--- 검토위원별 건수 ---")
        for reviewer_dir in sorted(target.iterdir()):
            if reviewer_dir.is_dir():
                count = len(list(reviewer_dir.iterdir()))
                print(f"  {reviewer_dir.name}: {count}건")

    # 분류된 관리번호 목록 출력 (접수 처리용)
    print()
    print("--- 접수 관리번호 목록 (웹에 붙여넣기용) ---")
    classified_nos = []
    for item in sorted(Path(args.source).iterdir()):
        mgmt_no = extract_mgmt_no(item.name)
        if mgmt_no and mgmt_no in assignment:
            classified_nos.append(mgmt_no)
    print("\n".join(classified_nos))


if __name__ == "__main__":
    main()
