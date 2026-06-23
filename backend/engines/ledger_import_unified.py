"""통합 관리대장 시트 Import 엔진

'통합 관리대장' 시트 기준 (3차수 통합본).
Row 3: 대분류 헤더
Row 4: 상세 컬럼명
Row 5~: 데이터 (3400+ 행)
"""

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models.building import Building
from models.phase_transition_log import PhaseTransitionLog
from models.review_stage import ReviewStage, PhaseType, ResultType
from services.audit import log_action
from services.phase_transition import transition_phase
from engines.column_mapping import col_letter_to_index

DATA_START_ROW = 5
GROUP_HEADER_ROW = 3
DETAIL_HEADER_ROW = 4
SHEET_NAME = "통합 관리대장"
SUPPLEMENT_SHEET_NAME = "통합 보완대장"
MAX_INLINE_WARNINGS = 200
PHASE_INDEX = {
    None: -1,
    "assigned": 0,
    "doc_received": 1,
    "preliminary": 2,
    "supplement_1_received": 3,
    "supplement_1": 4,
    "supplement_2_received": 5,
    "supplement_2": 6,
    "supplement_3_received": 7,
    "supplement_3": 8,
    "supplement_4_received": 9,
    "supplement_4": 10,
    "supplement_5_received": 11,
    "supplement_5": 12,
    "completed": 99,
}

# 통합 관리대장 열 매핑 (Row 4 기준)
BUILDING_COLUMN_MAP = {
    "A": "mgmt_no",
    "C": "building_type",        # 건축구분
    "E": "sido",                 # 시도명
    "F": "sigungu",              # 시군구명
    "G": "beopjeongdong",        # 법정동명
    "H": "land_type",            # 대지구분
    "I": "main_lot_no",          # 본번
    "J": "sub_lot_no",           # 부번
    "K": "special_lot_no",       # 특수지번
    "L": "building_name",        # 건물명
    "M": "gross_area",           # 연면적
    "N": "main_structure",       # 주구조
    "O": "other_structure",      # 기타구조
    "P": "main_usage",           # 주용도
    "Q": "other_usage",          # 기타용도
    "T": "height",               # 높이
    "U": "floors_above",         # 지상층수
    "V": "floors_below",         # 지하층수
    "AL": "is_special_structure", # 특수구조물 여부
    "AM": "is_high_rise",        # 고층 여부
    "AN": "is_multi_use",        # 다중이용건축물 여부
    "AR": "architect_firm",      # 건축사(소속)
    "AS": "architect_name",      # 건축사(성명)
    "AT": "struct_eng_firm",     # 책임구조기술자(소속)
    "AU": "struct_eng_name",     # 책임구조기술자(성명)
    "AZ": "high_risk_type",      # 고위험유형
    "AP": "remarks",             # 비고
}

# 예비판정 매핑
PRELIMINARY_MAP = {
    "BQ": "reviewer_name",       # 검토자
    "BS": "defect_type_1",       # 부적합유형-1
    "BT": "defect_type_2",       # 부적합유형-2
    "BU": "defect_type_3",       # 부적합유형-3
}

PRELIMINARY_DECISION_COLUMN = "BR"       # 1차검토의견(기술사회)
PRELIMINARY_ADMIN_RESULT_COLUMN = "BV"   # 예비판정 결과(관리원 입력)
PRELIMINARY_OPINION_COLUMN = "BW"        # 예비 검토의견

# 보완자료 검토(1차) 매핑
SUPPLEMENT_1_MAP = {
    "CI": "reviewer_name",       # 검토자
    "CJ": "result",              # 판정 결과
    "CK": "defect_type_1",       # 부적합유형-1
    "CL": "defect_type_2",       # 부적합유형-2
    "CM": "defect_type_3",       # 부적합유형-3
    "CN": "review_opinion",      # 검토의견
}

# 검토위원 열
REVIEWER_COLUMN = "DL"

# 최종 판정 열
FINAL_RESULT_COLUMN = "CW"

SUPPLEMENT_RESULT_COLUMNS = {
    1: "AP",
    2: "AX",
    3: "BF",
    4: "BN",
    5: "BV",
}

SUPPLEMENT_PHASES = {
    1: PhaseType.SUPPLEMENT_1,
    2: PhaseType.SUPPLEMENT_2,
    3: PhaseType.SUPPLEMENT_3,
    4: PhaseType.SUPPLEMENT_4,
    5: PhaseType.SUPPLEMENT_5,
}

RESULT_LABELS = {
    ResultType.PASS: "적합",
    ResultType.SIMPLE_ERROR: "단순오류",
    ResultType.RECALCULATE: "재계산",
}

FINAL_RESULT_MAP = {
    "원적합": "pass",
    "적합": "pass",
    "보완적합": "pass_supplement",
    "부적합": "fail",
    "부적합미회신": "fail_no_response",
    "부적합(미회신)": "fail_no_response",
    "대상제외": "excluded",
}


def _cell_value(row: tuple, col_letter: str):
    idx = col_letter_to_index(col_letter)
    if idx >= len(row):
        return None
    cell = row[idx]
    val = cell.value if hasattr(cell, "value") else cell
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None
    return val


def _clean_text(val) -> str:
    if val is None:
        return ""
    text = str(val).strip()
    text = re.sub(r"_?x000D_", "\n", text, flags=re.IGNORECASE)
    text = text.replace("\r", "\n")
    text = re.sub(r"\n\s*/\s*", "\n", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def _text_cell_value(row: tuple, col_letter: str) -> str | None:
    text = _clean_text(_cell_value(row, col_letter))
    return text or None


def _normalize_excel_text(value) -> str:
    return (
        _clean_text(value)
        .replace("\n", "")
        .replace(" ", "")
        .replace("\t", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )


def _is_blank_marker(value) -> bool:
    text = _normalize_excel_text(value)
    return text in ("", "-", "X", "x")


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _to_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _to_bool(val) -> bool | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("Y", "예", "○", "O", "1", "해당"):
        return True
    if s in ("N", "아니오", "×", "X", "0", "미해당"):
        return False
    return None


def _review_result_key(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, ResultType):
        return val.value
    text = _clean_text(val)
    if _is_blank_marker(text):
        return None
    mapping = {
        "적합": ResultType.PASS.value,
        "원적합": ResultType.PASS.value,
        "단순오류": ResultType.SIMPLE_ERROR.value,
        "경미": ResultType.SIMPLE_ERROR.value,
        "재계산": ResultType.RECALCULATE.value,
        "보완": ResultType.RECALCULATE.value,
        "부적합": ResultType.RECALCULATE.value,
    }
    return mapping.get(text, text)


def _review_result_label(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, ResultType):
        return RESULT_LABELS[val]
    text = _clean_text(val)
    key = _review_result_key(text)
    labels = {
        ResultType.PASS.value: "적합",
        ResultType.SIMPLE_ERROR.value: "단순오류",
        ResultType.RECALCULATE.value: "재계산",
    }
    return labels.get(key or "", text or None)


def _parse_result(val) -> ResultType | None:
    key = _review_result_key(val)
    mapping = {
        ResultType.PASS.value: ResultType.PASS,
        ResultType.SIMPLE_ERROR.value: ResultType.SIMPLE_ERROR,
        ResultType.RECALCULATE.value: ResultType.RECALCULATE,
    }
    return mapping.get(key or "")


def _parse_final_result(val) -> str | None:
    text = _clean_text(val)
    if _is_blank_marker(text):
        return None
    normalized = _normalize_excel_text(text)
    return FINAL_RESULT_MAP.get(normalized, text)


def _find_sheet(wb, preferred_name: str, *required_tokens: str):
    if preferred_name in wb.sheetnames:
        return wb[preferred_name]
    normalized_tokens = [_normalize_excel_text(token) for token in required_tokens]
    for sheet_name in wb.sheetnames:
        normalized_name = _normalize_excel_text(sheet_name)
        if all(token in normalized_name for token in normalized_tokens):
            return wb[sheet_name]
    return None


def _group_by_column(ws) -> dict[int, str]:
    groups: dict[int, str] = {}
    current = ""
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(GROUP_HEADER_ROW, col_idx).value
        if not _is_blank_marker(header):
            current = _normalize_excel_text(header)
        groups[col_idx] = current
    return groups


def _resolve_column(
    ws,
    *,
    fallback: str,
    detail_keywords: tuple[str, ...],
    group_keywords: tuple[str, ...] = (),
) -> str:
    detail_tokens = tuple(_normalize_excel_text(keyword) for keyword in detail_keywords)
    group_tokens = tuple(_normalize_excel_text(keyword) for keyword in group_keywords)
    groups = _group_by_column(ws)

    for col_idx in range(1, ws.max_column + 1):
        detail = _normalize_excel_text(ws.cell(DETAIL_HEADER_ROW, col_idx).value)
        if not detail:
            continue
        if any(token not in detail for token in detail_tokens):
            continue
        group = groups.get(col_idx, "")
        if all(token in group for token in group_tokens):
            return get_column_letter(col_idx)
    return fallback


def _append_warning(result: dict, message: str) -> None:
    result["warning_count"] += 1
    if len(result["warnings"]) < MAX_INLINE_WARNINGS:
        result["warnings"].append(message)
    elif len(result["warnings"]) == MAX_INLINE_WARNINGS:
        result["warnings"].append(
            f"알림이 {MAX_INLINE_WARNINGS}건을 초과하여 이후 항목은 감사 로그에서 확인해 주세요."
        )


def _find_stage(
    stage_map: dict[tuple[int, PhaseType], ReviewStage],
    building: Building,
    phase: PhaseType,
) -> ReviewStage | None:
    if building.id is None:
        return None
    return stage_map.get((building.id, phase))


def _apply_data(target, data: dict, *, skip_none: bool = False) -> None:
    for field_name, value in data.items():
        if skip_none and value is None:
            continue
        setattr(target, field_name, value)


def _transition_import_forward(
    db: Session,
    building: Building,
    *,
    to_phase: str,
    actor_user_id: int | None,
) -> None:
    current_rank = PHASE_INDEX.get(building.current_phase, -1)
    target_rank = PHASE_INDEX.get(to_phase, -1)
    if target_rank <= current_rank:
        return
    transition_phase(
        db,
        building,
        to_phase=to_phase,
        trigger="import",
        actor_user_id=actor_user_id,
        reason="ledger_import_unified",
    )


def _build_preliminary_data(row: tuple, columns: dict[str, str]) -> dict:
    prelim_data = {}
    for col_letter, field_name in PRELIMINARY_MAP.items():
        prelim_data[field_name] = _text_cell_value(row, col_letter)

    preliminary_decision = _text_cell_value(row, columns["preliminary_decision"])
    preliminary_admin_result = _text_cell_value(row, columns["preliminary_admin_result"])
    prelim_data["result"] = _parse_result(preliminary_decision) or _parse_result(preliminary_admin_result)
    prelim_data["review_opinion"] = _text_cell_value(row, columns["preliminary_opinion"])

    remarks = []
    if preliminary_decision and not _is_blank_marker(preliminary_decision):
        remarks.append(f"판정의견: {preliminary_decision}")
    if preliminary_admin_result and not _is_blank_marker(preliminary_admin_result):
        remarks.append(f"관리원 입력 예비판정 결과: {preliminary_admin_result}")
    prelim_data["stage_remarks"] = "\n".join(remarks) if remarks else None
    return prelim_data


def _apply_final_result(
    db: Session,
    result: dict,
    building: Building | None,
    *,
    mgmt_no: str | None = None,
    row_idx: int,
    final_raw: str | None,
    actor_user_id: int | None,
    dry_run: bool = False,
) -> None:
    final_result = _parse_final_result(final_raw)
    if final_result is None:
        return

    before_final = building.final_result if building is not None else None
    before_phase = building.current_phase if building is not None else None
    phase_changed = before_phase != "completed"
    final_changed = before_final != final_result
    if not phase_changed and not final_changed:
        return

    display_mgmt_no = mgmt_no or (building.mgmt_no if building is not None else "-")
    result["final_result_updated"] += 1
    if dry_run:
        _append_warning(
            result,
            f"{display_mgmt_no}: CW 최종완료 반영 예정 "
            f"({before_final or '-'} → {final_result}, 원문: {final_raw})",
        )
        return

    if building is None:
        return

    building.final_result = final_result
    building.current_phase = "completed"

    if phase_changed:
        db.add(PhaseTransitionLog(
            building_id=building.id,
            mgmt_no=building.mgmt_no,
            from_phase=before_phase,
            to_phase="completed",
            trigger="import",
            actor_user_id=actor_user_id,
            reason=f"ledger_import_unified:final_result:{final_result}",
        ))

    log_action(
        db,
        actor_user_id,
        "ledger_final_result_update",
        "building",
        building.id,
        before_data={
            "mgmt_no": building.mgmt_no,
            "current_phase": before_phase,
            "final_result": before_final,
        },
        after_data={
            "mgmt_no": building.mgmt_no,
            "row": row_idx,
            "current_phase": "completed",
            "final_result": final_result,
            "final_result_raw": final_raw,
        },
    )
    _append_warning(
        result,
        f"{display_mgmt_no}: CW 최종완료 반영 "
        f"({before_final or '-'} → {final_result}, 원문: {final_raw})",
    )


def _warn_preliminary_mismatch(
    db: Session,
    result: dict,
    building: Building,
    stage: ReviewStage | None,
    *,
    row_idx: int,
    excel_value: str | None,
    actor_user_id: int | None,
    dry_run: bool = False,
) -> None:
    if not excel_value or _is_blank_marker(excel_value):
        return
    db_value = stage.result if stage else None
    excel_key = _review_result_key(excel_value)
    db_key = _review_result_key(db_value)
    if excel_key == db_key:
        return

    db_label = _review_result_label(db_value) or "없음"
    message = (
        f"{building.mgmt_no}: BR 1차검토의견(기술사회)과 DB 예비판정 불일치 "
        f"(엑셀: {excel_value}, DB: {db_label})"
    )
    _append_warning(result, message)
    if dry_run:
        return
    log_action(
        db,
        actor_user_id,
        "ledger_preliminary_result_mismatch",
        "building",
        building.id,
        after_data={
            "mgmt_no": building.mgmt_no,
            "row": row_idx,
            "sheet": SHEET_NAME,
            "excel_column": "BR",
            "excel_value": excel_value,
            "db_value": db_label,
        },
    )


def _warn_supplement_mismatch(
    db: Session,
    result: dict,
    building: Building,
    stage: ReviewStage | None,
    *,
    row_idx: int,
    phase_no: int,
    column: str,
    excel_value: str | None,
    actor_user_id: int | None,
    dry_run: bool = False,
) -> None:
    if not excel_value or _is_blank_marker(excel_value):
        return
    db_value = stage.result if stage else None
    excel_key = _review_result_key(excel_value)
    db_key = _review_result_key(db_value)
    if excel_key == db_key:
        return

    db_label = _review_result_label(db_value) or "없음"
    message = (
        f"{building.mgmt_no}: 통합 보완대장 {phase_no}차 판정결과 불일치 "
        f"(엑셀 {column}: {excel_value}, DB: {db_label})"
    )
    _append_warning(result, message)
    if dry_run:
        return
    log_action(
        db,
        actor_user_id,
        "ledger_supplement_result_mismatch",
        "building",
        building.id,
        after_data={
            "mgmt_no": building.mgmt_no,
            "row": row_idx,
            "sheet": SUPPLEMENT_SHEET_NAME,
            "phase": f"supplement_{phase_no}",
            "excel_column": column,
            "excel_value": excel_value,
            "db_value": db_label,
        },
    )


def import_ledger_unified(
    file_path: str | Path,
    db: Session,
    actor_user_id: int | None = None,
    *,
    dry_run: bool = False,
) -> dict:
    """통합 관리대장 시트를 DB에 import (일괄 처리 최적화)"""
    wb = load_workbook(str(file_path), data_only=True, read_only=True)

    ws = _find_sheet(wb, SHEET_NAME, "통합", "관리대장")
    if ws is None:
        wb.close()
        return {"imported": 0, "skipped": 0, "errors": [f"시트 '{SHEET_NAME}'을 찾을 수 없습니다"]}

    supplement_ws = _find_sheet(wb, SUPPLEMENT_SHEET_NAME, "통합", "보완대장")
    columns = {
        "mgmt_no": _resolve_column(
            ws,
            fallback="A",
            detail_keywords=("관리번호",),
        ),
        "preliminary_decision": _resolve_column(
            ws,
            fallback=PRELIMINARY_DECISION_COLUMN,
            detail_keywords=("1차검토의견", "기술사회"),
            group_keywords=("예비판정",),
        ),
        "preliminary_admin_result": _resolve_column(
            ws,
            fallback=PRELIMINARY_ADMIN_RESULT_COLUMN,
            detail_keywords=("예비판정", "결과"),
            group_keywords=("예비판정",),
        ),
        "preliminary_opinion": _resolve_column(
            ws,
            fallback=PRELIMINARY_OPINION_COLUMN,
            detail_keywords=("예비", "검토의견"),
            group_keywords=("예비판정",),
        ),
        "final_result": _resolve_column(
            ws,
            fallback=FINAL_RESULT_COLUMN,
            detail_keywords=("최종", "판정결과"),
            group_keywords=("결과보고",),
        ),
    }

    # 1단계: 엑셀 데이터를 메모리에 수집
    rows_parsed = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        mgmt_no = _cell_value(row, columns["mgmt_no"])
        if not mgmt_no:
            continue
        mgmt_no = str(mgmt_no).strip()
        if not (len(mgmt_no) >= 9 and "-" in mgmt_no):
            continue
        rows_parsed.append((row_idx, mgmt_no, row))

    supplement_columns: dict[int, str] = {}
    supplement_rows = []
    if supplement_ws is not None:
        supplement_mgmt_no_col = _resolve_column(
            supplement_ws,
            fallback="A",
            detail_keywords=("관리번호",),
        )
        for phase_no, fallback in SUPPLEMENT_RESULT_COLUMNS.items():
            supplement_columns[phase_no] = _resolve_column(
                supplement_ws,
                fallback=fallback,
                detail_keywords=("판정", "결과"),
                group_keywords=(f"{phase_no}차",),
            )

        for row_idx, row in enumerate(
            supplement_ws.iter_rows(min_row=DATA_START_ROW, values_only=True),
            start=DATA_START_ROW,
        ):
            mgmt_no = _cell_value(row, supplement_mgmt_no_col)
            if not mgmt_no:
                continue
            mgmt_no = str(mgmt_no).strip()
            if not (len(mgmt_no) >= 9 and "-" in mgmt_no):
                continue
            supplement_rows.append((row_idx, mgmt_no, row))
    wb.close()

    if not rows_parsed and not supplement_rows:
        return {"imported": 0, "updated": 0, "skipped": 0, "errors": [], "warnings": [], "warning_count": 0}

    # 2단계: 기존 관리번호 일괄 조회
    all_mgmt_nos = list({r[1] for r in rows_parsed} | {r[1] for r in supplement_rows})
    # 1000개씩 분할 조회 (IN 절 제한 대응)
    existing_map: dict[str, Building] = {}
    for i in range(0, len(all_mgmt_nos), 1000):
        chunk = all_mgmt_nos[i:i+1000]
        existing = db.query(Building).filter(Building.mgmt_no.in_(chunk)).all()
        existing_map.update((building.mgmt_no, building) for building in existing)

    stage_map: dict[tuple[int, PhaseType], ReviewStage] = {}
    existing_ids = [building.id for building in existing_map.values() if building.id is not None]
    for i in range(0, len(existing_ids), 1000):
        chunk = existing_ids[i:i + 1000]
        stages = db.query(ReviewStage).filter(ReviewStage.building_id.in_(chunk)).all()
        stage_map.update(((stage.building_id, stage.phase), stage) for stage in stages)

    result = {
        "imported": 0,
        "updated": 0,
        "skipped": 0,
        "errors": [],
        "warnings": [],
        "warning_count": 0,
        "final_result_updated": 0,
        "sheet": ws.title,
        "supplement_sheet": supplement_ws.title if supplement_ws is not None else None,
        "mode": "validate" if dry_run else "import",
    }

    # 3단계: 일괄 생성 (배치 커밋)
    batch_count = 0
    for row_idx, mgmt_no, row in rows_parsed:
        # 건축물 기본정보
        building_data = {}
        for col_letter, field_name in BUILDING_COLUMN_MAP.items():
            val = _cell_value(row, col_letter)
            if field_name in ("gross_area", "height"):
                val = _to_float(val)
            elif field_name in ("floors_above", "floors_below"):
                val = _to_int(val)
            elif field_name in ("is_special_structure", "is_high_rise", "is_multi_use"):
                val = _to_bool(val)
            building_data[field_name] = val

        reviewer_name = _cell_value(row, REVIEWER_COLUMN)
        if reviewer_name:
            building_data["assigned_reviewer_name"] = str(reviewer_name).strip()

        building = existing_map.get(mgmt_no)
        is_new = building is None
        if is_new:
            if not dry_run:
                building = Building(**building_data)
                db.add(building)
                db.flush()
                existing_map[mgmt_no] = building
        else:
            if not dry_run:
                _apply_data(building, building_data, skip_none=True)

        # 예비검토
        prelim_data = _build_preliminary_data(row, columns)

        if any(v is not None for v in prelim_data.values()):
            stage = _find_stage(stage_map, building, PhaseType.PRELIMINARY) if building else None
            if not is_new and building is not None:
                _warn_preliminary_mismatch(
                    db,
                    result,
                    building,
                    stage,
                    row_idx=row_idx,
                    excel_value=_text_cell_value(row, columns["preliminary_decision"]),
                    actor_user_id=actor_user_id,
                    dry_run=dry_run,
                )
            if not dry_run and building is not None:
                if stage is None:
                    stage = ReviewStage(
                        building_id=building.id,
                        phase=PhaseType.PRELIMINARY,
                        phase_order=0,
                        **prelim_data,
                    )
                    db.add(stage)
                    db.flush()
                    stage_map[(building.id, PhaseType.PRELIMINARY)] = stage
                else:
                    _apply_data(stage, prelim_data)
                _transition_import_forward(
                    db,
                    building,
                    to_phase="preliminary",
                    actor_user_id=actor_user_id,
                )

        # 1차 보완
        supp1_data = {}
        for col_letter, field_name in SUPPLEMENT_1_MAP.items():
            val = _text_cell_value(row, col_letter)
            if field_name == "result":
                val = _parse_result(val)
            supp1_data[field_name] = val

        if any(v is not None for v in supp1_data.values()):
            stage = _find_stage(stage_map, building, PhaseType.SUPPLEMENT_1) if building else None
            if not dry_run and building is not None:
                if stage is None:
                    stage = ReviewStage(
                        building_id=building.id,
                        phase=PhaseType.SUPPLEMENT_1,
                        phase_order=1,
                        **supp1_data,
                    )
                    db.add(stage)
                    db.flush()
                    stage_map[(building.id, PhaseType.SUPPLEMENT_1)] = stage
                else:
                    _apply_data(stage, supp1_data)
                _transition_import_forward(
                    db,
                    building,
                    to_phase="supplement_1",
                    actor_user_id=actor_user_id,
                )

        _apply_final_result(
            db,
            result,
            building,
            mgmt_no=mgmt_no,
            row_idx=row_idx,
            final_raw=_text_cell_value(row, columns["final_result"]),
            actor_user_id=actor_user_id,
            dry_run=dry_run,
        )

        if is_new:
            result["imported"] += 1
        else:
            result["updated"] += 1
        batch_count += 1

        # 500건마다 중간 커밋 (메모리 관리)
        if not dry_run and batch_count % 500 == 0:
            db.commit()

    for row_idx, mgmt_no, row in supplement_rows:
        building = existing_map.get(mgmt_no)
        if building is None:
            _append_warning(result, f"{mgmt_no}: 통합 보완대장 행은 있으나 DB 건축물이 없습니다")
            if not dry_run:
                log_action(
                    db,
                    actor_user_id,
                    "ledger_supplement_building_missing",
                    "building",
                    None,
                    after_data={"mgmt_no": mgmt_no, "row": row_idx, "sheet": SUPPLEMENT_SHEET_NAME},
                )
            continue

        for phase_no, col_letter in supplement_columns.items():
            excel_value = _text_cell_value(row, col_letter)
            if not excel_value or _is_blank_marker(excel_value):
                continue
            phase = SUPPLEMENT_PHASES[phase_no]
            stage = _find_stage(stage_map, building, phase)
            _warn_supplement_mismatch(
                db,
                result,
                building,
                stage,
                row_idx=row_idx,
                phase_no=phase_no,
                column=col_letter,
                excel_value=excel_value,
                actor_user_id=actor_user_id,
                dry_run=dry_run,
            )

    if not dry_run:
        db.commit()
    return result
