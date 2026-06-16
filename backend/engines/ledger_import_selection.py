"""대상선정 결과 엑셀 → DB Import 엔진"""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.building import Building

HEADER_ROW = 1
DATA_START_ROW = 2

FIELD_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "mgmt_no": ("관리번호", "모니터링관리번호"),
    "building_type": ("건축구분",),
    "sido": ("시도명",),
    "sigungu": ("시군구명",),
    "beopjeongdong": ("법정동명",),
    "land_type": ("대지구분",),
    "main_lot_no": ("본번",),
    "sub_lot_no": ("부번",),
    "special_lot_no": ("특수지번",),
    "building_name": ("건물명",),
    "gross_area": ("연면적",),
    "main_structure": ("주구조",),
    "other_structure": ("기타구조",),
    "main_usage": ("주용도",),
    "other_usage": ("기타용도",),
    "height": ("높이",),
    "floors_above": ("지상층수",),
    "floors_below": ("지하층수",),
    "architect_name": ("설계자", "건축사성명", "건축사"),
    "architect_firm": ("설계사무소", "건축사소속"),
    "remarks": ("공사건물명", "공사명", "비고"),
    "is_special_structure": ("특수구조물여부", "특수구조물"),
    "is_high_rise": ("고층", "고층여부"),
    "is_multi_use": ("다중이용", "다중이용건축물여부"),
    "is_quasi_multi_use": ("준다중이용", "준다중이용시설", "준다중이용건축물여부"),
    "high_risk_type": ("고위험", "고위험유형"),
}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .replace("\n", "")
        .replace(" ", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )


def _cell_value(row: tuple, index: int | None):
    if index is None or index >= len(row):
        return None
    val = row[index].value
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None
    return val


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _to_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _to_bool(val) -> bool | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("Y", "예", "○", "O", "1", "True", "true", "해당"):
        return True
    if s in ("N", "아니오", "×", "X", "0", "False", "false", "미해당"):
        return False
    return None


def _to_high_risk_type(val) -> str | None:
    is_high_risk = _to_bool(val)
    if is_high_risk is True:
        return "고위험"
    if is_high_risk is False or val is None:
        return None
    return str(val).strip()


def _build_header_index(header_row: tuple) -> dict[str, int]:
    normalized_to_index = {
        _normalize_header(cell.value): idx
        for idx, cell in enumerate(header_row)
        if _normalize_header(cell.value)
    }

    header_index: dict[str, int] = {}
    for field_name, aliases in FIELD_HEADER_ALIASES.items():
        for alias in aliases:
            index = normalized_to_index.get(_normalize_header(alias))
            if index is not None:
                header_index[field_name] = index
                break
    return header_index


def _is_mgmt_no(value) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return len(text) >= 9 and "-" in text


def import_ledger_selection(file_path: str | Path, db: Session) -> dict:
    """3차수 대상선정 결과 형식의 엑셀 파일을 DB에 import."""
    wb = load_workbook(str(file_path), data_only=True, read_only=True)

    matched_sheet = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW), ())
        header_index = _build_header_index(header)
        if "mgmt_no" in header_index:
            matched_sheet = sheet_name
            break

    if matched_sheet is None:
        wb.close()
        return {"imported": 0, "skipped": 0, "errors": ["관리번호 헤더를 찾을 수 없습니다"]}

    ws = wb[matched_sheet]
    header = next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW), ())
    header_index = _build_header_index(header)

    rows_parsed = []
    for row in ws.iter_rows(min_row=DATA_START_ROW):
        mgmt_no = _cell_value(row, header_index.get("mgmt_no"))
        if not _is_mgmt_no(mgmt_no):
            continue
        rows_parsed.append((str(mgmt_no).strip(), row))
    wb.close()

    if not rows_parsed:
        return {"imported": 0, "skipped": 0, "errors": [], "sheet": matched_sheet}

    all_mgmt_nos = [mgmt_no for mgmt_no, _ in rows_parsed]
    existing_set: set[str] = set()
    for i in range(0, len(all_mgmt_nos), 1000):
        chunk = all_mgmt_nos[i:i + 1000]
        existing = db.query(Building.mgmt_no).filter(Building.mgmt_no.in_(chunk)).all()
        existing_set.update(r[0] for r in existing)

    result = {"imported": 0, "skipped": 0, "errors": [], "sheet": matched_sheet}
    for mgmt_no, row in rows_parsed:
        if mgmt_no in existing_set:
            result["skipped"] += 1
            continue

        building_data = {"mgmt_no": mgmt_no}
        for field_name in FIELD_HEADER_ALIASES:
            if field_name == "mgmt_no":
                continue
            val = _cell_value(row, header_index.get(field_name))
            if field_name in ("gross_area", "height"):
                val = _to_float(val)
            elif field_name in ("floors_above", "floors_below"):
                val = _to_int(val)
            elif field_name in (
                "is_special_structure",
                "is_high_rise",
                "is_multi_use",
                "is_quasi_multi_use",
            ):
                val = _to_bool(val)
            elif field_name == "high_risk_type":
                val = _to_high_risk_type(val)
            building_data[field_name] = val

        db.add(Building(**building_data))
        db.flush()
        result["imported"] += 1

    db.commit()
    return result
