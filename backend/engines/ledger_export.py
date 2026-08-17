"""DB → 통합관리대장 엑셀 Export 엔진"""

from pathlib import Path
from typing import Iterator

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from sqlalchemy.orm import Session

from models.building import Building
from models.review_stage import ReviewStage, PhaseType
from services.related_tech import has_related_tech_cooperation, is_related_tech_target
from engines.column_mapping import (
    BUILDING_COLUMN_MAP,
    PRELIMINARY_STAGE_MAP,
    SUPPLEMENT_SUBMIT_START_COLS,
    SUPPLEMENT_SUBMIT_OFFSETS,
    SUPPLEMENT_REVIEW_START_COLS,
    SUPPLEMENT_REVIEW_OFFSETS,
    FINAL_RESULT_COLUMN,
    col_letter_to_index,
    index_to_col_letter,
)

# 스타일 상수
HEADER_FONT = Font(bold=True, size=10)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# Row 1 대분류 헤더 정의
ROW1_HEADERS = {
    "C": "대상 건축물 개요(허가대장 DB)",
    "AC": "관계기술자 협력",
    "AE": "예비도서 접수",
    "AF": "예비판정",
    "AN": "보완서류 제출(1차)",
    "AS": "보완자료 검토(1차)",
    "AZ": "보완서류 제출(2차)",
    "BE": "보완자료 검토(2차)",
    "BL": "보완서류 제출(3차)",
    "BQ": "보완자료 검토(3차)",
    "BX": "보완서류 제출(4차)",
    "CC": "보완자료 검토(4차)",
    "CJ": "결과보고",
}

# Row 2 상세 컬럼명 (역매핑)
FIELD_TO_LABEL = {
    "mgmt_no": "모니터링\n관리번호",
    "building_type": "건축구분",
    "sido": "시도명",
    "sigungu": "시군구명",
    "beopjeongdong": "법정동명",
    "land_type": "대지구분",
    "main_lot_no": "본번",
    "sub_lot_no": "부번",
    "special_lot_no": "특수지번",
    "building_name": "건물명",
    "main_structure": "주구조",
    "other_structure": "기타구조",
    "main_usage": "주용도",
    "other_usage": "기타용도",
    "gross_area": "연면적",
    "height": "높이",
    "floors_above": "지상층수",
    "floors_below": "지하층수",
    "is_special_structure": "특수구조물 여부",
    "is_high_rise": "고층 여부",
    "is_multi_use": "다중이용건축물 여부",
    "remarks": "비고",
    "architect_firm": "건축사(소속)",
    "architect_name": "건축사(성명)",
    "struct_eng_firm": "책임구조기술자(소속)",
    "struct_eng_name": "책임구조기술자(성명)",
    "high_risk_type": "고위험유형",
    "related_tech_coop_target": "협력대상",
    "related_tech_coop": "협력여부",
    "doc_received_at": "도서접수일",
    "report_submitted_at": "검토서 제출일",
    "reviewer_name": "검토자",
    "review_opinion": "검토의견",
    "defect_type_1": "부적합유형-1",
    "defect_type_2": "부적합유형-2",
    "defect_type_3": "부적합유형-3",
    "result": "판정 결과",
    "stage_remarks": "비고",
    "objection_filed": "이의신청 제출",
    "objection_content": "이의신청\n검토내용",
    "objection_reason": "이의신청 사유",
}

RELATED_TECH_COLUMNS = {
    "AC": "related_tech_coop_target",
    "AD": "related_tech_coop",
}

RESULT_LABELS = {
    "pass": "적합",
    "supplement": "보완",
    "fail": "부적합",
    "minor": "경미",
}

# 최종판정(final_result) 코드 → CW열 한글 표기 역매핑.
# 실물 관리대장 CW열은 코드 값이 아닌 한글 라벨(원적합/보완적합/부적합(단순오류)…)을
# 사용하므로 export 시에도 한글 라벨로 기록한다. 매핑에 없는 값은 원본 그대로 둔다.
FINAL_RESULT_EXPORT_LABELS = {
    "pass": "원적합",
    "pass_supplement": "보완적합",
    "fail_simple_error": "부적합(단순오류)",
    "fail_recalculate": "부적합(재계산)",
    "fail_no_response": "부적합(미회신)",
    "excluded": "대상제외",
    "fail": "부적합",  # 레거시
}


def _format_value(val, field_name: str):
    """DB 값을 엑셀 출력용으로 변환"""
    if val is None:
        return ""
    if field_name == "result" and hasattr(val, "value"):
        return RESULT_LABELS.get(val.value, str(val.value))
    if isinstance(val, bool):
        return "Y" if val else "N"
    return val


def _related_tech_value(building: Building, field_name: str) -> bool:
    if field_name == "related_tech_coop_target":
        return is_related_tech_target(building)
    if field_name == "related_tech_coop":
        return has_related_tech_cooperation(building)
    return False


# 건물을 한 번에 몇 건씩 읽어 시트에 흘려보낼지. 전체 로드 시 메모리 급증(OOM) 방지.
EXPORT_CHUNK_SIZE = 300


def _sheet_width() -> int:
    """양식이 사용하는 마지막 열 번호(1-based)."""
    indices = [col_letter_to_index(FINAL_RESULT_COLUMN)]
    for mapping in (ROW1_HEADERS, BUILDING_COLUMN_MAP, RELATED_TECH_COLUMNS, PRELIMINARY_STAGE_MAP):
        indices.extend(col_letter_to_index(col_letter) for col_letter in mapping)
    for supp_no in range(1, 5):
        indices.append(
            col_letter_to_index(SUPPLEMENT_SUBMIT_START_COLS[supp_no])
            + max(SUPPLEMENT_SUBMIT_OFFSETS)
        )
        indices.append(
            col_letter_to_index(SUPPLEMENT_REVIEW_START_COLS[supp_no])
            + max(SUPPLEMENT_REVIEW_OFFSETS)
        )
    return max(indices) + 1


def _header_cell(ws, label: str, *, filled: bool = False):
    cell = WriteOnlyCell(ws, value=label)
    cell.font = HEADER_FONT
    if filled:
        cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    return cell


def _row1_cells(ws, width: int) -> list:
    """Row 1: 대분류 헤더"""
    row: list = [None] * width
    for col_letter, label in ROW1_HEADERS.items():
        row[col_letter_to_index(col_letter)] = _header_cell(ws, label, filled=True)
    return row


def _row2_cells(ws, width: int) -> list:
    """Row 2: 상세 컬럼 헤더"""
    row: list = [None] * width
    # A/B열은 스타일 없는 기본 라벨. A열은 아래 BUILDING_COLUMN_MAP에서 덮어쓴다.
    row[0] = "모니터링\n관리번호"
    row[1] = "검토\n위원"

    for mapping in (BUILDING_COLUMN_MAP, RELATED_TECH_COLUMNS, PRELIMINARY_STAGE_MAP):
        for col_letter, field_name in mapping.items():
            label = FIELD_TO_LABEL.get(field_name, field_name)
            row[col_letter_to_index(col_letter)] = _header_cell(ws, label)

    # 보완 단계 헤더 (1차~4차)
    for supp_no in range(1, 5):
        submit_start = col_letter_to_index(SUPPLEMENT_SUBMIT_START_COLS[supp_no])
        for offset, field_name in SUPPLEMENT_SUBMIT_OFFSETS.items():
            label = FIELD_TO_LABEL.get(field_name, field_name)
            row[submit_start + offset] = _header_cell(ws, label)

        review_start = col_letter_to_index(SUPPLEMENT_REVIEW_START_COLS[supp_no])
        for offset, field_name in SUPPLEMENT_REVIEW_OFFSETS.items():
            label = FIELD_TO_LABEL.get(field_name, field_name)
            row[review_start + offset] = _header_cell(ws, label)

    # 최종 판정 헤더는 글꼴만 적용
    final_cell = WriteOnlyCell(ws, value="최종\n판정결과")
    final_cell.font = HEADER_FONT
    row[col_letter_to_index(FINAL_RESULT_COLUMN)] = final_cell
    return row


def _building_row(building: Building, stages: list, width: int) -> list:
    """건물 1건 + 검토 단계들을 데이터 행 하나로 만든다."""
    row: list = [None] * width

    for col_letter, field_name in BUILDING_COLUMN_MAP.items():
        val = getattr(building, field_name, None)
        row[col_letter_to_index(col_letter)] = _format_value(val, field_name)

    for col_letter, field_name in RELATED_TECH_COLUMNS.items():
        val = _related_tech_value(building, field_name)
        row[col_letter_to_index(col_letter)] = _format_value(val, field_name)

    # 검토위원명 (B열)
    if building.assigned_reviewer_name:
        row[1] = building.assigned_reviewer_name
    elif building.reviewer and building.reviewer.user:
        row[1] = building.reviewer.user.name

    # 최종 판정 (코드 → 한글 라벨)
    if building.final_result:
        row[col_letter_to_index(FINAL_RESULT_COLUMN)] = FINAL_RESULT_EXPORT_LABELS.get(
            building.final_result, building.final_result
        )

    for stage in stages:
        if stage.phase == PhaseType.PRELIMINARY:
            # 예비검토
            for col_letter, field_name in PRELIMINARY_STAGE_MAP.items():
                val = getattr(stage, field_name, None)
                row[col_letter_to_index(col_letter)] = _format_value(val, field_name)
            continue

        # 보완 단계
        supp_no = stage.phase_order
        if supp_no < 1 or supp_no > 4:
            continue

        submit_start = col_letter_to_index(SUPPLEMENT_SUBMIT_START_COLS[supp_no])
        for offset, field_name in SUPPLEMENT_SUBMIT_OFFSETS.items():
            val = getattr(stage, field_name, None)
            row[submit_start + offset] = _format_value(val, field_name)

        review_start = col_letter_to_index(SUPPLEMENT_REVIEW_START_COLS[supp_no])
        for offset, field_name in SUPPLEMENT_REVIEW_OFFSETS.items():
            val = getattr(stage, field_name, None)
            row[review_start + offset] = _format_value(val, field_name)

    return row


def _iter_buildings_with_stages(
    db: Session,
    chunk_size: int,
) -> Iterator[tuple[Building, list]]:
    """건물을 청크 단위로 읽어 (건물, 단계 목록)을 차례로 넘긴다.

    전체 건물·단계를 한 번에 세션에 올리면 메모리가 급증하므로, 청크를 다 쓰면
    세션에서 분리(expunge)해 곧바로 회수되게 한다.
    """
    offset = 0
    while True:
        buildings = (
            db.query(Building)
            .order_by(Building.mgmt_no)
            .offset(offset)
            .limit(chunk_size)
            .all()
        )
        if not buildings:
            return

        stages = (
            db.query(ReviewStage)
            .filter(ReviewStage.building_id.in_([b.id for b in buildings]))
            .order_by(ReviewStage.phase_order)
            .all()
        )
        stages_by_building: dict[int, list] = {}
        for stage in stages:
            stages_by_building.setdefault(stage.building_id, []).append(stage)

        for building in buildings:
            yield building, stages_by_building.get(building.id, [])

        for stage in stages:
            db.expunge(stage)
        for building in buildings:
            db.expunge(building)

        if len(buildings) < chunk_size:
            return
        offset += chunk_size


def export_ledger(db: Session, output_path: str | Path) -> Path:
    """DB 데이터를 통합관리대장 형식의 엑셀로 `output_path`에 기록한다.

    write_only 모드라 append 한 행은 곧바로 파일로 흘러가고, 시트 전체를 메모리에
    들고 있지 않는다.

    Returns:
        Path: 기록된 엑셀 파일 경로
    """
    width = _sheet_width()
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="통합 관리대장")

    # 열 너비 조정 (주요 열)
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["K"].width = 25

    ws.append(_row1_cells(ws, width))
    ws.append(_row2_cells(ws, width))

    for building, stages in _iter_buildings_with_stages(db, EXPORT_CHUNK_SIZE):
        ws.append(_building_row(building, stages, width))

    output = Path(output_path)
    try:
        wb.save(str(output))
    finally:
        wb.close()
    return output
