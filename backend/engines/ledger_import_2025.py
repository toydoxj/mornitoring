"""2025년 관리대장 엑셀 → DB Import 엔진

'2025년 관리대장_1차수_250604(1443건)' 시트 기준.
Row 3: 대분류 헤더
Row 4: 상세 컬럼명
Row 5~: 데이터
"""

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.building import Building
from models.review_stage import ReviewStage, PhaseType, ResultType
from engines.column_mapping import col_letter_to_index

DATA_START_ROW = 5
SHEET_NAME = "2025년 관리대장_1차수_250604(1443건)"

# 2025년 시트 열 매핑 (Row 4 기준)
BUILDING_COLUMN_MAP_2025 = {
    "A": "mgmt_no",              # 모니터링 관리번호
    "D": "building_type",        # 건축구분
    "F": "sido",                 # 시도명
    "G": "sigungu",              # 시군구명
    "H": "beopjeongdong",        # 법정동명
    "I": "land_type",            # 대지구분
    "J": "main_lot_no",          # 본번
    "K": "sub_lot_no",           # 부번
    "M": "special_lot_no",       # 특수지번
    "N": "building_name",        # 건물명
    "P": "gross_area",           # 연면적
    "Q": "main_structure",       # 주구조
    "R": "other_structure",      # 기타구조
    "S": "main_usage",           # 주용도
    "T": "other_usage",          # 기타용도
    "W": "height",               # 높이
    "X": "floors_above",         # 지상층수
    "Y": "floors_below",         # 지하층수
    "AO": "is_special_structure", # 특수구조물 여부
    "AP": "is_high_rise",        # 고층 여부
    "AQ": "is_multi_use",        # 다중이용건축물 여부
    "AT": "architect_firm",      # 건축사(소속)
    "AU": "architect_name",      # 건축사(성명)
    "AV": "struct_eng_firm",     # 책임구조기술자(소속)
    "AW": "struct_eng_name",     # 책임구조기술자(성명)
    "BB": "high_risk_type",      # 고위험유형
    "BE": "remarks",             # 비고
}

# 예비판정 매핑
PRELIMINARY_MAP_2025 = {
    "BS": "reviewer_name",       # 검토자
    "BT": "result",              # 판정결과(예비)
    "BV": "defect_type_1",       # 부적합유형-1
    "BW": "defect_type_2",       # 부적합유형-2
    "BX": "defect_type_3",       # 부적합유형-3
    "BY": "review_opinion",      # 예비 검토의견
}

# 보완자료 검토(1차) 매핑
SUPPLEMENT_1_MAP_2025 = {
    "CK": "reviewer_name",       # 검토자
    "CL": "result",              # 판정 결과
    "CM": "defect_type_1",       # 부적합유형-1
    "CN": "defect_type_2",       # 부적합유형-2
    "CO": "defect_type_3",       # 부적합유형-3
    "CP": "review_opinion",      # 검토의견
}


def _cell_value(row: tuple, col_letter: str):
    idx = col_letter_to_index(col_letter)
    if idx >= len(row):
        return None
    val = row[idx].value
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None
    return val


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _to_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _to_bool(val) -> bool | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("Y", "예", "○", "O", "1", "해당"):
        return True
    if s in ("N", "아니오", "×", "X", "0", "미해당"):
        return False
    return None


def _parse_result(val) -> ResultType | None:
    if val is None:
        return None
    s = str(val).strip()
    mapping = {
        "적합": ResultType.PASS,
        "보완": ResultType.SUPPLEMENT,
        "부적합": ResultType.FAIL,
        "경미": ResultType.MINOR,
    }
    return mapping.get(s)


def import_ledger_2025(file_path: str | Path, db: Session, sheet_name: str | None = None) -> dict:
    """2025년 관리대장 엑셀 파일을 DB에 import

    Returns:
        {"imported": int, "skipped": int, "errors": list[str]}
    """
    wb = load_workbook(str(file_path), data_only=True, read_only=True)

    target_sheet = sheet_name or SHEET_NAME
    # 시트명이 정확히 일치하지 않을 수 있으므로 부분 매칭
    matched_sheet = None
    for sn in wb.sheetnames:
        if "관리대장" in sn and ("1차수" in sn or "1443" in sn or "2025" in sn):
            matched_sheet = sn
            break
    if not matched_sheet:
        # 첫 번째 시트 사용
        matched_sheet = wb.sheetnames[0]

    ws = wb[matched_sheet]
    result = {"imported": 0, "skipped": 0, "errors": [], "sheet": matched_sheet}

    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW), start=DATA_START_ROW):
        mgmt_no = _cell_value(row, "A")
        if not mgmt_no:
            continue

        mgmt_no = str(mgmt_no).strip()

        # 관리번호 형식 확인 (YYYY-NNNN)
        if not (len(mgmt_no) >= 9 and "-" in mgmt_no):
            continue

        # 중복 체크
        existing = db.query(Building).filter(Building.mgmt_no == mgmt_no).first()
        if existing:
            result["skipped"] += 1
            continue

        # 건축물 기본정보 파싱
        building_data = {}
        for col_letter, field_name in BUILDING_COLUMN_MAP_2025.items():
            val = _cell_value(row, col_letter)
            if field_name in ("gross_area", "height"):
                val = _to_float(val)
            elif field_name in ("floors_above", "floors_below"):
                val = _to_int(val)
            elif field_name in ("is_special_structure", "is_high_rise", "is_multi_use"):
                val = _to_bool(val)
            building_data[field_name] = val

        # 최종 판정
        final = _cell_value(row, "CY")
        if final:
            building_data["final_result"] = str(final)

        building = Building(**building_data)
        db.add(building)
        db.flush()

        # 예비검토 단계
        prelim_data = {}
        for col_letter, field_name in PRELIMINARY_MAP_2025.items():
            val = _cell_value(row, col_letter)
            if field_name == "result":
                val = _parse_result(val)
            prelim_data[field_name] = val

        if any(v is not None for v in prelim_data.values()):
            stage = ReviewStage(
                building_id=building.id,
                phase=PhaseType.PRELIMINARY,
                phase_order=0,
                **prelim_data,
            )
            db.add(stage)
            building.current_phase = "preliminary"

        # 1차 보완 검토
        supp1_data = {}
        for col_letter, field_name in SUPPLEMENT_1_MAP_2025.items():
            val = _cell_value(row, col_letter)
            if field_name == "result":
                val = _parse_result(val)
            supp1_data[field_name] = val

        if any(v is not None for v in supp1_data.values()):
            stage = ReviewStage(
                building_id=building.id,
                phase=PhaseType.SUPPLEMENT_1,
                phase_order=1,
                **supp1_data,
            )
            db.add(stage)
            building.current_phase = "supplement_1"

        result["imported"] += 1

    db.commit()
    wb.close()
    return result
