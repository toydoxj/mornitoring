"""통합 보완대장 단계와 DB 현재 단계 비교 엔진."""

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from sqlalchemy.orm import Session, selectinload

from models.building import Building
from models.phase_transition_log import PhaseTransitionLog
from models.reviewer import Reviewer
from models.user import User
from services.audit import log_action
from services.scope import building_visibility_filter

SUPPLEMENT_SHEET_NAME = "통합 보완대장"
MANAGEMENT_SHEET_NAME = "통합 관리대장"
DATA_START_ROW = 5
FINAL_RESULT_COLUMN = "CW"

PHASE_INDEX = {
    "assigned": 0,
    "doc_received": 1,
    "preliminary": 2,
    "supplement_1_received": 3,
    "supplement_1": 4,
    "supplement_2_received": 5,
    "supplement_2": 6,
    "supplement_3_received": 7,
    "supplement_3": 8,
    "supplement_4_received": 9,
    "supplement_4": 10,
    "supplement_5_received": 11,
    "supplement_5": 12,
    "completed": 99,
}

SUPPLEMENT_PHASE_COLUMNS = (
    {
        "round": 1,
        "doc_column": "S",
        "report_column": "AV",
        "doc_phase": "supplement_1_received",
        "report_phase": "supplement_1",
    },
    {
        "round": 2,
        "doc_column": "U",
        "report_column": "BD",
        "doc_phase": "supplement_2_received",
        "report_phase": "supplement_2",
    },
    {
        "round": 3,
        "doc_column": "W",
        "report_column": "BL",
        "doc_phase": "supplement_3_received",
        "report_phase": "supplement_3",
    },
    {
        "round": 4,
        "doc_column": "Y",
        "report_column": "BT",
        "doc_phase": "supplement_4_received",
        "report_phase": "supplement_4",
    },
    {
        "round": 5,
        "doc_column": "AA",
        "report_column": "CB",
        "doc_phase": "supplement_5_received",
        "report_phase": "supplement_5",
    },
)

FINAL_RESULT_MAP = {
    "원적합": "pass",
    "적합": "pass",
    "보완적합": "pass_supplement",
    "부적합": "fail",
    "부적합미회신": "fail_no_response",
    "대상제외": "excluded",
}


def _normalize_text(value) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .replace("\n", "")
        .replace(" ", "")
        .replace("\t", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )


def _is_blank_marker(value) -> bool:
    text = _normalize_text(value)
    return text in ("", "-", "X", "x")


def _cell_value(row: tuple, column: str):
    index = column_index_from_string(column) - 1
    if index >= len(row):
        return None
    return row[index]


def _has_value(value) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return not _is_blank_marker(value)
    return True


def _format_cell_value(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _text_cell_value(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if _is_blank_marker(text):
        return None
    return text


def _parse_final_result(value) -> str | None:
    text = _text_cell_value(value)
    if text is None:
        return None
    normalized = _normalize_text(text)
    return FINAL_RESULT_MAP.get(normalized, text)


def _find_supplement_sheet(workbook):
    if SUPPLEMENT_SHEET_NAME in workbook.sheetnames:
        return workbook[SUPPLEMENT_SHEET_NAME]
    for sheet_name in workbook.sheetnames:
        normalized = _normalize_text(sheet_name)
        if "통합" in normalized and "보완대장" in normalized:
            return workbook[sheet_name]
    return None


def _find_management_sheet(workbook):
    if MANAGEMENT_SHEET_NAME in workbook.sheetnames:
        return workbook[MANAGEMENT_SHEET_NAME]
    for sheet_name in workbook.sheetnames:
        normalized = _normalize_text(sheet_name)
        if "통합" in normalized and "관리대장" in normalized and "보완" not in normalized:
            return workbook[sheet_name]
    return None


def _resolve_excel_phase(row: tuple) -> dict[str, object]:
    """지정 열의 제출 여부를 보고 엑셀상 최신 보완 단계를 산출한다."""
    rounds: list[dict[str, object]] = []
    latest: dict[str, object] | None = None

    for definition in SUPPLEMENT_PHASE_COLUMNS:
        doc_value = _cell_value(row, str(definition["doc_column"]))
        report_value = _cell_value(row, str(definition["report_column"]))
        doc_submitted = _has_value(doc_value)
        report_submitted = _has_value(report_value)
        round_no = int(definition["round"])

        rounds.append(
            {
                "round": round_no,
                "doc_column": definition["doc_column"],
                "report_column": definition["report_column"],
                "doc_submitted": doc_submitted,
                "report_submitted": report_submitted,
                "doc_value": _format_cell_value(doc_value),
                "report_value": _format_cell_value(report_value),
            }
        )

        if doc_submitted:
            latest = {
                "phase": definition["doc_phase"],
                "round": round_no,
                "column": definition["doc_column"],
                "value": _format_cell_value(doc_value),
                "label": f"{round_no}차 보완도서 제출",
            }
        if report_submitted:
            latest = {
                "phase": definition["report_phase"],
                "round": round_no,
                "column": definition["report_column"],
                "value": _format_cell_value(report_value),
                "label": f"{round_no}차 보완검토서 제출",
            }

    return {
        "phase": latest["phase"] if latest else None,
        "evidence_round": latest["round"] if latest else None,
        "evidence_column": latest["column"] if latest else None,
        "evidence_value": latest["value"] if latest else None,
        "evidence_label": latest["label"] if latest else None,
        "rounds": rounds,
    }


def _phase_gap(db_phase: str | None, excel_phase: str | None) -> int | None:
    if db_phase not in PHASE_INDEX or excel_phase not in PHASE_INDEX:
        return None
    return PHASE_INDEX[db_phase] - PHASE_INDEX[excel_phase]


def _phase_direction(gap: int | None) -> str:
    if gap is None:
        return "unknown"
    if gap == 0:
        return "same"
    if gap > 0:
        return "db_ahead"
    return "excel_ahead"


def _status_for(building: Building | None, excel_phase: str | None) -> str:
    if building is None:
        return "missing_db"
    if excel_phase is None:
        return "excel_phase_missing"
    if building.current_phase == excel_phase:
        return "matched"
    return "mismatch"


def _final_result_status(
    building: Building | None,
    excel_final_result: str | None,
    *,
    checked: bool,
) -> str:
    if not checked:
        return "not_checked"
    if building is None:
        return "missing_db"
    if excel_final_result is None:
        return "excel_final_result_missing"
    if building.final_result == excel_final_result:
        return "matched"
    return "mismatch"


def _reviewer_name(building: Building) -> str | None:
    if building.reviewer and building.reviewer.user:
        return building.reviewer.user.name
    return building.assigned_reviewer_name


def _fetch_buildings(
    db: Session,
    mgmt_nos: list[str],
    current_user: User,
) -> dict[str, Building]:
    visibility = building_visibility_filter(current_user)
    building_map: dict[str, Building] = {}
    for start in range(0, len(mgmt_nos), 1000):
        chunk = mgmt_nos[start:start + 1000]
        query = (
            db.query(Building)
            .options(selectinload(Building.reviewer).selectinload(Reviewer.user))
            .filter(Building.mgmt_no.in_(chunk))
        )
        if visibility is not None:
            query = query.filter(visibility)
        for building in query.all():
            building_map[building.mgmt_no] = building
    return building_map


def _read_final_result_rows(workbook) -> tuple[str | None, dict[str, dict[str, object]]]:
    worksheet = _find_management_sheet(workbook)
    if worksheet is None:
        return None, {}

    rows: dict[str, dict[str, object]] = {}
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=DATA_START_ROW, values_only=True),
        start=DATA_START_ROW,
    ):
        mgmt_no_value = _cell_value(row, "A")
        if not _has_value(mgmt_no_value):
            continue
        mgmt_no = str(mgmt_no_value).strip()
        if not mgmt_no:
            continue

        final_raw = _text_cell_value(_cell_value(row, FINAL_RESULT_COLUMN))
        rows[mgmt_no] = {
            "management_row_number": row_number,
            "excel_final_result": _parse_final_result(final_raw),
            "excel_final_result_raw": final_raw,
            "final_result_column": FINAL_RESULT_COLUMN,
        }
    return worksheet.title, rows


def _apply_final_result_update(
    db: Session,
    building: Building,
    *,
    row_number: int,
    final_result: str,
    final_result_raw: str | None,
    actor_user_id: int | None,
) -> None:
    before_final = building.final_result
    before_phase = building.current_phase
    final_changed = before_final != final_result
    phase_changed = before_phase != "completed"
    if not final_changed and not phase_changed:
        return

    building.final_result = final_result
    building.current_phase = "completed"

    if phase_changed:
        db.add(PhaseTransitionLog(
            building_id=building.id,
            mgmt_no=building.mgmt_no,
            from_phase=before_phase,
            to_phase="completed",
            trigger="import",
            actor_user_id=actor_user_id,
            reason=f"ledger_phase_compare:final_result:{final_result}",
        ))

    log_action(
        db,
        actor_user_id,
        "ledger_final_result_update",
        "building",
        building.id,
        before_data={
            "mgmt_no": building.mgmt_no,
            "current_phase": before_phase,
            "final_result": before_final,
        },
        after_data={
            "mgmt_no": building.mgmt_no,
            "row": row_number,
            "sheet": MANAGEMENT_SHEET_NAME,
            "excel_column": FINAL_RESULT_COLUMN,
            "current_phase": "completed",
            "final_result": final_result,
            "final_result_raw": final_result_raw,
        },
    )


def compare_supplement_phase_with_db(
    file_path: str | Path,
    db: Session,
    *,
    current_user: User,
) -> dict[str, object]:
    """통합 보완대장 제출 열 기준 최신 단계와 DB 현재 단계를 비교한다."""
    workbook = load_workbook(str(file_path), data_only=True, read_only=True)
    try:
        worksheet = _find_supplement_sheet(workbook)
        if worksheet is None:
            raise ValueError(f"시트 '{SUPPLEMENT_SHEET_NAME}'을 찾을 수 없습니다")

        excel_rows: list[dict[str, object]] = []
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=DATA_START_ROW, values_only=True),
            start=DATA_START_ROW,
        ):
            mgmt_no_value = _cell_value(row, "A")
            if not _has_value(mgmt_no_value):
                continue
            mgmt_no = str(mgmt_no_value).strip()
            if not mgmt_no:
                continue
            excel_phase = _resolve_excel_phase(row)
            excel_rows.append(
                {
                    "row_number": row_number,
                    "mgmt_no": mgmt_no,
                    **excel_phase,
                }
            )
        management_sheet, final_result_rows = _read_final_result_rows(workbook)
    finally:
        workbook.close()

    phase_row_map = {str(row["mgmt_no"]): row for row in excel_rows}
    def row_sort_key(mgmt_no: str) -> tuple[int, str]:
        phase_row = phase_row_map.get(mgmt_no)
        final_row = final_result_rows.get(mgmt_no)
        row_number = (
            phase_row["row_number"]
            if phase_row
            else final_row["management_row_number"] if final_row else 999999
        )
        return int(row_number), mgmt_no

    all_mgmt_nos = sorted(
        set(phase_row_map.keys()) | set(final_result_rows.keys()),
        key=row_sort_key,
    )
    building_map = _fetch_buildings(
        db,
        all_mgmt_nos,
        current_user,
    )

    items: list[dict[str, object]] = []
    summary = {
        "total_rows": len(all_mgmt_nos),
        "matched": 0,
        "mismatched": 0,
        "missing_db": 0,
        "excel_phase_missing": 0,
        "compared": 0,
        "final_result_compared": 0,
        "final_result_matched": 0,
        "final_result_mismatched": 0,
        "final_result_missing_db": 0,
        "excel_final_result_missing": 0,
    }

    for mgmt_no in all_mgmt_nos:
        row = phase_row_map.get(mgmt_no)
        building = building_map.get(mgmt_no)
        excel_phase = row["phase"] if row and isinstance(row["phase"], str) else None
        status = _status_for(building, excel_phase) if row else "not_checked"
        db_phase = building.current_phase if building else None
        gap = _phase_gap(db_phase, excel_phase)

        if status == "matched":
            summary["matched"] += 1
            summary["compared"] += 1
        elif status == "mismatch":
            summary["mismatched"] += 1
            summary["compared"] += 1
        elif status == "missing_db":
            summary["missing_db"] += 1
        elif status == "excel_phase_missing":
            summary["excel_phase_missing"] += 1

        final_row = final_result_rows.get(mgmt_no)
        excel_final_result = (
            final_row["excel_final_result"]
            if final_row and isinstance(final_row["excel_final_result"], str)
            else None
        )
        final_status = _final_result_status(
            building,
            excel_final_result,
            checked=final_row is not None,
        )
        if final_status == "matched":
            summary["final_result_matched"] += 1
            summary["final_result_compared"] += 1
        elif final_status == "mismatch":
            summary["final_result_mismatched"] += 1
            summary["final_result_compared"] += 1
        elif final_status == "missing_db":
            summary["final_result_missing_db"] += 1
        elif final_status == "excel_final_result_missing":
            summary["excel_final_result_missing"] += 1

        items.append(
            {
                "row_number": row["row_number"] if row else final_row["management_row_number"],
                "management_row_number": final_row["management_row_number"] if final_row else None,
                "mgmt_no": mgmt_no,
                "building_id": building.id if building else None,
                "building_name": building.building_name if building else None,
                "reviewer_name": _reviewer_name(building) if building else None,
                "excel_phase": excel_phase,
                "db_phase": db_phase,
                "excel_final_result": excel_final_result,
                "excel_final_result_raw": final_row["excel_final_result_raw"] if final_row else None,
                "db_final_result": building.final_result if building else None,
                "final_result_status": final_status,
                "final_result_matched": final_status == "matched" if final_status != "not_checked" else None,
                "final_result_column": final_row["final_result_column"] if final_row else None,
                "status": status,
                "matched": status == "matched",
                "phase_gap": gap,
                "phase_direction": _phase_direction(gap),
                "evidence_round": row["evidence_round"] if row else None,
                "evidence_column": row["evidence_column"] if row else None,
                "evidence_value": row["evidence_value"] if row else None,
                "evidence_label": row["evidence_label"] if row else None,
                "rounds": row["rounds"] if row else [],
            }
        )

    return {
        **summary,
        "sheet": SUPPLEMENT_SHEET_NAME,
        "management_sheet": management_sheet,
        "items": items,
    }


def apply_final_results_from_ledger(
    file_path: str | Path,
    db: Session,
    *,
    actor_user_id: int | None,
) -> dict[str, object]:
    """통합 관리대장 CW열 최종판정만 DB에 반영한다."""
    workbook = load_workbook(str(file_path), data_only=True, read_only=True)
    try:
        sheet_name, final_result_rows = _read_final_result_rows(workbook)
        if sheet_name is None:
            raise ValueError(f"시트 '{MANAGEMENT_SHEET_NAME}'을 찾을 수 없습니다")
    finally:
        workbook.close()

    building_map: dict[str, Building] = {}
    mgmt_nos = sorted(final_result_rows.keys())
    for start in range(0, len(mgmt_nos), 1000):
        chunk = mgmt_nos[start:start + 1000]
        for building in db.query(Building).filter(Building.mgmt_no.in_(chunk)).all():
            building_map[building.mgmt_no] = building

    warnings: list[str] = []
    result: dict[str, object] = {
        "sheet": sheet_name,
        "final_result_column": FINAL_RESULT_COLUMN,
        "total_rows": len(final_result_rows),
        "updated": 0,
        "matched": 0,
        "missing_db": 0,
        "excel_final_result_missing": 0,
        "warnings": warnings,
    }

    for mgmt_no, row in final_result_rows.items():
        building = building_map.get(mgmt_no)
        final_result = row["excel_final_result"]
        if building is None:
            result["missing_db"] = int(result["missing_db"]) + 1
            warnings.append(f"{mgmt_no}: DB 건축물이 없어 최종판정을 반영하지 않았습니다.")
            continue
        if not isinstance(final_result, str) or not final_result:
            result["excel_final_result_missing"] = int(result["excel_final_result_missing"]) + 1
            continue
        if building.final_result == final_result:
            result["matched"] = int(result["matched"]) + 1
            continue

        _apply_final_result_update(
            db,
            building,
            row_number=int(row["management_row_number"]),
            final_result=final_result,
            final_result_raw=row["excel_final_result_raw"] if isinstance(row["excel_final_result_raw"], str) else None,
            actor_user_id=actor_user_id,
        )
        result["updated"] = int(result["updated"]) + 1

    return result
