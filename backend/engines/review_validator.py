"""검토서 유효성 검증 엔진

샘플: .doc/2025-0005.xlsm 참조
검토서는 단일 시트(.xlsm)로 구성

셀 위치:
  C4: 관리번호
  F4: 검토위원
  H4: 검토결과
  C5: 절차 (예: "1차 적정성 검토")
  F7: 건축사 소속
  H7: 건축사 성명
  F8: 책임구조기술자 소속
  H8: 책임구조기술자 성명
  F9: 도면작성자 소속 (추정 위치 — 실제 검토서 양식에 맞게 확인 필요)
  H9: 도면작성자 성명
  F11: 주요 구조형식
  F12: 내진등급 (특/I/II)
  E13~F13: 구조도면 작성자 자격 (건축사/건축구조기술사/기타)
  E14~F14: 고위험 유형
  C9: 전이구조 (O/X) — 우측에 "필로티" 있으면 필로티 구분
  C8~C15: 유형별 상세검토 (공법/전이구조/면진제진/특수전단벽/무량판/캔틸래버/장스팬/고층)
  D79: 적정성 검토 결과
  G81: 판정결과 부적합 유형 1
  G83: 판정결과 부적합 유형 2 (또는 이후 행)
  시트명: "검토서 (1차)" 등
"""

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from engines.review_opinion_parser import parse_review_opinions

MGMT_NO_PATTERN = re.compile(r"^\d{4}-\d{4}$")

REVIEW_RESULT_ALIASES = {
    "적합": "적합",
    "단순오류": "단순오류",
    "경미": "단순오류",
    "재계산": "재계산",
    "보완": "재계산",
    "부적합": "재계산",
}


def _expected_phase_labels(
    expected_phase: str | None,
) -> tuple[str, tuple[str, ...], str] | None:
    """업로드 단계에 맞는 절차(C5) 값과 시트명 차수 라벨을 반환한다."""
    if expected_phase == "preliminary":
        return ("예비검토", ("1차 적정성 검토",), "1차")

    match = re.fullmatch(r"supplement_(\d+)", expected_phase or "")
    if match:
        supplement_order = match.group(1)
        return (
            f"보완검토({supplement_order}차)",
            (f"2차 적정성 검토({supplement_order})", "2차 적정성 검토"),
            "2차",
        )

    return None


@dataclass
class ValidationResult:
    is_valid: bool = True
    mgmt_no: str = ""
    reviewer_name: str = ""
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    extracted_data: dict = field(default_factory=dict)

    def add_error(self, msg: str):
        self.errors.append(msg)
        self.is_valid = False

    def add_warning(self, msg: str):
        self.warnings.append(msg)


def _cell_str(ws, coord: str) -> str:
    """셀 값을 문자열로 반환 (_x000D_ 등 특수문자 정리)"""
    val = ws[coord].value
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace("_x000D_", "").replace("_x000d_", "").replace("\r", "")
    return s


def _find_result_row(ws, start_row: int = 1, max_row: int | None = None) -> int:
    """적정성 검토 결과 행 찾기 (행이 추가되면 이동할 수 있음)"""
    end_row = max_row or ws.max_row + 1
    for row in range(start_row, end_row):
        val = ws.cell(row=row, column=2).value  # B열
        if val and "적정성 검토 결과" in str(val):
            return row
    return 79  # 기본값


def _find_defect_type_rows(ws, start_row: int = 1, max_row: int | None = None) -> list[int]:
    """판정결과 부적합 유형 행 찾기 (F열에 "판정결과" 라벨이 있는 행부터 최대 6행)"""
    end_row = max_row or ws.max_row + 1
    label_row = None
    for row in range(start_row, end_row):
        val = ws.cell(row=row, column=6).value  # F열
        if val and "판정결과" in str(val) and "부적합" in str(val):
            label_row = row
            break

    if not label_row:
        return []

    rows = []
    for row in range(label_row, min(label_row + 20, end_row)):
        val = ws.cell(row=row, column=7).value  # G열
        if val and str(val).strip():
            rows.append(row)
        if len(rows) >= 3:
            break
    return rows


def _expected_review_result(opinion_parse) -> tuple[str, str]:
    """상세의견/심각도 기준으로 기대 검토결과를 계산."""
    if not opinion_parse.entries:
        return "적합", "상세내용이 없습니다"
    counts = opinion_parse.severity_counts
    if counts.get("L3", 0) > 0 or counts.get("L4", 0) > 0:
        return "재계산", "심각도 L3/L4가 1건 이상 있습니다"
    return "단순오류", "상세내용은 있으나 심각도 L3/L4가 없습니다"


def _normalized_review_result(value: str) -> str:
    return REVIEW_RESULT_ALIASES.get(value.strip(), value.strip())


def extract_mgmt_no_from_filename(filename: str) -> str | None:
    """파일명에서 관리번호 추출 (예: 2025-0005.xlsm)"""
    stem = Path(filename).stem
    parts = stem.split("_")
    candidate = parts[0].strip()
    if MGMT_NO_PATTERN.match(candidate):
        return candidate
    return None


def validate_review_file(
    file_path: str | Path,
    filename: str,
    expected_mgmt_no: str | None,
    submitter_name: str | None,
    expected_phase: str | None = None,
    submitter_label: str = "로그인 사용자",
) -> ValidationResult:
    """검토서 파일 유효성 검증

    검증 항목:
    1. 파일명 = 관리번호.xlsm
    2. 관리번호 일치 (파일명 vs 내부 C4)
    3. 검토위원 일치 (기대 검토위원 vs F4)
    4. 적정성 검토 결과가 "적합"이면 부적합유형1은 "적합", 나머지 빈칸
    5. 적정성 검토 결과에 내용이 있는데 부적합유형이 "적합" 또는 비어있는 경우
    6. 차수 라벨 검증 (시트명 + 절차 C5)
       - expected_phase=preliminary → "1차 적정성 검토" / "검토서(1차)"
       - expected_phase=supplement_N → "2차 적정성 검토(N)" 또는 "2차 적정성 검토" / "검토서(2차)"
    7. 시트가 2개 이상이면 안됨
    """
    result = ValidationResult()

    # 1. 파일 확장자 검증 — .xlsm만 허용
    suffix = Path(filename).suffix.lower()
    if suffix != ".xlsm":
        result.add_error(f"지원하지 않는 파일 형식입니다: {suffix} (.xlsm 파일만 업로드할 수 있습니다)")
        return result

    # 파일명 관리번호 확인
    file_mgmt_no = extract_mgmt_no_from_filename(filename)
    if not file_mgmt_no:
        result.add_error(
            f"파일명에서 관리번호를 인식할 수 없습니다: '{filename}'. "
            "파일명은 '관리번호.xlsm' 형식이어야 합니다. (예: 2025-0001.xlsm)"
        )
        return result
    result.mgmt_no = file_mgmt_no

    # URL 관리번호와 비교
    if expected_mgmt_no and file_mgmt_no != expected_mgmt_no:
        result.add_error(
            f"파일명의 관리번호({file_mgmt_no})가 "
            f"검토 대상 관리번호({expected_mgmt_no})와 일치하지 않습니다."
        )

    try:
        wb = load_workbook(str(file_path), data_only=True, read_only=True)
    except Exception as e:
        result.add_error(f"엑셀 파일을 열 수 없습니다: {str(e)}")
        return result

    # 7. 시트 수 검증
    if len(wb.sheetnames) > 1:
        result.add_error(f"시트가 {len(wb.sheetnames)}개입니다. 검토서는 1개의 시트만 있어야 합니다.")

    # 6. 시트명 검증
    sheet_name = wb.sheetnames[0]
    if "검토서" not in sheet_name:
        result.add_warning(f"시트명이 '{sheet_name}'입니다. '검토서(1차)' 형식이어야 합니다.")

    ws = wb[sheet_name]

    # 2. 내부 관리번호 (C4) 확인
    internal_mgmt_no = _cell_str(ws, "C4")
    if internal_mgmt_no and MGMT_NO_PATTERN.match(internal_mgmt_no):
        if internal_mgmt_no != file_mgmt_no:
            result.add_error(
                f"검토서 내부 관리번호({internal_mgmt_no})가 "
                f"파일명의 관리번호({file_mgmt_no})와 일치하지 않습니다."
            )
    else:
        result.add_warning(f"검토서 내부 관리번호(C4)가 비어있거나 형식이 다릅니다: '{internal_mgmt_no}'")

    # 3. 검토위원 (F4) 확인
    internal_reviewer = _cell_str(ws, "F4")
    result.reviewer_name = internal_reviewer
    if submitter_name and internal_reviewer and internal_reviewer != submitter_name:
        result.add_error(
            f"검토서 내부 검토위원({internal_reviewer})이 "
            f"{submitter_label}({submitter_name})와 일치하지 않습니다."
        )

    # 6. 절차 (C5) 및 시트명 검증 — expected_phase에 맞춰 차수 검증
    #   예비검토(preliminary)     → "1차 적정성 검토" / "검토서(1차)"
    #   보완검토(supplement_N)    → "2차 적정성 검토(N)" 또는 "2차 적정성 검토" / "검토서(2차)"
    procedure = _cell_str(ws, "C5")
    expected_labels = _expected_phase_labels(expected_phase)

    if expected_labels:
        round_label, expected_procedures, expected_sheet_round = expected_labels
        if procedure and procedure not in expected_procedures:
            procedure_label = " 또는 ".join(f"'{value}'" for value in expected_procedures)
            result.add_error(
                f"{round_label} 업로드인데 절차(C5)가 '{procedure}'입니다. "
                f"{procedure_label}여야 합니다."
            )
        if expected_sheet_round not in sheet_name:
            result.add_error(
                f"{round_label} 업로드인데 시트명이 '{sheet_name}'입니다. "
                f"'검토서({expected_sheet_round})' 형식이어야 합니다."
            )
    else:
        # phase 정보 없으면 기존 휴리스틱 (1차/2차 라벨 기반)
        if "1차" in sheet_name or "1차" in procedure:
            if procedure and "1차" not in procedure:
                result.add_error(f"예비검토인데 절차(C5)가 '{procedure}'입니다.")
            if "1차" not in sheet_name:
                result.add_error(f"예비검토인데 시트명이 '{sheet_name}'입니다.")
        elif "2차" in sheet_name or "2차" in procedure:
            if procedure and "2차" not in procedure:
                result.add_error(f"보완검토인데 절차(C5)가 '{procedure}'입니다.")
            if "2차" not in sheet_name:
                result.add_error(f"보완검토인데 시트명이 '{sheet_name}'입니다.")

    # 상세의견 + 심각도 파싱. 저장용 검토의견은 D81 수식값이 아니라 상세의견에서 구성한다.
    opinion_parse = parse_review_opinions(ws)
    for error in opinion_parse.errors:
        result.add_error(error)

    # 적정성 검토 결과 행 찾기
    result_row = _find_result_row(ws)
    review_result_text = _cell_str(ws, f"D{result_row}")

    # 판정결과 부적합 유형 찾기
    defect_rows = _find_defect_type_rows(ws, start_row=result_row)
    defect_type_1 = _cell_str(ws, f"G{defect_rows[0]}") if len(defect_rows) > 0 else ""
    defect_type_2 = _cell_str(ws, f"G{defect_rows[1]}") if len(defect_rows) > 1 else ""
    defect_type_3 = _cell_str(ws, f"G{defect_rows[2]}") if len(defect_rows) > 2 else ""

    # 4. 적정성 검토 결과가 "적합"이면 부적합유형 검증
    review_result_value = _cell_str(ws, "H4")  # 검토결과
    expected_review_result, expected_reason = _expected_review_result(opinion_parse)
    normalized_review_result = _normalized_review_result(review_result_value)
    if normalized_review_result != expected_review_result:
        current_label = review_result_value or "빈값"
        result.add_error(
            f"검토결과(H4)는 상세의견/심각도 기준 '{expected_review_result}'이어야 합니다. "
            f"현재 값: '{current_label}' ({expected_reason})"
        )

    if review_result_value == "적합":
        if defect_type_1 and defect_type_1 != "적합":
            result.add_error(
                f"검토결과가 '적합'인데 판정결과 부적합유형 1이 '{defect_type_1}'입니다. '적합'이어야 합니다."
            )
        if defect_type_2:
            result.add_error(
                f"검토결과가 '적합'인데 판정결과 부적합유형 2에 '{defect_type_2}'가 입력되어 있습니다. 빈칸이어야 합니다."
            )
        if defect_type_3:
            result.add_error(
                f"검토결과가 '적합'인데 판정결과 부적합유형 3에 '{defect_type_3}'가 입력되어 있습니다. 빈칸이어야 합니다."
            )

    # 5. 적정성 검토 결과에 내용이 있는데 부적합유형이 비어있는 경우
    if review_result_text and review_result_text != "적합":
        if not defect_type_1 or defect_type_1 == "적합":
            result.add_error(
                "적정성 검토 결과에 내용이 기입되어 있는데 판정결과 부적합유형이 '적합' 또는 비어있습니다."
            )

    # 데이터 추출
    # 전이구조 필로티 확인
    transfer_structure = _cell_str(ws, "C9")
    is_piloti = False
    d9_val = _cell_str(ws, "D9")
    if "필로티" in d9_val:
        is_piloti = True

    result.extracted_data = {
        "mgmt_no": internal_mgmt_no,
        "reviewer_name": internal_reviewer,
        "review_result": review_result_value,
        "procedure": procedure,
        "architect_firm": _cell_str(ws, "F7"),
        "architect_name": _cell_str(ws, "H7"),
        "struct_eng_firm": _cell_str(ws, "F8"),
        "struct_eng_name": _cell_str(ws, "H8"),
        "main_structure_type": _cell_str(ws, "F11"),
        "seismic_level": _cell_str(ws, "F12"),
        "struct_drawing_qual": _cell_str(ws, "F13"),
        "high_risk_type": _cell_str(ws, "F14"),
        "defect_type_1": defect_type_1,
        "defect_type_2": defect_type_2,
        "defect_type_3": defect_type_3,
        "review_opinion": opinion_parse.formatted_text,
        "severity_counts": opinion_parse.severity_counts,
        "category_severity_counts": [
            {
                "category": item.category,
                "severity": item.severity,
                "count": item.count,
            }
            for item in opinion_parse.category_severity_counts
        ],
        "opinion_entries": [
            {
                "row": entry.row,
                "category": entry.section,
                "severity": entry.severity,
                "content": entry.content,
            }
            for entry in opinion_parse.entries
        ],
        # 유형별 상세검토
        "type_construction_method": _cell_str(ws, "C8"),     # 공법
        "type_transfer_structure": transfer_structure,         # 전이구조
        "type_is_piloti": is_piloti,                          # 필로티 여부
        "type_seismic_isolation": _cell_str(ws, "C10"),       # 면진&제진
        "type_special_shear_wall": _cell_str(ws, "C11"),      # 특수전단벽
        "type_flat_plate": _cell_str(ws, "C12"),              # 무량판
        "type_cantilever": _cell_str(ws, "C13"),              # 캔틸래버
        "type_long_span": _cell_str(ws, "C14"),               # 장스팬
        "type_high_rise": _cell_str(ws, "C15"),               # 고층
        "sheet_name": sheet_name,
        "sheet_count": len(wb.sheetnames),
    }

    # 제출전 확인 경고
    if not result.extracted_data["architect_firm"] and not result.extracted_data["architect_name"]:
        result.add_warning("건축사 소속/성명이 입력되지 않았습니다.")
    if not result.extracted_data["struct_eng_firm"] and not result.extracted_data["struct_eng_name"]:
        result.add_warning("책임구조기술자 소속/성명이 입력되지 않았습니다.")
    # 내진등급 (F12) — 특/I/II 중 하나여야 함
    _seismic = result.extracted_data["seismic_level"]
    if not _seismic:
        result.add_warning("내진등급이 입력되지 않았습니다.")
    elif _seismic not in ("특", "I", "II"):
        result.add_warning(
            f"내진등급이 '{_seismic}'입니다. 특/I/II 중 하나여야 합니다."
        )
    # 도면작성자 자격 (F13) — 건축사/건축구조기술사/기타 중 하나여야 함
    _qual = result.extracted_data["struct_drawing_qual"]
    if _qual and _qual not in ("건축사", "건축구조기술사", "기타"):
        result.add_warning(
            f"도면작성자 자격이 '{_qual}'입니다. 건축사/건축구조기술사/기타 중 하나여야 합니다."
        )
    if not result.extracted_data["main_structure_type"]:
        result.add_warning("주요 구조형식이 입력되지 않았습니다.")
    if not result.extracted_data["high_risk_type"]:
        result.add_warning("고위험 유형이 입력되지 않았습니다.")
    if not review_result_text and not opinion_parse.formatted_text:
        result.add_warning("적정성 검토 결과가 비어있습니다.")

    wb.close()
    return result
