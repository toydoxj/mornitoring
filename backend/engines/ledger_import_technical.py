"""2026년 기술사회 배포용 관리대장 엑셀 → DB Import 엔진

시트명: 관리대장
Row 3: 대분류 헤더
Row 4: 상세 컬럼명
Row 5~: 데이터
"""

import re
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.building import Building
from models.review_opinion_detail import ReviewOpinionDetail
from models.review_severity_summary import ReviewSeveritySummary
from models.review_stage import ReviewStage, PhaseType, ResultType
from services.phase_transition import transition_phase
from engines.column_mapping import col_letter_to_index
from engines.opinion_text import clean_opinion_detail_content

DATA_START_ROW = 5
SHEET_NAME = "관리대장"
SEVERITY_LABELS = ("L0", "L1", "L2", "L3", "L4")
PHASE_INDEX = {
    None: -1,
    "assigned": 0,
    "doc_received": 1,
    "preliminary": 2,
    "supplement_1_received": 3,
    "supplement_1": 4,
    "supplement_2_received": 5,
    "supplement_2": 6,
    "supplement_3_received": 7,
    "supplement_3": 8,
    "supplement_4_received": 9,
    "supplement_4": 10,
    "supplement_5_received": 11,
    "supplement_5": 12,
}

# 2026년 기술사회 배포용 관리대장 열 매핑 (Row 4 기준)
BUILDING_COLUMN_MAP_TECHNICAL = {
    "A": "mgmt_no",              # 모니터링 관리번호
    "F": "building_type",        # 건축구분
    "H": "sido",                 # 시도명
    "I": "sigungu",              # 시군구명
    "J": "beopjeongdong",        # 법정동명
    "K": "land_type",            # 대지구분
    "L": "main_lot_no",          # 본번
    "M": "sub_lot_no",           # 부번
    "N": "special_lot_no",       # 특수지번
    "O": "building_name",        # 건물명
    "P": "gross_area",           # 연면적
    "Q": "main_structure",       # 주구조
    "R": "other_structure",      # 기타구조
    "S": "main_usage",           # 주용도
    "T": "other_usage",          # 기타용도
    "W": "height",               # 높이
    "X": "floors_above",         # 지상층수
    "Y": "floors_below",         # 지하층수
    "AD": "architect_name",      # 설계자
    "AE": "architect_firm",      # 설계사무소
    "AO": "is_special_structure", # 특수구조물 여부
    "AP": "is_high_rise",        # 고층 여부
    "AQ": "is_multi_use",        # 다중이용건축물 여부
    "AS": "remarks",             # 비고
}

PRELIMINARY_MAP_TECHNICAL = {
    "AT": "reviewer_name",       # 검토자
    "AV": "defect_type_1",       # 부적합유형-1
    "AW": "defect_type_2",       # 부적합유형-2
    "AX": "defect_type_3",       # 부적합유형-3
    "AZ": "review_opinion",      # 예비 검토의견
}

SUPPLEMENT_1_MAP_TECHNICAL = {
    "BG": "reviewer_name",       # 검토자
    "BH": "result",              # 판정 결과
    "BI": "defect_type_1",       # 부적합유형-1
    "BJ": "defect_type_2",       # 부적합유형-2
    "BK": "defect_type_3",       # 부적합유형-3
    "BL": "review_opinion",      # 보완자료 판정결과 검토의견
    "BM": "stage_remarks",       # 비고
}

ASSIGNED_REVIEWER_COLUMNS = ("AT", "BG", "B")
HIGH_RISK_COLUMN = "AR"
PRELIMINARY_DECISION_COLUMN = "AU"
PRELIMINARY_ADMIN_RESULT_COLUMN = "AY"
UNCLASSIFIED_SEVERITY = "NA"


def _cell_value(row: tuple, col_letter: str):
    idx = col_letter_to_index(col_letter)
    if idx >= len(row):
        return None
    val = row[idx].value
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None
    return val


def _text_cell_value(row: tuple, col_letter: str) -> str | None:
    val = _cell_value(row, col_letter)
    text = _clean_text(val)
    return text or None


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _to_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _to_bool(val) -> bool | None:
    if val is None:
        return None
    s = str(val).strip()
    if s in ("Y", "예", "○", "O", "o", "1", "True", "true", "해당"):
        return True
    if s in ("N", "아니오", "×", "X", "x", "0", "False", "false", "미해당", "-"):
        return False
    return None


def _to_high_risk_type(val) -> str | None:
    is_high_risk = _to_bool(val)
    if is_high_risk is True:
        return "고위험"
    if is_high_risk is False or val is None:
        return None
    return str(val).strip()


def _clean_text(val) -> str:
    if val is None:
        return ""
    text = str(val).strip()
    text = re.sub(r"_?x000D_", "\n", text, flags=re.IGNORECASE)
    text = text.replace("\r", "\n")
    text = re.sub(r"\n\s*/\s*", "\n", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n[ \t]+", "\n", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def _parse_result(val) -> ResultType | None:
    if val is None:
        return None
    s = str(val).strip()
    mapping = {
        "적합": ResultType.PASS,
        "단순오류": ResultType.SIMPLE_ERROR,
        "경미": ResultType.SIMPLE_ERROR,
        "재계산": ResultType.RECALCULATE,
        "보완": ResultType.RECALCULATE,
        "부적합": ResultType.RECALCULATE,
    }
    return mapping.get(s)


def _is_mgmt_no(value) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return len(text) >= 9 and "-" in text


def _first_present(row: tuple, col_letters: tuple[str, ...]) -> str | None:
    for col_letter in col_letters:
        value = _text_cell_value(row, col_letter)
        if value:
            return value
    return None


def _split_numbered_opinions(body: str) -> list[str]:
    body = body.strip()
    if not body:
        return []

    parts = re.split(r"(?m)(?=^\s*\d+\s*[.)．]\s*)", body)
    items = []
    for part in parts:
        item = re.sub(r"^\s*\d+\s*[.)．]\s*", "", part).strip()
        if item:
            items.append(re.sub(r"\s+", " ", item))
    if items:
        return items

    return [
        re.sub(r"\s+", " ", line).strip()
        for line in body.splitlines()
        if line.strip()
    ]


def _parse_ledger_opinion_entries(text: str) -> list[dict[str, str | int | None]]:
    text = _clean_text(text)
    if not text:
        return []

    matches = list(re.finditer(r"\[([^\]]+)\]", text))
    if not matches:
        return [
            {"row": None, "category": "기타의견", "content": item}
            for item in _split_numbered_opinions(text)
        ]

    entries: list[dict[str, str | int | None]] = []
    for idx, match in enumerate(matches):
        category = re.sub(r"\s+", " ", match.group(1)).strip()
        start = match.end()
        end = matches[idx + 1].start() if idx + 1 < len(matches) else len(text)
        body = text[start:end].strip()
        for item in _split_numbered_opinions(body):
            entries.append({"row": None, "category": category, "content": item})
    return entries


def _replace_ledger_opinion_details(
    db: Session,
    stage: ReviewStage,
    *,
    phase: PhaseType,
    opinion_text: str | None,
) -> None:
    """기술사회 의견은 미분류로 보관하고, 수동 지정 전까지 L0~L4 통계에 넣지 않는다."""
    if stage.id is None:
        db.add(stage)
        db.flush()

    if stage.s3_file_key:
        return

    db.query(ReviewSeveritySummary).filter(
        ReviewSeveritySummary.stage_id == stage.id
    ).delete(synchronize_session="fetch")
    db.query(ReviewOpinionDetail).filter(
        ReviewOpinionDetail.stage_id == stage.id
    ).delete(synchronize_session="fetch")

    for label in SEVERITY_LABELS:
        setattr(stage, f"severity_{label.lower()}_count", 0)

    phase_value = phase.value if hasattr(phase, "value") else str(phase)
    phase_group = "preliminary" if phase == PhaseType.PRELIMINARY else "supplement"
    for index, entry in enumerate(_parse_ledger_opinion_entries(opinion_text or ""), start=1):
        category = str(entry.get("category") or "").strip()
        content = clean_opinion_detail_content(entry.get("content"))
        if not category or not content:
            continue
        db.add(ReviewOpinionDetail(
            stage_id=stage.id,
            phase=phase_value,
            phase_group=phase_group,
            row_number=index,
            category=category,
            severity=UNCLASSIFIED_SEVERITY,
            content=content,
        ))


def _find_sheet(wb) -> str | None:
    if SHEET_NAME in wb.sheetnames:
        return SHEET_NAME
    for sheet_name in wb.sheetnames:
        if "관리대장" in sheet_name:
            return sheet_name
    return None


def _find_stage(building: Building, phase: PhaseType) -> ReviewStage | None:
    for stage in building.stages:
        if stage.phase == phase:
            return stage
    return None


def _apply_building_data(building: Building, building_data: dict) -> None:
    for field_name, value in building_data.items():
        setattr(building, field_name, value)


def _apply_stage_data(stage: ReviewStage, stage_data: dict) -> None:
    for field_name, value in stage_data.items():
        setattr(stage, field_name, value)


def _transition_import_forward(
    db: Session,
    building: Building,
    *,
    to_phase: str,
    actor_user_id: int | None,
) -> None:
    current_rank = PHASE_INDEX.get(building.current_phase, -1)
    target_rank = PHASE_INDEX.get(to_phase, -1)
    if target_rank <= current_rank:
        return
    transition_phase(
        db,
        building,
        to_phase=to_phase,
        trigger="import",
        actor_user_id=actor_user_id,
        reason="ledger_import_technical",
    )


def import_ledger_technical(
    file_path: str | Path,
    db: Session,
    actor_user_id: int | None = None,
) -> dict:
    """2026년 기술사회 배포용 관리대장 파일을 DB에 import한다."""
    wb = load_workbook(str(file_path), data_only=True, read_only=True)
    matched_sheet = _find_sheet(wb)

    if matched_sheet is None:
        wb.close()
        return {"imported": 0, "updated": 0, "skipped": 0, "errors": ["관리대장 시트를 찾을 수 없습니다"]}

    ws = wb[matched_sheet]
    rows_parsed = []
    for row in ws.iter_rows(min_row=DATA_START_ROW):
        mgmt_no = _cell_value(row, "A")
        if not _is_mgmt_no(mgmt_no):
            continue
        rows_parsed.append((str(mgmt_no).strip(), row))
    wb.close()

    if not rows_parsed:
        return {"imported": 0, "updated": 0, "skipped": 0, "errors": [], "sheet": matched_sheet}

    all_mgmt_nos = [mgmt_no for mgmt_no, _ in rows_parsed]
    existing_map: dict[str, Building] = {}
    for i in range(0, len(all_mgmt_nos), 1000):
        chunk = all_mgmt_nos[i:i + 1000]
        existing = db.query(Building).filter(Building.mgmt_no.in_(chunk)).all()
        existing_map.update((building.mgmt_no, building) for building in existing)

    result = {"imported": 0, "updated": 0, "skipped": 0, "errors": [], "sheet": matched_sheet}
    batch_count = 0

    for mgmt_no, row in rows_parsed:
        building_data = {}
        for col_letter, field_name in BUILDING_COLUMN_MAP_TECHNICAL.items():
            val = _text_cell_value(row, col_letter)
            if field_name in ("gross_area", "height"):
                val = _to_float(_cell_value(row, col_letter))
            elif field_name in ("floors_above", "floors_below"):
                val = _to_int(_cell_value(row, col_letter))
            elif field_name in ("is_special_structure", "is_high_rise", "is_multi_use"):
                val = _to_bool(_cell_value(row, col_letter))
            building_data[field_name] = val

        assigned_reviewer_name = _first_present(row, ASSIGNED_REVIEWER_COLUMNS)
        if assigned_reviewer_name:
            building_data["assigned_reviewer_name"] = assigned_reviewer_name

        high_risk_type = _to_high_risk_type(_cell_value(row, HIGH_RISK_COLUMN))
        building_data["high_risk_type"] = high_risk_type

        building = existing_map.get(mgmt_no)
        is_new = building is None
        if building is None:
            building = Building(**building_data)
            db.add(building)
            db.flush()
            existing_map[mgmt_no] = building
        else:
            _apply_building_data(building, building_data)

        prelim_data = {}
        for col_letter, field_name in PRELIMINARY_MAP_TECHNICAL.items():
            val = _text_cell_value(row, col_letter)
            prelim_data[field_name] = val

        preliminary_decision = _text_cell_value(row, PRELIMINARY_DECISION_COLUMN)
        preliminary_admin_result = _text_cell_value(row, PRELIMINARY_ADMIN_RESULT_COLUMN)
        prelim_result = _parse_result(preliminary_decision) or _parse_result(preliminary_admin_result)
        prelim_data["result"] = prelim_result
        remarks = []
        if preliminary_decision:
            remarks.append(f"판정의견: {preliminary_decision}")
        if preliminary_admin_result:
            remarks.append(f"관리원 입력 예비판정 결과: {preliminary_admin_result}")
        prelim_data["stage_remarks"] = "\n".join(remarks) if remarks else None

        if any(v is not None for v in prelim_data.values()):
            stage = _find_stage(building, PhaseType.PRELIMINARY)
            if stage is None:
                stage = ReviewStage(
                    building_id=building.id,
                    phase=PhaseType.PRELIMINARY,
                    phase_order=0,
                    **prelim_data,
                )
                db.add(stage)
            else:
                _apply_stage_data(stage, prelim_data)
            db.flush()
            _replace_ledger_opinion_details(
                db,
                stage,
                phase=PhaseType.PRELIMINARY,
                opinion_text=prelim_data.get("review_opinion"),
            )
            _transition_import_forward(
                db,
                building,
                to_phase="preliminary",
                actor_user_id=actor_user_id,
            )

        supp1_data = {}
        for col_letter, field_name in SUPPLEMENT_1_MAP_TECHNICAL.items():
            val = _text_cell_value(row, col_letter)
            if field_name == "result":
                val = _parse_result(_cell_value(row, col_letter))
            supp1_data[field_name] = val

        if any(v is not None for v in supp1_data.values()):
            stage = _find_stage(building, PhaseType.SUPPLEMENT_1)
            if stage is None:
                stage = ReviewStage(
                    building_id=building.id,
                    phase=PhaseType.SUPPLEMENT_1,
                    phase_order=1,
                    **supp1_data,
                )
                db.add(stage)
            else:
                _apply_stage_data(stage, supp1_data)
            db.flush()
            _replace_ledger_opinion_details(
                db,
                stage,
                phase=PhaseType.SUPPLEMENT_1,
                opinion_text=supp1_data.get("review_opinion"),
            )
            _transition_import_forward(
                db,
                building,
                to_phase="supplement_1",
                actor_user_id=actor_user_id,
            )

        if is_new:
            result["imported"] += 1
        else:
            result["updated"] += 1
        batch_count += 1

        if batch_count % 500 == 0:
            db.commit()

    db.commit()
    return result
