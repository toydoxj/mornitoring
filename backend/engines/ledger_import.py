"""통합관리대장 엑셀 → DB Import 엔진"""

from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.building import Building
from models.review_stage import ReviewStage, PhaseType, ResultType
from engines.column_mapping import (
    BUILDING_COLUMN_MAP,
    REVIEWER_COLUMN,
    PRELIMINARY_STAGE_MAP,
    SUPPLEMENT_SUBMIT_START_COLS,
    SUPPLEMENT_SUBMIT_OFFSETS,
    SUPPLEMENT_REVIEW_START_COLS,
    SUPPLEMENT_REVIEW_OFFSETS,
    FINAL_RESULT_COLUMN,
    col_letter_to_index,
)

# 데이터 시작 행 (Row 1~2: 헤더, Row 3~: 데이터)
DATA_START_ROW = 3
SHEET_NAME = "통합 관리대장"


def _cell_value(row: tuple, col_letter: str):
    """행 데이터에서 열 문자 기준으로 값 추출"""
    idx = col_letter_to_index(col_letter)
    if idx >= len(row):
        return None
    val = row[idx].value
    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None
    return val


def _to_date(val) -> date | None:
    """다양한 형식의 날짜 값을 date로 변환"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _to_bool(val) -> bool | None:
    """여부 값을 bool로 변환"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("Y", "예", "○", "O", "1", "True", "true"):
        return True
    if s in ("N", "아니오", "×", "X", "0", "False", "false"):
        return False
    return None


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _to_int(val) -> int | None:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _parse_result(val) -> ResultType | None:
    """판정 결과 문자열을 ResultType으로 변환"""
    if val is None:
        return None
    s = str(val).strip()
    mapping = {
        "적합": ResultType.PASS,
        "보완": ResultType.SUPPLEMENT,
        "부적합": ResultType.FAIL,
        "경미": ResultType.MINOR,
    }
    return mapping.get(s)


def _offset_col(base_col: str, offset: int) -> str:
    """기준 열로부터 offset만큼 이동한 열 문자 반환"""
    from engines.column_mapping import index_to_col_letter
    base_idx = col_letter_to_index(base_col)
    return index_to_col_letter(base_idx + offset)


def import_ledger(file_path: str | Path, db: Session) -> dict:
    """통합관리대장 엑셀 파일을 DB에 import

    Returns:
        {"imported": int, "skipped": int, "errors": list[str]}
    """
    wb = load_workbook(str(file_path), data_only=True, read_only=True)

    if SHEET_NAME not in wb.sheetnames:
        return {"imported": 0, "skipped": 0, "errors": [f"시트 '{SHEET_NAME}'을 찾을 수 없습니다"]}

    ws = wb[SHEET_NAME]
    result = {"imported": 0, "skipped": 0, "errors": []}

    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW), start=DATA_START_ROW):
        mgmt_no = _cell_value(row, "A")
        if not mgmt_no:
            continue

        mgmt_no = str(mgmt_no).strip()

        # 중복 체크
        existing = db.query(Building).filter(Building.mgmt_no == mgmt_no).first()
        if existing:
            result["skipped"] += 1
            continue

        # 건축물 기본정보 파싱
        building_data = {}
        for col_letter, field_name in BUILDING_COLUMN_MAP.items():
            val = _cell_value(row, col_letter)
            if field_name in ("gross_area", "height"):
                val = _to_float(val)
            elif field_name in ("floors_above", "floors_below"):
                val = _to_int(val)
            elif field_name in ("is_special_structure", "is_high_rise", "is_multi_use",
                                "related_tech_coop", "drawing_creation"):
                val = _to_bool(val)
            building_data[field_name] = val

        # 최종 판정
        final = _cell_value(row, FINAL_RESULT_COLUMN)
        if final:
            building_data["final_result"] = str(final)

        building = Building(**building_data)
        db.add(building)
        db.flush()  # ID 할당

        # 예비검토 단계 파싱
        prelim_data = {}
        for col_letter, field_name in PRELIMINARY_STAGE_MAP.items():
            val = _cell_value(row, col_letter)
            if field_name in ("doc_received_at", "report_submitted_at"):
                val = _to_date(val)
            elif field_name == "result":
                val = _parse_result(val)
            prelim_data[field_name] = val

        # 예비검토에 데이터가 있으면 stage 생성
        if any(v is not None for v in prelim_data.values()):
            stage = ReviewStage(
                building_id=building.id,
                phase=PhaseType.PRELIMINARY,
                phase_order=0,
                **prelim_data,
            )
            db.add(stage)
            building.current_phase = "preliminary"

        # 보완 단계 파싱 (1차~4차)
        phase_types = [PhaseType.SUPPLEMENT_1, PhaseType.SUPPLEMENT_2,
                       PhaseType.SUPPLEMENT_3, PhaseType.SUPPLEMENT_4]

        for supp_no in range(1, 5):
            submit_start = SUPPLEMENT_SUBMIT_START_COLS.get(supp_no)
            review_start = SUPPLEMENT_REVIEW_START_COLS.get(supp_no)
            if not submit_start or not review_start:
                continue

            stage_data: dict = {}

            # 보완 제출 정보
            for offset, field_name in SUPPLEMENT_SUBMIT_OFFSETS.items():
                col = _offset_col(submit_start, offset)
                val = _cell_value(row, col)
                if field_name in ("doc_received_at",):
                    val = _to_date(val)
                elif field_name == "objection_filed":
                    val = _to_bool(val)
                stage_data[field_name] = val

            # 보완 검토 정보
            for offset, field_name in SUPPLEMENT_REVIEW_OFFSETS.items():
                col = _offset_col(review_start, offset)
                val = _cell_value(row, col)
                if field_name in ("report_submitted_at",):
                    val = _to_date(val)
                elif field_name == "result":
                    val = _parse_result(val)
                stage_data[field_name] = val

            # 데이터가 있는 경우만 stage 생성
            if any(v is not None for v in stage_data.values()):
                stage = ReviewStage(
                    building_id=building.id,
                    phase=phase_types[supp_no - 1],
                    phase_order=supp_no,
                    **stage_data,
                )
                db.add(stage)
                building.current_phase = phase_types[supp_no - 1].value

        result["imported"] += 1

    db.commit()
    wb.close()
    return result
