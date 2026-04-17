"""검토위원 배정 엑셀 처리 엔진

엑셀 형식:
- A열: 관리번호
- B열: 검토위원 이름
"""

from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from models.building import Building
from models.reviewer import Reviewer
from models.user import User, UserRole


def preview_assignment(file_path: str | Path, db: Session) -> dict:
    """배정 엑셀을 미리보기 (변경사항 확인용) - 일괄 쿼리 최적화"""
    wb = load_workbook(str(file_path), data_only=True, read_only=True)
    ws = wb.active

    # 엑셀 데이터 먼저 수집
    rows_data = []
    for row in ws.iter_rows(min_row=2):
        mgmt_no_cell = row[0].value
        reviewer_name_cell = row[1].value if len(row) > 1 else None
        if not mgmt_no_cell or not reviewer_name_cell:
            continue
        rows_data.append((str(mgmt_no_cell).strip(), str(reviewer_name_cell).strip()))
    wb.close()

    if not rows_data:
        return {"changes": [], "summary": {"new": 0, "changed": 0, "same": 0, "not_found": 0, "reviewer_not_found": 0}}

    # DB 일괄 조회
    mgmt_nos = [r[0] for r in rows_data]
    buildings = db.query(Building).filter(Building.mgmt_no.in_(mgmt_nos)).all()
    building_map: dict[str, Building] = {b.mgmt_no: b for b in buildings}

    reviewer_names = list(set(r[1] for r in rows_data))
    reviewers = db.query(User).filter(User.name.in_(reviewer_names)).all()
    reviewer_map: dict[str, User] = {u.name: u for u in reviewers}

    changes = []
    summary = {"new": 0, "changed": 0, "same": 0, "not_found": 0, "reviewer_not_found": 0}

    for mgmt_no, reviewer_name in rows_data:
        building = building_map.get(mgmt_no)
        if not building:
            changes.append({"mgmt_no": mgmt_no, "reviewer_name": reviewer_name, "current_reviewer": None, "status": "not_found"})
            summary["not_found"] += 1
            continue

        current_name = None
        if building.reviewer and building.reviewer.user:
            current_name = building.reviewer.user.name

        registered = reviewer_name in reviewer_map

        if current_name == reviewer_name:
            status = "same"
        elif current_name:
            status = "changed"
        else:
            status = "new"

        changes.append({
            "mgmt_no": mgmt_no,
            "reviewer_name": reviewer_name,
            "current_reviewer": current_name,
            "status": status,
            "registered": registered,
        })
        summary[status] += 1

    # 미등록 검토위원 수
    unregistered = set(r[1] for r in rows_data if r[1] not in reviewer_map and building_map.get(r[0]))
    summary["unregistered"] = len(unregistered)

    return {"changes": changes, "summary": summary}


def apply_assignment(file_path: str | Path, db: Session) -> dict:
    """배정 엑셀을 실제 적용 - 일괄 쿼리 최적화"""
    wb = load_workbook(str(file_path), data_only=True, read_only=True)
    ws = wb.active

    rows_data = []
    for row in ws.iter_rows(min_row=2):
        mgmt_no_cell = row[0].value
        reviewer_name_cell = row[1].value if len(row) > 1 else None
        if not mgmt_no_cell or not reviewer_name_cell:
            continue
        rows_data.append((str(mgmt_no_cell).strip(), str(reviewer_name_cell).strip()))
    wb.close()

    if not rows_data:
        return {"applied": 0, "skipped": 0, "errors": []}

    # DB 일괄 조회
    mgmt_nos = [r[0] for r in rows_data]
    buildings = db.query(Building).filter(Building.mgmt_no.in_(mgmt_nos)).all()
    building_map = {b.mgmt_no: b for b in buildings}

    reviewer_names = list(set(r[1] for r in rows_data))
    users = db.query(User).filter(User.name.in_(reviewer_names), User.role == UserRole.REVIEWER).all()
    user_map = {u.name: u for u in users}

    # Reviewer 레코드 일괄 조회/생성
    user_ids = [u.id for u in users]
    existing_reviewers = db.query(Reviewer).filter(Reviewer.user_id.in_(user_ids)).all()
    reviewer_by_user = {r.user_id: r for r in existing_reviewers}

    for user in users:
        if user.id not in reviewer_by_user:
            reviewer = Reviewer(user_id=user.id)
            db.add(reviewer)
            db.flush()
            reviewer_by_user[user.id] = reviewer

    applied = 0
    skipped = 0
    errors = []

    for mgmt_no, reviewer_name in rows_data:
        building = building_map.get(mgmt_no)
        if not building:
            errors.append(f"{mgmt_no}: 관리번호를 찾을 수 없습니다")
            skipped += 1
            continue

        # 이름은 항상 저장
        building.assigned_reviewer_name = reviewer_name

        # 등록된 사용자면 reviewer_id도 연결
        user = user_map.get(reviewer_name)
        if user and user.id in reviewer_by_user:
            building.reviewer_id = reviewer_by_user[user.id].id

        # 단계 미설정 건물은 배정완료 단계로 설정
        if not building.current_phase:
            building.current_phase = "assigned"

        applied += 1

    db.commit()
    return {"applied": applied, "skipped": skipped, "errors": errors}
