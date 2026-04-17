"""검토서 내용 자동 추출 엔진

검토서 엑셀에서 판정 결과, 부적합 유형, 검토 의견 등을 추출하여
review_stages 테이블에 반영한다.

셀 위치 (2025-0005.xlsm 기준):
  H4: 검토결과
  D79: 적정성 검토 결과 (행이 이동할 수 있음)
  G81~: 판정결과 부적합 유형 1~3
"""

import re
from pathlib import Path

from openpyxl import load_workbook

from models.review_stage import ResultType

RESULT_MAP = {
    "적합": ResultType.PASS,
    "보완": ResultType.SUPPLEMENT,
    "부적합": ResultType.FAIL,
    "경미": ResultType.MINOR,
    "재계산": ResultType.SUPPLEMENT,  # 재계산도 보완으로 처리
}


def _cell_str(ws, coord: str) -> str:
    val = ws[coord].value
    if val is None:
        return ""
    return str(val).strip()


def _find_defect_type_values(ws, start_row: int = 79, max_row: int = 95) -> list[str]:
    """판정결과 부적합 유형 값 추출"""
    # 적정성 검토 결과 행 찾기
    result_row = start_row
    for row in range(start_row, max_row):
        val = ws.cell(row=row, column=2).value
        if val and "적정성 검토 결과" in str(val):
            result_row = row
            break

    # G열에서 부적합 유형 찾기
    defects = []
    for row in range(result_row + 2, max_row):
        val = ws.cell(row=row, column=7).value  # G열
        if val and str(val).strip():
            defects.append(str(val).strip())
        if len(defects) >= 3:
            break

    return defects


def extract_review_data(file_path: str | Path) -> dict:
    """검토서 엑셀에서 주요 데이터를 추출"""
    data: dict = {
        "result": None,
        "defect_type_1": None,
        "defect_type_2": None,
        "defect_type_3": None,
        "review_opinion": None,
        "reviewer_name": None,
    }

    try:
        wb = load_workbook(str(file_path), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        # 검토결과 (H4)
        result_val = _cell_str(ws, "H4")
        if result_val:
            data["result"] = RESULT_MAP.get(result_val)

        # 검토위원 (F4)
        reviewer = _cell_str(ws, "F4")
        if reviewer:
            data["reviewer_name"] = reviewer

        # 적정성 검토 결과 (D79 근처)
        for row in range(79, 95):
            val = ws.cell(row=row, column=2).value
            if val and "적정성 검토 결과" in str(val):
                opinion = ws.cell(row=row, column=4).value  # D열
                if opinion:
                    data["review_opinion"] = str(opinion).strip()
                break

        # 판정결과 부적합 유형
        defects = _find_defect_type_values(ws)
        if len(defects) > 0:
            data["defect_type_1"] = defects[0]
        if len(defects) > 1:
            data["defect_type_2"] = defects[1]
        if len(defects) > 2:
            data["defect_type_3"] = defects[2]

        wb.close()

    except Exception:
        pass

    return data
