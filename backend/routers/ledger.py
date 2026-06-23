"""통합관리대장 엑셀 Import/Export 라우터"""

import logging
from pathlib import Path

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from database import get_db
from dependencies import stream_upload_to_tempfile
from models.user import User, UserRole
from routers.auth import get_current_user, require_roles
from services.audit import log_action
from services.phase_transition import InvalidPhaseTransition
from engines.ledger_import import import_ledger
from engines.ledger_import_2025 import import_ledger_2025
from engines.ledger_import_unified import import_ledger_unified
from engines.ledger_import_technical import import_ledger_technical
from engines.ledger_import_selection import import_ledger_selection
from engines.ledger_export import export_ledger

router = APIRouter()
logger = logging.getLogger(__name__)


def _normalize_excel_text(value) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .replace("\n", "")
        .replace(" ", "")
        .replace("(", "")
        .replace(")", "")
        .strip()
    )


def _normalized_row_values(ws, row_no: int) -> list[str]:
    return [
        _normalize_excel_text(cell.value)
        for cell in next(ws.iter_rows(min_row=row_no, max_row=row_no), ())
        if _normalize_excel_text(cell.value)
    ]


def _is_technical_ledger_sheet(sheet_name: str, row3_values: list[str], row4_values: list[str]) -> bool:
    normalized_sheet_name = _normalize_excel_text(sheet_name)
    if "관리대장" not in normalized_sheet_name:
        return False

    row3_set = set(row3_values)
    row4_set = set(row4_values)
    return (
        ("모니터링관리번호" in row4_set or "관리번호" in row4_set)
        and "건축구분" in row4_set
        and "예비판정" in row3_set
        and "예비판정결과관리원입력" in row4_set
    )


def _detect_format(file_path: Path) -> str:
    """엑셀 파일의 시트 구조를 분석하여 import 방식 결정"""
    wb = load_workbook(str(file_path), data_only=True, read_only=True)
    sheet_names = wb.sheetnames

    has_2025 = False
    for sn in sheet_names:
        ws = wb[sn]
        normalized_sheet_name = _normalize_excel_text(sn)
        row1_values = _normalized_row_values(ws, 1)
        row3_values = _normalized_row_values(ws, 3)
        row4_values = _normalized_row_values(ws, 4)

        if "대상선정" in sn or (
            "관리번호" in row1_values and "건축구분" in row1_values
        ):
            wb.close()
            return "selection"

        if "관리대장" in sn and ("1차수" in sn or "1443" in sn or "2025" in sn):
            has_2025 = True
            continue

        if "통합관리대장" in normalized_sheet_name:
            # 통합 관리대장 시트의 Row 4 A열 확인하여 신형/구형 구분
            row4_a = _normalize_excel_text(ws.cell(row=4, column=1).value)
            wb.close()
            if row4_a and "관리번호" in row4_a:
                return "unified_new"  # 3차수 통합본 (Row4 헤더, Row5 데이터)
            return "unified_old"      # 기존 형식 (Row2 헤더, Row3 데이터)

        if _is_technical_ledger_sheet(sn, row3_values, row4_values):
            wb.close()
            return "technical"

    wb.close()
    if has_2025:
        return "2025"
    return "unified_old"


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(UserRole.CHIEF_SECRETARY)),
):
    """통합관리대장 엑셀 파일을 DB에 import (총괄간사 전용, 형식 자동 감지)"""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx)만 업로드 가능합니다")

    tmp_path = await stream_upload_to_tempfile(file, max_mb=20, suffix=".xlsx")  # 통합관리대장은 큼

    try:
        fmt = _detect_format(tmp_path)
        if fmt == "selection":
            result = import_ledger_selection(tmp_path, db)
        elif fmt == "technical":
            result = import_ledger_technical(tmp_path, db, actor_user_id=current_user.id)
        elif fmt == "unified_new":
            result = import_ledger_unified(tmp_path, db, actor_user_id=current_user.id)
        elif fmt == "2025":
            result = import_ledger_2025(tmp_path, db, actor_user_id=current_user.id)
        else:
            result = import_ledger(tmp_path, db, actor_user_id=current_user.id)
        summary = result if isinstance(result, dict) else {"result": str(result)}
        if isinstance(result, dict) and (
            result.get("imported")
            or result.get("updated")
            or result.get("final_result_updated")
        ):
            from routers.buildings import clear_stats_cache
            clear_stats_cache()
        log_action(
            db, current_user.id, "upload", "ledger",
            after_data={"filename": file.filename, "format": fmt, **summary},
        )
        db.commit()
        return result
    except InvalidPhaseTransition as exc:
        # 현행 importer 는 신규 건물에 전진 전환만 수행하므로 정상 데이터로는
        # 도달하지 않지만, 향후 기존 건물 갱신이 추가될 때 부분 커밋 상태로
        # 500 이 나가는 것을 막기 위한 방어선.
        db.rollback()
        raise HTTPException(status_code=400, detail=f"단계 전환 규칙 위반: {exc}")
    except Exception as exc:
        db.rollback()
        logger.exception("관리대장 업로드 실패")
        raise HTTPException(
            status_code=500,
            detail=f"관리대장 업로드 실패: {type(exc).__name__}: {exc}",
        )
    finally:
        tmp_path.unlink(missing_ok=True)


@router.get("/export")
def export_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(
            UserRole.TEAM_LEADER,
            UserRole.CHIEF_SECRETARY,
            UserRole.SECRETARY,
            UserRole.MANAGER,
        )
    ),
):
    """DB 데이터를 통합관리대장 형식의 엑셀로 export (팀장/총괄간사/간사/관리원)"""
    output = export_ledger(db)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=management_ledger.xlsx"},
    )
