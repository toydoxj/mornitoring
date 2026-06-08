"""사용자 관리 라우터"""

import re
import secrets
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

from email_validator import EmailNotValidError, validate_email
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from openpyxl import load_workbook
from pydantic import BaseModel, EmailStr, Field
from sqlalchemy import case, func
from sqlalchemy.orm import Session

from config import settings
from database import get_db
from dependencies import stream_upload_to_tempfile
from logging_config import log_event
from models.building import Building
from models.password_setup_token import PasswordSetupToken, TokenPurpose
from models.reviewer import Reviewer
from models.user import User, UserRole
from routers.auth import get_current_user, get_password_hash, require_roles
from services.invite import InviteResult, send_invites
from services.kakao import get_kakao_identity_status, get_kakao_token_status
from services.reviewer_link import ensure_reviewer_link

router = APIRouter()


# 혼동 문자(l/i, 0/o, 1) 제외한 알파벳·숫자
_PW_LETTERS = "abcdefghjkmnpqrstuvwxyz"  # 23자 (i, l 제외)
_PW_DIGITS = "23456789"                   # 8자 (0, 1 제외)


def _generate_initial_password() -> str:
    """일회용 초기 비밀번호 (영문 소문자 4 + 숫자 4 = 8자).

    전달·입력 편의 위해 혼동 문자(i/l, 0/o, 1) 제외.
    엔트로피 약 2^30으로 must_change_password=True 전제 하에 충분.
    """
    letters = "".join(secrets.choice(_PW_LETTERS) for _ in range(4))
    digits = "".join(secrets.choice(_PW_DIGITS) for _ in range(4))
    return letters + digits


ROLE_MAP = {
    "팀장": UserRole.TEAM_LEADER,
    "총괄간사": UserRole.CHIEF_SECRETARY,
    "간사": UserRole.SECRETARY,
    "관리원": UserRole.MANAGER,
    "검토위원": UserRole.REVIEWER,
    "team_leader": UserRole.TEAM_LEADER,
    "chief_secretary": UserRole.CHIEF_SECRETARY,
    "secretary": UserRole.SECRETARY,
    "manager": UserRole.MANAGER,
    "reviewer": UserRole.REVIEWER,
}


@dataclass
class BulkUserImportRow:
    row_idx: int
    name: str
    email: str
    role: UserRole
    phone: str | None = None
    group_no: int | None = None
    specialty: str | None = None


def _cell_to_text(value: object) -> str | None:
    """엑셀 셀 값을 사용자 입력 문자열로 정규화."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
    elif isinstance(value, float) and value.is_integer():
        text = str(int(value)).strip()
    else:
        text = str(value).strip()
    return text or None


def _normalize_header(value: object) -> str:
    """헤더 비교용 문자열 정규화: 공백/개행 제거 + 소문자."""
    text = _cell_to_text(value)
    if not text:
        return ""
    return re.sub(r"\s+", "", text).lower()


def _field_from_header(value: object) -> str | None:
    """현재 명단/기존 템플릿의 헤더명을 내부 필드명으로 매핑."""
    header = _normalize_header(value)
    if not header:
        return None
    if header in {"이름", "성명", "회원명", "name"}:
        return "name"
    if "이메일" in header or header in {"email", "e-mail", "mail"}:
        return "email"
    if header in {"역할", "권한", "role"}:
        return "role"
    if (
        "휴대전화" in header
        or "휴대폰" in header
        or "전화번호" in header
        or header in {"phone", "mobile", "tel"}
    ):
        return "phone"
    if header in {"조", "조번호", "group", "groupno", "group_no"}:
        return "group_no"
    if "특수분야" in header or "전문분야" in header or header == "specialty":
        return "specialty"
    return None


def _detect_user_import_columns(ws) -> tuple[int, dict[str, int]]:
    """이름/이메일 헤더가 있는 행을 찾아 필드별 컬럼 인덱스를 반환."""
    max_scan_row = min(ws.max_row or 1, 10)
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_row, values_only=True),
        start=1,
    ):
        columns: dict[str, int] = {}
        for col_idx, value in enumerate(row):
            field = _field_from_header(value)
            if field and field not in columns:
                columns[field] = col_idx
        if "name" in columns and "email" in columns:
            return row_idx, columns
    raise HTTPException(
        status_code=400,
        detail=(
            "엑셀에서 이름/이메일 헤더를 찾을 수 없습니다. "
            "지원 헤더: 회원명 또는 이름, 이메일"
        ),
    )


def _row_value(row: tuple[object, ...], columns: dict[str, int], field: str) -> str | None:
    col_idx = columns.get(field)
    if col_idx is None or col_idx >= len(row):
        return None
    return _cell_to_text(row[col_idx])


def _normalize_email(value: str) -> str:
    """일괄등록용 이메일 검증 및 정규화."""
    try:
        info = validate_email(value, check_deliverability=False)
    except EmailNotValidError as exc:
        raise ValueError(str(exc)) from exc
    return info.normalized.lower()


def _parse_group_no(value: str | None) -> int | None:
    """'6조', '6 조' 같은 엑셀 값을 1~7 정수로 변환."""
    if not value:
        return None
    match = re.search(r"[1-7]", value)
    if not match:
        raise ValueError("조는 1~7 범위여야 합니다")
    return int(match.group(0))


def _parse_user_import_rows(ws) -> tuple[list[BulkUserImportRow], list[str]]:
    """사용자 일괄등록 엑셀을 헤더 기반으로 파싱."""
    header_row_idx, columns = _detect_user_import_columns(ws)
    rows: list[BulkUserImportRow] = []
    errors: list[str] = []

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row_idx + 1, values_only=True),
        start=header_row_idx + 1,
    ):
        if not any(_cell_to_text(value) for value in row):
            continue

        name = _row_value(row, columns, "name")
        email_raw = _row_value(row, columns, "email")
        if not name and not email_raw:
            continue
        if not name or not email_raw:
            errors.append(f"{row_idx}행: 이름과 이메일은 필수입니다")
            continue

        try:
            email = _normalize_email(email_raw)
        except ValueError as exc:
            errors.append(f"{row_idx}행: 이메일 형식 오류({exc})")
            continue

        role_str = _row_value(row, columns, "role")
        role = ROLE_MAP.get(role_str, UserRole.REVIEWER) if role_str else UserRole.REVIEWER
        phone = _row_value(row, columns, "phone")
        specialty = _row_value(row, columns, "specialty")
        try:
            group_no = _parse_group_no(_row_value(row, columns, "group_no"))
        except ValueError as exc:
            errors.append(f"{row_idx}행: {exc}")
            continue

        rows.append(BulkUserImportRow(
            row_idx=row_idx,
            name=name,
            email=email,
            role=role,
            phone=phone,
            group_no=group_no,
            specialty=specialty,
        ))

    return rows, errors


def _resolve_group_no(user: User, reviewer_by_user: dict[int, Reviewer]) -> int | None:
    """역할별 group_no 단일 진실 분기.

    - REVIEWER: Reviewer.group_no (없으면 None)
    - 그 외:    User.group_no
    """
    if user.role == UserRole.REVIEWER:
        r = reviewer_by_user.get(user.id)
        return r.group_no if r else None
    return user.group_no


def _set_group_no(db: Session, user: User, group_no: int | None) -> None:
    """역할별 group_no 저장. 검토위원이면 Reviewer 행을 보장 + 갱신.

    1~7 범위 검증은 Pydantic 단에서 끝났으므로 여기선 값만 반영.
    """
    if user.role == UserRole.REVIEWER:
        reviewer = (
            db.query(Reviewer).filter(Reviewer.user_id == user.id).first()
        )
        if reviewer is None:
            reviewer = Reviewer(user_id=user.id, group_no=group_no)
            db.add(reviewer)
        else:
            reviewer.group_no = group_no
    else:
        user.group_no = group_no


def _apply_import_profile(db: Session, user: User, row: BulkUserImportRow) -> None:
    """일괄등록 행의 부가 정보를 역할에 맞는 테이블에 저장."""
    if user.role == UserRole.REVIEWER or row.group_no is not None:
        _set_group_no(db, user, row.group_no)
    if user.role != UserRole.REVIEWER or row.specialty is None:
        return
    reviewer = db.query(Reviewer).filter(Reviewer.user_id == user.id).first()
    if reviewer is None:
        reviewer = Reviewer(user_id=user.id)
        db.add(reviewer)
        db.flush()
    reviewer.specialty = row.specialty


def _find_user_by_email(db: Session, email: str) -> User | None:
    """이메일 대소문자 차이로 중복 계정이 생기지 않게 조회한다."""
    normalized = email.lower()
    return db.query(User).filter(func.lower(User.email) == normalized).first()


def _consume_open_setup_tokens(db: Session, user_id: int) -> None:
    """재등록 시 기존 초대/비밀번호 설정 링크를 모두 무효화한다."""
    now = datetime.now(timezone.utc)
    tokens = (
        db.query(PasswordSetupToken)
        .filter(
            PasswordSetupToken.user_id == user_id,
            PasswordSetupToken.consumed_at.is_(None),
        )
        .all()
    )
    for token in tokens:
        token.consumed_at = now


def _reset_kakao_state(user: User) -> None:
    """삭제 전 카카오 인증/매칭 상태가 재등록 계정에 남지 않도록 초기화한다."""
    user.kakao_id = None
    user.kakao_uuid = None
    user.kakao_login_uuid = None
    user.kakao_access_token = None
    user.kakao_refresh_token = None
    user.kakao_token_expires_at = None
    user.kakao_scopes_ok = None
    user.kakao_scopes_checked_at = None


def _restore_inactive_user_for_registration(
    db: Session,
    user: User,
    *,
    name: str,
    email: str,
    role: UserRole,
    phone: str | None,
    group_no: int | None,
    initial_password: str,
) -> None:
    """삭제(비활성)된 계정을 신규 등록 요청으로 재활성화한다."""
    user.name = name
    user.email = email
    user.role = role
    user.phone = phone
    user.password_hash = get_password_hash(initial_password)
    user.must_change_password = True
    user.is_active = True
    user.group_no = group_no if role != UserRole.REVIEWER else None
    _reset_kakao_state(user)
    _consume_open_setup_tokens(db, user.id)


# --- Pydantic 스키마 ---

class UserCreate(BaseModel):
    name: str
    email: EmailStr
    role: UserRole
    phone: str | None = None
    # 조 (1~7). 검토위원이면 Reviewer.group_no에, 그 외엔 User.group_no에 저장.
    group_no: int | None = Field(default=None, ge=1, le=7)


class UserUpdate(BaseModel):
    """사용자 정보 수정. 알 수 없는 필드는 명시적으로 거부."""
    model_config = {"extra": "forbid"}

    name: str | None = None
    phone: str | None = None
    role: UserRole | None = None
    is_active: bool | None = None
    # 조 (1~7 또는 null로 해제). 역할에 따라 적절한 곳에 저장됨.
    group_no: int | None = Field(default=None, ge=1, le=7)


class UserResponse(BaseModel):
    id: int
    name: str
    email: str
    role: UserRole
    phone: str | None = None
    is_active: bool
    # 카카오 상태
    kakao_linked: bool = False        # 카카오 로그인 완료(kakao_id 존재)
    kakao_matched: bool = False       # 친구 매칭 완료(kakao_uuid 존재)
    kakao_uuid: str | None = None
    kakao_login_uuid: str | None = None
    kakao_identity_status: str = "not_linked"
    # 카카오 토큰 상태 — not_linked/valid/refresh_needed/refresh_unavailable/invalid/...
    kakao_token_status: str | None = None
    kakao_token_expires_at: str | None = None
    # 카카오 동의 캐시 — ok | insufficient | unknown
    kakao_scopes_status: str | None = None
    kakao_scopes_checked_at: str | None = None
    # 비밀번호 셋업 상태 (목록 조회 시에만 채워짐)
    # setup_completed | pending | expired | not_invited
    setup_status: str | None = None
    last_invite_sent_at: str | None = None  # 마지막 토큰 발급 시각 (재발송 판단용)
    # 조 번호 (1~7 또는 null). 검토위원은 Reviewer.group_no, 그 외엔 User.group_no 기준.
    group_no: int | None = None

    model_config = {"from_attributes": True}


class UserCreateResponse(UserResponse):
    """사용자 신규 생성 응답 — 일회용 초기 비밀번호 포함."""
    initial_password: str


class BulkImportAccount(BaseModel):
    email: str
    name: str
    initial_password: str


class BulkImportResponse(BaseModel):
    created: int
    skipped: int
    errors: list[str] = []
    accounts: list[BulkImportAccount] = []
    # auto_send_invite=True로 호출된 경우만 채워짐
    invite_summary: "BulkInviteSummary | None" = None
    invite_results: list["BulkInviteResultItem"] = []


class ResetPasswordResponse(BaseModel):
    message: str
    initial_password: str


class SendInviteResponse(BaseModel):
    """단건 초대 발송 결과."""
    delivery: str  # "kakao" | "manual"
    setup_url: str | None  # manual/error에만 노출 (kakao 성공 시 None)
    expires_at: str
    purpose: str
    error: str | None = None


class BulkSendInviteRequest(BaseModel):
    user_ids: list[int]
    purpose: TokenPurpose = TokenPurpose.INITIAL_SETUP


class BulkInviteResultItem(BaseModel):
    user_id: int
    name: str
    delivery: str  # "kakao" | "manual"
    expires_at: str
    setup_url: str | None  # manual/error에만
    error: str | None


class BulkInviteSummary(BaseModel):
    total: int
    kakao_sent: int
    manual: int
    failed: int
    sender_error: str | None  # 카카오 발신자 토큰 미준비 등 공통 사유


class BulkSendInviteResponse(BaseModel):
    summary: BulkInviteSummary
    results: list[BulkInviteResultItem]


class ConsentReminderResponse(BaseModel):
    """카카오 동의 재안내 발송 결과."""
    delivery: str  # "kakao" | "manual"
    login_url: str
    error: str | None = None


class UserListResponse(BaseModel):
    items: list[UserResponse]
    total: int


# --- 엔드포인트 ---

@router.get("", response_model=UserListResponse)
def list_users(
    role: UserRole | None = None,
    include_inactive: bool = Query(False, description="비활성 사용자 포함 여부"),
    setup_status: str | None = Query(
        None, description="필터: setup_completed/pending/expired/not_invited"
    ),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(
            UserRole.TEAM_LEADER,
            UserRole.CHIEF_SECRETARY,
            UserRole.MANAGER,
        )
    ),
):
    """사용자 목록 조회.

    팀장/총괄간사는 운영 상태까지 보고, 관리원은 이름/조/권한/전화번호/이메일
    조회 용도로만 사용한다.
    """
    from datetime import datetime, timezone
    from models.password_setup_token import PasswordSetupToken
    from services.password_setup import _ensure_aware_utc

    is_manager = current_user.role == UserRole.MANAGER
    if is_manager:
        include_inactive = False
        setup_status = None

    query = db.query(User)
    if not include_inactive:
        query = query.filter(User.is_active.is_(True))
    if role:
        query = query.filter(User.role == role)

    # 정렬: 역할 순서(팀장→총괄간사→간사→검토위원) 후 이름 가나다.
    # SQL CASE 로 직접 정렬해 페이지네이션이 정확히 적용되게 한다.
    # (case({...}, value=...) 형태는 Enum 컬럼과의 매핑이 dialect 별로 차이가 있어
    #  명시적 WHEN 비교 형태로 작성)
    role_rank = case(
        (User.role == UserRole.TEAM_LEADER, 0),
        (User.role == UserRole.CHIEF_SECRETARY, 1),
        (User.role == UserRole.SECRETARY, 2),
        (User.role == UserRole.MANAGER, 3),
        (User.role == UserRole.REVIEWER, 4),
        else_=99,
    )
    query = query.order_by(role_rank, User.name)

    # 페이지네이션 전에 setup_status 필터를 적용하려면 모든 사용자에 대해
    # 상태를 계산해야 한다(60명 규모라 부담 없음). 필터 미지정 시는 페이지만 잘라낸다.
    if setup_status:
        candidates = query.all()
    else:
        total_pre = query.count()
        candidates = query.offset((page - 1) * size).limit(size).all()

    user_ids = [u.id for u in candidates]
    latest_tokens: dict[int, PasswordSetupToken] = {}
    reviewer_by_user: dict[int, Reviewer] = {}
    if user_ids:
        token_rows = (
            db.query(PasswordSetupToken)
            .filter(PasswordSetupToken.user_id.in_(user_ids))
            .order_by(
                PasswordSetupToken.user_id,
                PasswordSetupToken.created_at.desc(),
            )
            .all()
        )
        for t in token_rows:
            if t.user_id not in latest_tokens:
                latest_tokens[t.user_id] = t

        # 검토위원 group_no 통합 노출용 일괄 조회 (N+1 회피).
        for r in (
            db.query(Reviewer).filter(Reviewer.user_id.in_(user_ids)).all()
        ):
            reviewer_by_user[r.user_id] = r

    now = datetime.now(timezone.utc)

    def _compute_status(user: User) -> tuple[str, str | None]:
        """(setup_status, last_invite_sent_at_iso). 우선순위: 완료 > 토큰 상태."""
        if not user.must_change_password:
            tok = latest_tokens.get(user.id)
            return "setup_completed", tok.created_at.isoformat() if tok else None
        tok = latest_tokens.get(user.id)
        if tok is None:
            return "not_invited", None
        if tok.consumed_at is not None:
            # 토큰 소비됐는데 must_change_password=true인 비정상 상태.
            # 완료로 가리지 말고 재발송 필요로 표시한다 (운영자 액션 트리거).
            return "not_invited", tok.created_at.isoformat()
        if _ensure_aware_utc(tok.expires_at) <= now:
            return "expired", tok.created_at.isoformat()
        return "pending", tok.created_at.isoformat()

    items_all = [
        (u, *_compute_status(u))
        for u in candidates
    ]

    if setup_status:
        items_all = [t for t in items_all if t[1] == setup_status]
        total = len(items_all)
        items_all = items_all[(page - 1) * size : (page - 1) * size + size]
    else:
        total = total_pre

    def _scopes_status(u: User) -> str | None:
        if u.kakao_scopes_ok is True:
            return "ok"
        if u.kakao_scopes_ok is False:
            return "insufficient"
        return "unknown"

    items: list[UserResponse] = []
    for (u, status, last_sent) in items_all:
        if is_manager:
            items.append(UserResponse(
                id=u.id,
                name=u.name,
                email=u.email,
                role=u.role,
                phone=u.phone,
                is_active=u.is_active,
                group_no=_resolve_group_no(u, reviewer_by_user),
            ))
            continue
        kakao_token_status, kakao_token_expires_at = get_kakao_token_status(u)
        items.append(UserResponse(
            id=u.id, name=u.name, email=u.email, role=u.role,
            phone=u.phone, is_active=u.is_active,
            kakao_linked=bool(u.kakao_id),
            kakao_matched=bool(u.kakao_uuid),
            kakao_uuid=u.kakao_uuid,
            kakao_login_uuid=u.kakao_login_uuid,
            kakao_identity_status=get_kakao_identity_status(u),
            kakao_token_status=kakao_token_status,
            kakao_token_expires_at=kakao_token_expires_at,
            kakao_scopes_status=_scopes_status(u),
            kakao_scopes_checked_at=(
                u.kakao_scopes_checked_at.isoformat()
                if u.kakao_scopes_checked_at else None
            ),
            setup_status=status,
            last_invite_sent_at=last_sent,
            group_no=_resolve_group_no(u, reviewer_by_user),
        ))
    return UserListResponse(items=items, total=total)


@router.post("", response_model=UserCreateResponse, status_code=201)
def create_user(
    body: UserCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(UserRole.TEAM_LEADER, UserRole.CHIEF_SECRETARY)
    ),
):
    """사용자 등록 (팀장/총괄간사). 일회용 초기 비밀번호를 응답으로 반환."""
    email = str(body.email).lower()
    existing = _find_user_by_email(db, email)
    if existing and existing.is_active:
        raise HTTPException(status_code=409, detail="이미 등록된 이메일입니다")

    initial_password = _generate_initial_password()
    if existing:
        user = existing
        _restore_inactive_user_for_registration(
            db,
            user,
            name=body.name,
            email=email,
            role=body.role,
            phone=body.phone,
            group_no=body.group_no,
            initial_password=initial_password,
        )
    else:
        user = User(
            name=body.name,
            email=email,
            role=body.role,
            phone=body.phone,
            password_hash=get_password_hash(initial_password),
            must_change_password=True,
            # 검토위원이 아닌 경우의 group_no 만 User 컬럼에 직접 저장.
            group_no=body.group_no if body.role != UserRole.REVIEWER else None,
        )
        db.add(user)
        db.flush()

    # Reviewer 행 보장 + 배정된 건물 reviewer_id 자동 백필
    ensure_reviewer_link(db, user)
    if body.role == UserRole.REVIEWER:
        # ensure_reviewer_link 직후라 Reviewer 행이 보장됨.
        _set_group_no(db, user, body.group_no)
    db.commit()
    db.refresh(user)
    reviewer = (
        db.query(Reviewer).filter(Reviewer.user_id == user.id).first()
        if user.role == UserRole.REVIEWER else None
    )
    reviewer_map: dict[int, Reviewer] = {reviewer.user_id: reviewer} if reviewer else {}
    kakao_token_status, kakao_token_expires_at = get_kakao_token_status(user)
    return UserCreateResponse(
        id=user.id,
        name=user.name,
        email=user.email,
        role=user.role,
        phone=user.phone,
        is_active=user.is_active,
        kakao_linked=bool(user.kakao_id),
        kakao_matched=bool(user.kakao_uuid),
        kakao_uuid=user.kakao_uuid,
        kakao_login_uuid=user.kakao_login_uuid,
        kakao_identity_status=get_kakao_identity_status(user),
        kakao_token_status=kakao_token_status,
        kakao_token_expires_at=kakao_token_expires_at,
        initial_password=initial_password,
        group_no=_resolve_group_no(user, reviewer_map),
    )


@router.post("/import-excel", response_model=BulkImportResponse)
async def import_users_excel(
    file: UploadFile = File(...),
    auto_send_invite: bool = Query(False, description="등록 후 카카오/수동 초대 자동 발송"),
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(UserRole.TEAM_LEADER, UserRole.CHIEF_SECRETARY)
    ),
):
    """사용자 일괄 등록 (엑셀). 각 신규 계정의 일회용 초기 비밀번호를 응답에 포함.

    엑셀 형식: 헤더명 기반 자동 매핑.
    - 기존 템플릿: 이름, 이메일, 역할, 전화번호
    - 검토위원 명단: 조, 회원명, 휴대전화번호, 특수분야, 이메일
    auto_send_invite=true인 경우 신규 계정에 한해 send_invites를 호출하고
    invite_summary/invite_results를 응답에 포함한다.
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx)만 업로드 가능합니다")

    tmp_path = await stream_upload_to_tempfile(file, max_mb=10, suffix=".xlsx")

    try:
        try:
            wb = load_workbook(str(tmp_path), data_only=True, read_only=True)
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail="엑셀 파일을 읽을 수 없습니다. .xlsx 형식을 확인해주세요",
            ) from exc
        ws = wb.active
        import_rows, errors = _parse_user_import_rows(ws)

        created = 0
        skipped = 0
        accounts: list[BulkImportAccount] = []
        created_users: list[User] = []
        created_entries: list[tuple[User, BulkUserImportRow]] = []
        existing_users_by_email = {
            u.email.lower(): u for u in db.query(User).all()
        }
        seen_emails: set[str] = set()

        for row in import_rows:
            if row.email in seen_emails:
                skipped += 1
                continue
            seen_emails.add(row.email)

            initial_password = _generate_initial_password()
            existing = existing_users_by_email.get(row.email)
            if existing and existing.is_active:
                skipped += 1
                continue
            if existing:
                user = existing
                _restore_inactive_user_for_registration(
                    db,
                    user,
                    name=row.name,
                    email=row.email,
                    role=row.role,
                    phone=row.phone,
                    group_no=row.group_no,
                    initial_password=initial_password,
                )
            else:
                user = User(
                    name=row.name,
                    email=row.email,
                    role=row.role,
                    phone=row.phone,
                    password_hash=get_password_hash(initial_password),
                    must_change_password=True,
                    group_no=row.group_no if row.role != UserRole.REVIEWER else None,
                )
                db.add(user)
            accounts.append(BulkImportAccount(
                email=row.email, name=row.name, initial_password=initial_password,
            ))
            created += 1
            created_users.append(user)
            created_entries.append((user, row))

        # flush로 신규 user.id 확정 후 사용자별 Reviewer 자동 연결
        db.flush()
        for u, row in created_entries:
            ensure_reviewer_link(db, u)
            _apply_import_profile(db, u, row)

        db.commit()
        # commit 후 user.id 재확보
        for u in created_users:
            db.refresh(u)
        wb.close()

        invite_summary = None
        invite_results: list[BulkInviteResultItem] = []
        if auto_send_invite and created_users:
            summary, results = await send_invites(
                db,
                sender=current_user,
                targets=created_users,
                purpose=TokenPurpose.INITIAL_SETUP,
            )
            invite_summary = BulkInviteSummary(
                total=summary.total,
                kakao_sent=summary.kakao_sent,
                manual=summary.manual,
                failed=summary.failed,
                sender_error=summary.sender_error,
            )
            invite_results = [
                BulkInviteResultItem(
                    user_id=r.user_id,
                    name=r.name,
                    delivery=r.delivery,
                    expires_at=r.expires_at,
                    setup_url=r.setup_url,
                    error=r.error,
                )
                for r in results
            ]

        return BulkImportResponse(
            created=created,
            skipped=skipped,
            errors=errors,
            accounts=accounts,
            invite_summary=invite_summary,
            invite_results=invite_results,
        )

    finally:
        tmp_path.unlink(missing_ok=True)


@router.get("/{user_id}", response_model=UserResponse)
def get_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """사용자 상세 조회 — 본인 또는 운영 조회 권한자만.

    REVIEWER/SECRETARY가 임의의 user_id로 다른 사용자 정보(이메일/전화번호 등)를
    조회하지 못하도록 방어. 존재 자체를 노출하지 않기 위해 권한 미달 시 404.
    관리원은 이름/조/권한/전화번호/이메일 조회 용도라 운영 상태 필드는 비워서 반환한다.
    """
    is_admin = current_user.role in (
        UserRole.TEAM_LEADER,
        UserRole.CHIEF_SECRETARY,
        UserRole.MANAGER,
    )
    is_self = current_user.id == user_id
    if not (is_admin or is_self):
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    reviewer = (
        db.query(Reviewer).filter(Reviewer.user_id == user.id).first()
        if user.role == UserRole.REVIEWER else None
    )
    reviewer_map: dict[int, Reviewer] = {reviewer.user_id: reviewer} if reviewer else {}
    if current_user.role == UserRole.MANAGER and not is_self:
        return UserResponse(
            id=user.id,
            name=user.name,
            email=user.email,
            role=user.role,
            phone=user.phone,
            is_active=user.is_active,
            group_no=_resolve_group_no(user, reviewer_map),
        )
    kakao_token_status, kakao_token_expires_at = get_kakao_token_status(user)
    return UserResponse(
        id=user.id, name=user.name, email=user.email, role=user.role,
        phone=user.phone, is_active=user.is_active,
        kakao_linked=bool(user.kakao_id),
        kakao_matched=bool(user.kakao_uuid),
        kakao_uuid=user.kakao_uuid,
        kakao_login_uuid=user.kakao_login_uuid,
        kakao_identity_status=get_kakao_identity_status(user),
        kakao_token_status=kakao_token_status,
        kakao_token_expires_at=kakao_token_expires_at,
        group_no=_resolve_group_no(user, reviewer_map),
    )


@router.patch("/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    body: UserUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(UserRole.TEAM_LEADER, UserRole.CHIEF_SECRETARY)
    ),
):
    """사용자 정보 수정 (팀장/총괄간사만)"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")

    update_data = body.model_dump(exclude_unset=True)
    # group_no 는 역할에 따라 Reviewer.group_no 또는 User.group_no 에 저장.
    group_no_provided = "group_no" in update_data
    new_group_no = update_data.pop("group_no", None) if group_no_provided else None

    for key, value in update_data.items():
        setattr(user, key, value)
    if group_no_provided:
        _set_group_no(db, user, new_group_no)

    db.commit()
    db.refresh(user)
    # 응답에 통합 group_no 노출 — 갱신 후 reviewer를 다시 조회.
    reviewer = (
        db.query(Reviewer).filter(Reviewer.user_id == user.id).first()
        if user.role == UserRole.REVIEWER else None
    )
    reviewer_map: dict[int, Reviewer] = {reviewer.user_id: reviewer} if reviewer else {}
    kakao_token_status, kakao_token_expires_at = get_kakao_token_status(user)
    return UserResponse(
        id=user.id, name=user.name, email=user.email, role=user.role,
        phone=user.phone, is_active=user.is_active,
        kakao_linked=bool(user.kakao_id),
        kakao_matched=bool(user.kakao_uuid),
        kakao_uuid=user.kakao_uuid,
        kakao_login_uuid=user.kakao_login_uuid,
        kakao_identity_status=get_kakao_identity_status(user),
        kakao_token_status=kakao_token_status,
        kakao_token_expires_at=kakao_token_expires_at,
        group_no=_resolve_group_no(user, reviewer_map),
    )


@router.delete("/{user_id}", status_code=204)
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(UserRole.TEAM_LEADER, UserRole.CHIEF_SECRETARY)
    ),
):
    """사용자 삭제 (팀장/총괄간사).

    실제 사용자 행을 지우면 초대 토큰, 감사 로그, 게시글 작성자 등 FK 참조와 충돌한다.
    운영 이력 보존을 위해 삭제 요청은 비활성화로 처리하고, 기본 목록에서는 제외한다.
    """
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="본인 계정은 삭제할 수 없습니다")

    # 검토위원 배정 연결 해제
    reviewer = db.query(Reviewer).filter(Reviewer.user_id == user.id).first()
    if reviewer:
        db.query(Building).filter(Building.reviewer_id == reviewer.id).update(
            {"reviewer_id": None}
        )

    user.is_active = False
    db.commit()


@router.post("/{user_id}/send-invite", response_model=SendInviteResponse)
async def send_invite(
    user_id: int,
    purpose: TokenPurpose = TokenPurpose.INITIAL_SETUP,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(UserRole.TEAM_LEADER, UserRole.CHIEF_SECRETARY)
    ),
):
    """단건 초대 발송. 내부적으로 bulk 흐름과 같은 services/invite.send_invites 호출."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    if not user.is_active:
        raise HTTPException(status_code=400, detail="비활성 사용자입니다")

    _, results = await send_invites(
        db, sender=current_user, targets=[user], purpose=purpose
    )
    r = results[0]
    return SendInviteResponse(
        delivery=r.delivery,
        setup_url=r.setup_url,
        expires_at=r.expires_at,
        purpose=purpose.value,
        error=r.error,
    )


@router.post("/bulk-send-invite", response_model=BulkSendInviteResponse)
async def bulk_send_invite(
    body: BulkSendInviteRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(UserRole.TEAM_LEADER, UserRole.CHIEF_SECRETARY)
    ),
):
    """다중 사용자 일괄 초대 발송.

    - 비활성/존재하지 않는 사용자 ID는 결과에서 제외 (에러 카운트로 들어감)
    - 카카오 매칭자는 카카오 메시지 발송, 미매칭은 manual setup_url 반환
    - 카카오 발신자 토큰 미준비 시 모두 manual fallback
    - best-effort: 일부 실패가 전체 롤백을 일으키지 않음
    """
    if not body.user_ids:
        raise HTTPException(status_code=400, detail="user_ids가 비어 있습니다")
    if len(body.user_ids) > 200:
        raise HTTPException(status_code=400, detail="한 번에 최대 200명까지 발송 가능합니다")

    users = (
        db.query(User)
        .filter(User.id.in_(body.user_ids), User.is_active.is_(True))
        .all()
    )
    user_by_id = {u.id: u for u in users}
    targets = [user_by_id[uid] for uid in body.user_ids if uid in user_by_id]
    skipped = [uid for uid in body.user_ids if uid not in user_by_id]

    summary, results = await send_invites(
        db, sender=current_user, targets=targets, purpose=body.purpose
    )

    # 누락된 user_id는 실패 결과로 추가
    items = [
        BulkInviteResultItem(
            user_id=r.user_id,
            name=r.name,
            delivery=r.delivery,
            expires_at=r.expires_at,
            setup_url=r.setup_url,
            error=r.error,
        )
        for r in results
    ]
    for uid in skipped:
        items.append(
            BulkInviteResultItem(
                user_id=uid,
                name=f"#{uid}",
                delivery="manual",
                expires_at="",
                setup_url=None,
                error="사용자를 찾을 수 없거나 비활성 상태입니다",
            )
        )

    return BulkSendInviteResponse(
        summary=BulkInviteSummary(
            total=summary.total + len(skipped),
            kakao_sent=summary.kakao_sent,
            manual=summary.manual,
            failed=summary.failed + len(skipped),
            sender_error=summary.sender_error,
        ),
        results=items,
    )


@router.post("/{user_id}/send-consent-reminder", response_model=ConsentReminderResponse)
async def send_consent_reminder(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(UserRole.TEAM_LEADER, UserRole.CHIEF_SECRETARY)
    ),
):
    """카카오 동의 부족 사용자에게 재동의 안내 메시지 발송.

    - 카카오 매칭(`kakao_uuid`)된 사용자: 발신 관리자의 카카오 토큰으로 친구 메시지 발송
    - 미매칭 또는 발송 실패: manual — 운영자가 다른 채널로 login URL 안내
    - 사용자가 login URL에 진입하면 프론트에서 카카오 추가동의 흐름을 자동 시작함
    """
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    if not user.is_active:
        raise HTTPException(status_code=400, detail="비활성 사용자입니다")

    login_url = f"{settings.frontend_base_url.rstrip('/')}/login?kakao=consent"
    delivery = "manual"
    error: str | None = None

    if user.kakao_uuid:
        try:
            from services.kakao import ensure_valid_token, send_message_to_friends

            access_token = await ensure_valid_token(current_user, db)
            title = "건축구조안전 모니터링 — 카카오 동의 갱신 요청"
            description = (
                f"{user.name}님, 시스템에서 카카오 알림을 받으려면 "
                f"카카오 로그인 후 추가 동의가 필요합니다. 아래 링크를 눌러 다시 로그인해주세요."
            )
            result = await send_message_to_friends(
                access_token=access_token,
                receiver_uuids=[user.kakao_uuid],
                title=title,
                description=description,
                link_url=login_url,
            )
            if "error" in result:
                error = str(result.get("detail", "발송 실패"))
                log_event(
                    "error", "consent_reminder_kakao_failed",
                    user_id=user.id, reason=error,
                )
            else:
                delivery = "kakao"
                log_event(
                    "info", "consent_reminder_kakao_sent",
                    user_id=user.id,
                )
        except Exception as exc:
            error = f"카카오 발송 오류: {exc}"
            log_event(
                "error", "consent_reminder_kakao_exception",
                user_id=user.id, reason=str(exc),
            )

    return ConsentReminderResponse(
        delivery=delivery,
        login_url=login_url,
        error=error,
    )


@router.post("/{user_id}/reset-password", response_model=ResetPasswordResponse)
def reset_password(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(
        require_roles(UserRole.TEAM_LEADER, UserRole.CHIEF_SECRETARY)
    ),
):
    """비밀번호 초기화 (팀장/총괄간사). 일회용 초기 비밀번호를 응답으로 반환."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")

    initial_password = _generate_initial_password()
    user.password_hash = get_password_hash(initial_password)
    user.must_change_password = True
    db.commit()
    return ResetPasswordResponse(
        message=f"{user.name}의 비밀번호가 초기화되었습니다",
        initial_password=initial_password,
    )
