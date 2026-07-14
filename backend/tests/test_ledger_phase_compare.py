from datetime import date

from openpyxl import Workbook

from engines.ledger_phase_compare import (
    apply_final_results_from_ledger,
    compare_supplement_phase_with_db,
)
from models.phase_transition_log import PhaseTransitionLog
from models.user import UserRole


def _make_supplement_workbook(
    path,
    rows: list[dict[str, object]],
    management_rows: list[dict[str, object]] | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "통합 보완대장"
    ws["A4"] = "모니터링\n관리번호"
    for col in ("S", "U", "W", "Y", "AA"):
        ws[f"{col}4"] = "접수일자"
    for col in ("AV", "BD", "BL", "BT", "CB"):
        ws[f"{col}4"] = "보완자료검토서\n제출일"

    for row_no, row in enumerate(rows, start=5):
        for col, value in row.items():
            ws[f"{col}{row_no}"] = value

    if management_rows is not None:
        management_ws = wb.create_sheet("통합 관리대장")
        management_ws["A4"] = "모니터링\n관리번호"
        management_ws["CW4"] = "최종\n판정결과"
        for row_no, row in enumerate(management_rows, start=5):
            for col, value in row.items():
                management_ws[f"{col}{row_no}"] = value

    wb.save(path)


def test_compare_supplement_phase_with_db_detects_match_and_mismatch(
    db_session,
    make_user,
    make_building,
    tmp_path,
):
    user, _ = make_user(UserRole.CHIEF_SECRETARY)
    matched = make_building(mgmt_no="2026-0001")
    matched.current_phase = "supplement_2_received"
    matched.assigned_reviewer_name = "검토자1"
    matched.final_result = "pass"

    mismatched = make_building(mgmt_no="2026-0002")
    mismatched.current_phase = "supplement_1_received"
    mismatched.final_result = None

    completed = make_building(mgmt_no="2026-0004")
    completed.current_phase = "completed"
    completed.final_result = "fail_simple_error"

    db_session.commit()

    path = tmp_path / "phase_compare.xlsx"
    _make_supplement_workbook(
        path,
        [
            {"A": "2026-0001", "S": date(2026, 4, 1), "AV": date(2026, 4, 5), "U": date(2026, 4, 10)},
            {"A": "2026-0002", "S": date(2026, 4, 1), "AV": date(2026, 4, 5)},
            {"A": "2026-9999", "S": date(2026, 4, 1)},
            {"A": "2026-0003"},
            {"A": "2026-0004", "S": date(2026, 4, 1), "AV": date(2026, 4, 5)},
        ],
        management_rows=[
            {"A": "2026-0001", "CW": "적합"},
            {"A": "2026-0002", "CW": "부적합"},
            {"A": "2026-0004", "CW": "부적합"},
        ],
    )

    result = compare_supplement_phase_with_db(path, db_session, current_user=user)

    assert result["total_rows"] == 5
    # 단계 비교: CW 명기 행(0001·0002·0004)은 엑셀 단계가 completed로 바뀌어
    # DB completed(0004)만 일치, supplement 단계(0001·0002)는 불일치가 된다.
    assert result["phase_compare"]["matched"] == 1
    assert result["phase_compare"]["mismatched"] == 2
    assert result["phase_compare"]["missing_db"] == 2
    # 판정 비교: 0001·0004 일치, 0002 불일치
    assert result["final_result_compare"]["matched"] == 2
    assert result["final_result_compare"]["mismatched"] == 1

    items = {item["mgmt_no"]: item for item in result["items"]}
    # 0001: CW="적합" → 엑셀 단계 completed, 근거는 CW열. DB는 supplement_2_received → 단계 불일치, 판정 일치
    assert items["2026-0001"]["cw_completed"] is True
    assert items["2026-0001"]["status"] == "mismatch"
    assert items["2026-0001"]["excel_phase"] == "completed"
    assert items["2026-0001"]["evidence_column"] == "CW"
    assert items["2026-0001"]["final_result_status"] == "matched"
    assert items["2026-0001"]["reviewer_name"] == "검토자1"

    # 0002: CW="부적합" → 엑셀 단계 completed. DB는 supplement_1_received → 단계 불일치, 판정도 불일치
    assert items["2026-0002"]["status"] == "mismatch"
    assert items["2026-0002"]["excel_phase"] == "completed"
    assert items["2026-0002"]["db_phase"] == "supplement_1_received"
    assert items["2026-0002"]["phase_direction"] == "excel_ahead"
    assert items["2026-0002"]["final_result_status"] == "mismatch"
    assert items["2026-0002"]["excel_final_result"] == "fail_simple_error"
    assert items["2026-0002"]["final_result_column"] == "CW"

    assert items["2026-9999"]["status"] == "missing_db"

    # 0004: CW="부적합"(=DB fail_simple_error) → 엑셀 단계 completed, DB도 completed → 단계·판정 모두 일치
    assert items["2026-0004"]["status"] == "matched"
    assert items["2026-0004"]["matched"] is True
    assert items["2026-0004"]["cw_completed"] is True
    assert items["2026-0004"]["db_phase"] == "completed"
    assert items["2026-0004"]["excel_phase"] == "completed"
    assert items["2026-0004"]["final_result_status"] == "matched"
    assert items["2026-0004"]["excel_final_result"] == "fail_simple_error"


def test_phase_compare_endpoint_returns_summary(
    client,
    db_session,
    make_user,
    make_building,
    tmp_path,
):
    _, headers = make_user(UserRole.MANAGER)
    building = make_building(mgmt_no="2026-0101")
    building.current_phase = "supplement_1"
    db_session.commit()

    path = tmp_path / "phase_compare.xlsx"
    _make_supplement_workbook(
        path,
        [{"A": "2026-0101", "S": date(2026, 5, 1), "AV": date(2026, 5, 7)}],
    )

    with path.open("rb") as file:
        response = client.post(
            "/api/ledger/phase-compare",
            files={
                "file": (
                    "phase_compare.xlsx",
                    file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers=headers,
        )

    assert response.status_code == 200
    data = response.json()
    assert data["phase_compare"]["matched"] == 1
    assert data["items"][0]["mgmt_no"] == "2026-0101"
    assert data["items"][0]["status"] == "matched"
    assert data["items"][0]["final_result_status"] == "not_checked"


def test_apply_final_results_from_ledger_updates_only_mismatched_final_result(
    db_session,
    make_user,
    make_building,
    tmp_path,
):
    user, _ = make_user(UserRole.CHIEF_SECRETARY)
    building = make_building(mgmt_no="2026-0201")
    building.current_phase = "supplement_2"
    building.final_result = "pass"
    matched = make_building(mgmt_no="2026-0202")
    matched.current_phase = "supplement_1"
    matched.final_result = "fail_simple_error"
    transfer = make_building(mgmt_no="2026-0203")
    transfer.current_phase = "preliminary"
    transfer.final_result = None
    db_session.commit()

    path = tmp_path / "final_result_apply.xlsx"
    _make_supplement_workbook(
        path,
        [],
        management_rows=[
            {"A": "2026-0201", "CW": "부적합"},
            {"A": "2026-0202", "CW": "부적합"},
            {"A": "2026-0203", "CW": "차수이관"},
            {"A": "2026-9999", "CW": "적합"},
        ],
    )

    result = apply_final_results_from_ledger(
        path,
        db_session,
        actor_user_id=user.id,
    )
    db_session.commit()

    assert result["updated"] == 1
    assert result["matched"] == 1
    assert result["missing_db"] == 1
    assert result["excel_final_result_missing"] == 1

    db_session.refresh(building)
    db_session.refresh(matched)
    db_session.refresh(transfer)
    assert building.final_result == "fail_simple_error"
    assert building.current_phase == "completed"
    assert matched.final_result == "fail_simple_error"
    assert matched.current_phase == "supplement_1"
    assert transfer.final_result is None
    assert transfer.current_phase == "preliminary"

    phase_log = db_session.query(PhaseTransitionLog).filter_by(mgmt_no="2026-0201").one()
    assert phase_log.from_phase == "supplement_2"
    assert phase_log.to_phase == "completed"


def test_final_results_apply_endpoint_returns_summary(
    client,
    db_session,
    make_user,
    make_building,
    tmp_path,
):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    building = make_building(mgmt_no="2026-0301")
    building.current_phase = "preliminary"
    building.final_result = None
    db_session.commit()

    path = tmp_path / "final_result_endpoint.xlsx"
    _make_supplement_workbook(
        path,
        [],
        management_rows=[{"A": "2026-0301", "CW": "보완적합"}],
    )

    with path.open("rb") as file:
        response = client.post(
            "/api/ledger/final-results/apply",
            files={
                "file": (
                    "final_result_endpoint.xlsx",
                    file,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            headers=headers,
        )

    assert response.status_code == 200
    data = response.json()
    assert data["updated"] == 1
    db_session.refresh(building)
    assert building.final_result == "pass_supplement"
    assert building.current_phase == "completed"


def test_parse_final_result_maps_six_categories():
    from engines.ledger_phase_compare import _parse_final_result

    assert _parse_final_result("원적합") == "pass"
    assert _parse_final_result("적합") == "pass"
    assert _parse_final_result("보완적합") == "pass_supplement"
    # 구분 없는 "부적합"은 단순오류로 (정책 결정 1)
    assert _parse_final_result("부적합") == "fail_simple_error"
    assert _parse_final_result("부적합(단순오류)") == "fail_simple_error"
    assert _parse_final_result("부적합\n(단순오류)") == "fail_simple_error"
    assert _parse_final_result("부적합(재계산)") == "fail_recalculate"
    assert _parse_final_result("부적합\n(재계산)") == "fail_recalculate"
    assert _parse_final_result("부적합(미회신)") == "fail_no_response"
    assert _parse_final_result("대상제외") == "excluded"
    # 이관 계열은 최종 완료가 아니므로 제외 (None)
    assert _parse_final_result("차수이관") is None
    assert _parse_final_result("재보완(3차수 이관)") is None
    assert _parse_final_result("재보완\n(3차수 이관)") is None
    # 6분류에 없는 값은 원문 fallback 없이 None (잠재 결함 수정)
    assert _parse_final_result("알수없는값") is None
    assert _parse_final_result("재보완") is None


def test_compare_treats_cw_result_as_completed_and_falls_back_on_reamendment(
    db_session,
    make_user,
    make_building,
    tmp_path,
):
    user, _ = make_user(UserRole.CHIEF_SECRETARY)
    # CW 최종판정 명기 + 보완대장 3차 제출 → 엑셀 단계는 completed (제출 열 무시)
    cw_done = make_building(mgmt_no="2026-1001")
    cw_done.current_phase = "supplement_3"
    cw_done.final_result = "pass_supplement"
    # 재보완 이관 행 → CW 파싱 None, 엑셀 단계는 보완대장 제출 열 기준
    reamend = make_building(mgmt_no="2026-1002")
    reamend.current_phase = "supplement_1_received"
    reamend.final_result = None
    # CW 미기재 → 기존 동작 유지
    plain = make_building(mgmt_no="2026-1003")
    plain.current_phase = "supplement_1"
    plain.final_result = None
    db_session.commit()

    path = tmp_path / "cw_rule.xlsx"
    _make_supplement_workbook(
        path,
        [
            {
                "A": "2026-1001",
                "S": date(2026, 4, 1),
                "AV": date(2026, 4, 5),
                "U": date(2026, 4, 10),
                "BD": date(2026, 4, 15),
                "W": date(2026, 4, 20),
                "BL": date(2026, 4, 25),
            },
            {"A": "2026-1002", "S": date(2026, 4, 1)},
            {"A": "2026-1003", "S": date(2026, 4, 1), "AV": date(2026, 4, 5)},
        ],
        management_rows=[
            {"A": "2026-1001", "CW": "보완적합"},
            {"A": "2026-1002", "CW": "재보완(3차수 이관)"},
        ],
    )

    result = compare_supplement_phase_with_db(path, db_session, current_user=user)
    items = {item["mgmt_no"]: item for item in result["items"]}

    # 2026-1001: CW 완료 → 3차 제출 열 무시하고 엑셀 단계 completed, 근거는 CW열
    assert items["2026-1001"]["cw_completed"] is True
    assert items["2026-1001"]["excel_phase"] == "completed"
    assert items["2026-1001"]["status"] == "mismatch"  # DB는 supplement_3
    assert items["2026-1001"]["excel_final_result"] == "pass_supplement"
    assert items["2026-1001"]["final_result_status"] == "matched"
    assert items["2026-1001"]["evidence_column"] == "CW"

    # 2026-1002: 재보완 이관 → CW 파싱 None, 엑셀 단계는 제출 열 기준
    assert items["2026-1002"]["cw_completed"] is False
    assert items["2026-1002"]["excel_phase"] == "supplement_1_received"
    assert items["2026-1002"]["status"] == "matched"  # DB도 supplement_1_received
    assert items["2026-1002"]["excel_final_result"] is None
    assert items["2026-1002"]["final_result_status"] == "excel_final_result_missing"

    # 2026-1003: CW 미기재 → 기존 동작 (제출 열 기준 supplement_1)
    assert items["2026-1003"]["cw_completed"] is False
    assert items["2026-1003"]["excel_phase"] == "supplement_1"
    assert items["2026-1003"]["status"] == "matched"
    assert items["2026-1003"]["final_result_status"] == "not_checked"
