"""검토서 시트 구성 검증 회귀 테스트.

실제 업로드 실패 사례(2026-2708.xlsm): 검토위원이 그래프를 별도 시트로 옮기면서
생긴 Chartsheet가 첫 번째 시트가 되었다. 시트 수 오류를 기록하고도 계속 진행해
`ws[coord]`에서 TypeError('Chartsheet' object is not subscriptable) → 500이 났다.
그래서 시트가 2개 이상이면 셀 접근 전에 중단해야 한다.

(차트 시트 자체는 openpyxl로 만들면 재로드 시 깨져서 픽스처로 쓸 수 없다.
 여기서는 "셀 접근 전에 중단한다"는 핵심 동작을 워크시트 2개로 검증한다.)
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from engines.review_validator import validate_review_file


def _write_review_book(path: Path, *, extra_sheet: bool) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "검토서 (1차)"
    ws["C4"] = "2026-2708"
    ws["F4"] = "홍길동"
    ws["H4"] = "적합"
    ws["C5"] = "1차 적정성 검토"
    if extra_sheet:
        wb.create_sheet("Chart1")
    wb.save(path)


def _validate(path: Path):
    return validate_review_file(
        file_path=path,
        filename="2026-2708.xlsm",
        expected_mgmt_no="2026-2708",
        submitter_name="홍길동",
        expected_phase="preliminary",
    )


@pytest.fixture
def review_path(tmp_path: Path) -> Path:
    return tmp_path / "2026-2708.xlsm"


def test_시트가_2개면_셀검증_전에_중단한다(review_path: Path):
    _write_review_book(review_path, extra_sheet=True)

    result = _validate(review_path)

    assert result.is_valid is False
    assert len(result.errors) == 1, "시트 수 오류만 남기고 즉시 중단해야 한다"
    assert "시트가 2개입니다" in result.errors[0]
    # 어떤 시트를 지워야 하는지 알 수 있어야 한다
    assert "Chart1" in result.errors[0]


def test_시트가_1개면_시트수_오류가_없다(review_path: Path):
    _write_review_book(review_path, extra_sheet=False)

    result = _validate(review_path)

    assert not any("시트가" in e for e in result.errors)
