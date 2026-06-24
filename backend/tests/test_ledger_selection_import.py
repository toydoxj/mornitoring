import io

from openpyxl import Workbook

from engines.ledger_import_selection import import_ledger_selection
from engines.ledger_import_technical import import_ledger_technical
from engines.ledger_import_unified import import_ledger_unified
from models.audit_log import AuditLog
from models.building import Building
from models.phase_transition_log import PhaseTransitionLog
from models.review_opinion_detail import ReviewOpinionDetail
from models.review_severity_summary import ReviewSeveritySummary
from models.review_stage import PhaseType, ResultType, ReviewStage
from models.user import UserRole
from routers.ledger import _detect_format


def test_detect_format_selection_result(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "3차수 대상선정"
    ws.append(["관리번호", None, "건축구분", "시도명"])
    ws.append(["2026-0363", "강정임", "신축", "강원특별자치도"])
    path = tmp_path / "selection.xlsx"
    wb.save(path)

    assert _detect_format(path) == "selection"


def test_import_ledger_selection_reads_header_based_format(db_session, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "3차수 대상선정"
    ws.append([
        "관리번호",
        None,
        "건축구분",
        "대장구분",
        "시도명",
        "시군구명",
        "법정동명",
        "대지구분",
        "본번",
        "부번",
        "특수지번",
        "건물명",
        "연면적",
        "주구조",
        "기타구조",
        "주용도",
        "기타용도",
        "주지붕",
        "기타지붕",
        "높이",
        "지상층수",
        "지하층수",
        "설계자",
        "설계사무소",
        "공사(건물)명",
        "특수구조물 여부",
        "고층",
        "다중이용",
        "고위험",
        "준다중이용",
    ])
    ws.append([
        "2026-0363",
        "강정임",
        "신축",
        "일반",
        "강원특별자치도",
        "평창군",
        "미탄면 회동리",
        "대지",
        "1",
        "1",
        None,
        "은하수전망대 주건축물제1동",
        "639.63",
        "철골철근콘크리트구조",
        "철골철근콘크리트조",
        "관광휴게시설",
        None,
        "(철근)콘크리트",
        None,
        "13.15",
        "1",
        "0",
        "주식회사네임리스건축사사무소",
        "(주) 네임리스 건축사사무소",
        "은하수전망대",
        "O",
        None,
        None,
        "O",
        "O",
    ])
    path = tmp_path / "selection.xlsx"
    wb.save(path)

    result = import_ledger_selection(path, db_session)

    assert result == {
        "imported": 1,
        "skipped": 0,
        "errors": [],
        "sheet": "3차수 대상선정",
    }
    building = db_session.query(Building).filter_by(mgmt_no="2026-0363").one()
    assert building.building_type == "신축"
    assert building.sido == "강원특별자치도"
    assert building.sigungu == "평창군"
    assert float(building.gross_area) == 639.63
    assert float(building.height) == 13.15
    assert building.floors_above == 1
    assert building.floors_below == 0
    assert building.architect_name == "주식회사네임리스건축사사무소"
    assert building.architect_firm == "(주) 네임리스 건축사사무소"
    assert building.remarks == "은하수전망대"
    assert building.is_special_structure is True
    assert building.is_high_rise is None
    assert building.is_multi_use is None
    assert building.is_quasi_multi_use is True
    assert building.high_risk_type == "고위험"


def test_detect_format_technical_ledger_with_management_sheet_name(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "관리대장"
    ws["A3"] = "대상 건축물 개요(허가대장 DB)"
    ws["AT3"] = "예비판정"
    ws["BG3"] = "보완자료 검토"
    ws["A4"] = "모니터링\n관리번호"
    ws["F4"] = "건축구분"
    ws["AY4"] = "예비판정 결과   (관리원 입력)"
    ws["A5"] = "2026-0001"

    path = tmp_path / "technical.xlsx"
    wb.save(path)

    assert _detect_format(path) == "technical"


def test_detect_format_prefers_unified_for_integrated_management_sheet(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "통합 관리대장"
    ws["A3"] = "대상 건축물 개요(허가대장 DB)"
    ws["BQ3"] = "예비판정"
    ws["A4"] = "모니터링\n관리번호"
    ws["C4"] = "건축구분"
    ws["BV4"] = "예비판정 결과   (관리원 입력)"
    ws["A5"] = "2026-0001"

    path = tmp_path / "unified.xlsx"
    wb.save(path)

    assert _detect_format(path) == "unified_new"


def test_import_ledger_technical_reads_2026_distribution_format(db_session, tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "관리대장"

    ws["A3"] = "대상 건축물 개요(허가대장 DB)"
    ws["AT3"] = "예비판정"
    ws["BG3"] = "보완자료 검토"
    headers = {
        "A": "모니터링\n관리번호",
        "F": "건축구분",
        "H": "시도명",
        "I": "시군구명",
        "J": "법정동명",
        "K": "대지구분",
        "L": "본번",
        "M": "부번",
        "N": "특수지번",
        "O": "건물명",
        "P": "연면적",
        "Q": "주구조",
        "R": "기타구조",
        "S": "주용도",
        "T": "기타용도",
        "W": "높이",
        "X": "지상층수",
        "Y": "지하층수",
        "AD": "설계자",
        "AE": "설계사무소",
        "AO": "특수구조물 여부",
        "AP": "고층 여부",
        "AQ": "다중이용건축물 여부",
        "AR": "(고위험공종 여부)",
        "AS": "비고",
        "AT": "검토자",
        "AU": "1차검토의견\n(기술사회)",
        "AV": "부적합유형-1",
        "AW": "부적합유형-2",
        "AX": "부적합유형-3",
        "AY": "예비판정 결과   (관리원 입력)",
        "AZ": "예비 검토의견",
        "BG": "검토자",
        "BH": "판정 결과\n(이의신청\n반영)",
        "BI": "판정 결과\n부적합유형-1",
        "BJ": "판정 결과\n부적합유형-2",
        "BK": "판정 결과\n부적합유형-3",
        "BL": "보완자료 판정결과 검토의견\n(이의신청반영)",
        "BM": "비   고",
    }
    for col, header in headers.items():
        ws[f"{col}4"] = header

    values = {
        "A": "2026-0001",
        "B": "신호영",
        "F": "신축",
        "H": "경상북도",
        "I": "경산시",
        "J": "진량읍 문천리",
        "K": "대지",
        "L": "918",
        "M": "1",
        "N": "특수",
        "O": "테스트 건물",
        "P": "283.14",
        "Q": "일반철골구조",
        "R": "기타구조",
        "S": "공장",
        "T": "사무실",
        "W": "8.13",
        "X": "2",
        "Y": "0",
        "AD": "김규환",
        "AE": "동현 건축사사무소",
        "AO": "O",
        "AP": "-",
        "AR": "O",
        "AS": "비고 메모",
        "AT": "신호영",
        "AU": "단순오류",
        "AV": "하중오류-활하중",
        "AW": "도서작성 오류-도서 누락",
        "AX": "기타오류",
        "AY": "보완",
        "AZ": "[하중 적정성]\n1. 활하중 확인 필요\n[구조도면 작성 적정성]\n1. 구조도면 누락 확인",
        "BG": "신호영",
        "BH": "단순오류",
        "BI": "도서작성 오류-도서 불일치",
        "BJ": "도서작성 오류-도서 누락",
        "BK": "-",
        "BL": "보완 의견",
        "BM": "보완 비고",
    }
    for col, value in values.items():
        ws[f"{col}5"] = value

    path = tmp_path / "technical.xlsx"
    wb.save(path)

    result = import_ledger_technical(path, db_session)

    assert result == {
        "imported": 1,
        "updated": 0,
        "skipped": 0,
        "errors": [],
        "sheet": "관리대장",
    }
    building = db_session.query(Building).filter_by(mgmt_no="2026-0001").one()
    assert building.building_type == "신축"
    assert building.sido == "경상북도"
    assert building.sigungu == "경산시"
    assert float(building.gross_area) == 283.14
    assert float(building.height) == 8.13
    assert building.floors_above == 2
    assert building.floors_below == 0
    assert building.architect_name == "김규환"
    assert building.architect_firm == "동현 건축사사무소"
    assert building.assigned_reviewer_name == "신호영"
    assert building.is_special_structure is True
    assert building.is_high_rise is False
    assert building.high_risk_type == "고위험"
    assert building.current_phase == "supplement_1"

    stages = (
        db_session.query(ReviewStage)
        .filter_by(building_id=building.id)
        .order_by(ReviewStage.phase_order)
        .all()
    )
    assert [stage.phase for stage in stages] == [
        PhaseType.PRELIMINARY,
        PhaseType.SUPPLEMENT_1,
    ]
    assert stages[0].result == ResultType.SIMPLE_ERROR
    assert stages[0].review_opinion == (
        "[하중 적정성]\n1. 활하중 확인 필요\n"
        "[구조도면 작성 적정성]\n1. 구조도면 누락 확인"
    )
    assert stages[0].stage_remarks == "판정의견: 단순오류\n관리원 입력 예비판정 결과: 보완"
    assert stages[0].severity_l0_count == 0
    assert stages[0].severity_l3_count == 0
    assert stages[1].result == ResultType.SIMPLE_ERROR
    assert stages[1].review_opinion == "보완 의견"
    assert stages[1].stage_remarks == "보완 비고"
    assert stages[1].severity_l0_count == 0
    assert stages[1].severity_l3_count == 0

    summary_rows = (
        db_session.query(ReviewSeveritySummary)
        .join(ReviewStage, ReviewSeveritySummary.stage_id == ReviewStage.id)
        .filter(ReviewStage.building_id == building.id)
        .order_by(ReviewStage.phase_order, ReviewSeveritySummary.category)
        .all()
    )
    assert summary_rows == []

    detail_rows = (
        db_session.query(ReviewOpinionDetail)
        .join(ReviewStage, ReviewOpinionDetail.stage_id == ReviewStage.id)
        .filter(ReviewStage.building_id == building.id)
        .order_by(ReviewStage.phase_order, ReviewOpinionDetail.row_number)
        .all()
    )
    assert [
        (row.phase_group, row.category, row.severity, row.content)
        for row in detail_rows
    ] == [
        ("preliminary", "하중 적정성", "NA", "활하중 확인 필요"),
        ("preliminary", "구조도면 작성 적정성", "NA", "구조도면 누락 확인"),
        ("supplement", "기타의견", "NA", "보완 의견"),
    ]


def test_import_ledger_technical_updates_existing_rows_and_cleans_excel_newlines(
    db_session,
    tmp_path,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "관리대장"
    ws["A3"] = "대상 건축물 개요(허가대장 DB)"
    ws["AT3"] = "예비판정"
    ws["BG3"] = "보완자료 검토"
    ws["A4"] = "모니터링\n관리번호"
    ws["F4"] = "건축구분"
    ws["H4"] = "시도명"
    ws["O4"] = "건물명"
    ws["AT4"] = "검토자"
    ws["AU4"] = "1차검토의견\n(기술사회)"
    ws["AY4"] = "예비판정 결과   (관리원 입력)"
    ws["AZ4"] = "예비 검토의견"

    ws["A5"] = "2026-0001"
    ws["F5"] = "신축"
    ws["H5"] = "경상북도"
    ws["O5"] = "첫 건물명"
    ws["AT5"] = "신호영"
    ws["AU5"] = "단순오류"
    ws["AY5"] = "보완"
    ws["AZ5"] = "[하중 적정성]_x000D_ / 1. 최초 의견"

    path = tmp_path / "technical.xlsx"
    wb.save(path)

    first_result = import_ledger_technical(path, db_session)
    assert first_result["imported"] == 1
    assert first_result["updated"] == 0
    first_stage = (
        db_session.query(ReviewStage)
        .join(Building, ReviewStage.building_id == Building.id)
        .filter(Building.mgmt_no == "2026-0001", ReviewStage.phase == PhaseType.PRELIMINARY)
        .one()
    )
    first_stage.severity_l3_count = 1
    db_session.add(ReviewOpinionDetail(
        stage_id=first_stage.id,
        phase=PhaseType.PRELIMINARY.value,
        phase_group="preliminary",
        row_number=1,
        category="하중 적정성",
        severity="L3",
        content="기존 자동 분류",
    ))
    db_session.add(ReviewSeveritySummary(
        stage_id=first_stage.id,
        category="하중 적정성",
        severity="L3",
        count=1,
    ))
    db_session.commit()

    ws["O5"] = "수정된 건물명"
    ws["AU5"] = "재계산"
    ws["AZ5"] = "[하중 적정성]_x000D_ / 1. 수정 의견"
    wb.save(path)

    second_result = import_ledger_technical(path, db_session)
    assert second_result["imported"] == 0
    assert second_result["updated"] == 1
    assert second_result["skipped"] == 0

    building = db_session.query(Building).filter_by(mgmt_no="2026-0001").one()
    assert building.building_name == "수정된 건물명"
    stage = (
        db_session.query(ReviewStage)
        .filter_by(building_id=building.id, phase=PhaseType.PRELIMINARY)
        .one()
    )
    assert stage.result == ResultType.RECALCULATE
    assert "_x000D_" not in (stage.review_opinion or "")
    assert stage.review_opinion == "[하중 적정성]\n1. 수정 의견"
    assert stage.stage_remarks == "판정의견: 재계산\n관리원 입력 예비판정 결과: 보완"
    assert stage.severity_l3_count == 0
    assert stage.severity_l0_count == 0

    detail = db_session.query(ReviewOpinionDetail).filter_by(stage_id=stage.id).one()
    assert detail.category == "하중 적정성"
    assert detail.severity == "NA"
    assert detail.content == "수정 의견"
    assert db_session.query(ReviewSeveritySummary).filter_by(stage_id=stage.id).count() == 0


def test_opinion_detail_severity_can_be_assigned_manually(
    client,
    db_session,
    make_user,
    make_building,
):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    building = make_building(mgmt_no="OPINION-SEV-001")
    stage = ReviewStage(
        building_id=building.id,
        phase=PhaseType.PRELIMINARY,
        phase_order=0,
        result=ResultType.SIMPLE_ERROR,
    )
    db_session.add(stage)
    db_session.flush()
    detail = ReviewOpinionDetail(
        stage_id=stage.id,
        phase=PhaseType.PRELIMINARY.value,
        phase_group="preliminary",
        row_number=1,
        category="하중 적정성",
        severity="NA",
        content="활하중 확인 필요",
    )
    db_session.add(detail)
    db_session.commit()

    list_res = client.get(
        "/api/reviews/opinion-details",
        headers=headers,
        params={"severity": "NA", "search": "OPINION-SEV-001"},
    )
    assert list_res.status_code == 200
    assert list_res.json()["total"] == 1

    patch_res = client.patch(
        f"/api/reviews/opinion-details/{detail.id}/severity",
        headers=headers,
        json={"severity": "L2"},
    )
    assert patch_res.status_code == 200
    assert patch_res.json()["severity"] == "L2"

    db_session.refresh(stage)
    db_session.refresh(detail)
    assert stage.severity_l2_count == 1
    assert detail.severity == "L2"
    summary = db_session.query(ReviewSeveritySummary).filter_by(stage_id=stage.id).one()
    assert summary.category == "하중 적정성"
    assert summary.severity == "L2"
    assert summary.count == 1

    reset_res = client.patch(
        f"/api/reviews/opinion-details/{detail.id}/severity",
        headers=headers,
        json={"severity": "NA"},
    )
    assert reset_res.status_code == 200
    db_session.refresh(stage)
    db_session.refresh(detail)
    assert stage.severity_l2_count == 0
    assert detail.severity == "NA"
    assert db_session.query(ReviewSeveritySummary).filter_by(stage_id=stage.id).count() == 0


def test_import_ledger_unified_finalizes_from_result_report_and_warns_preliminary_mismatch(
    db_session,
    tmp_path,
):
    building = Building(
        mgmt_no="2026-9001",
        building_name="기존 건물",
        current_phase="preliminary",
        final_result="pass",
    )
    db_session.add(building)
    db_session.flush()
    stage = ReviewStage(
        building_id=building.id,
        phase=PhaseType.PRELIMINARY,
        phase_order=0,
        result=ResultType.PASS,
        review_opinion="기존 의견",
    )
    db_session.add(stage)
    db_session.commit()

    wb = Workbook()
    supp_ws = wb.active
    supp_ws.title = "통합 보완대장"
    supp_ws["A4"] = "모니터링\n관리번호"

    ws = wb.create_sheet("통합 관리대장")
    ws["A4"] = "모니터링\n관리번호"
    ws["C4"] = "건축구분"
    ws["BQ3"] = "예비판정"
    ws["BQ4"] = "검토자"
    ws["BX4"] = "1차검토의견\n(기술사회)"
    ws["BY4"] = "예비판정 결과   (관리원 입력)"
    ws["BZ4"] = "예비 검토의견"
    ws["CY3"] = "결과보고"
    ws["CY4"] = "최종\n판정결과"

    ws["A5"] = "2026-9001"
    ws["C5"] = "신축"
    ws["BQ5"] = "홍길동"
    ws["BX5"] = "재계산"
    ws["BY5"] = "보완"
    ws["BZ5"] = "변경된 예비 검토의견"
    ws["CY5"] = "부적합"

    path = tmp_path / "unified_shifted.xlsx"
    wb.save(path)

    result = import_ledger_unified(path, db_session)

    assert result["imported"] == 0
    assert result["updated"] == 1
    assert result["final_result_updated"] == 1
    assert result["warning_count"] == 2
    assert any("BR 1차검토의견" in warning for warning in result["warnings"])
    assert any("CW 최종완료 반영" in warning for warning in result["warnings"])

    db_session.refresh(building)
    db_session.refresh(stage)
    assert building.current_phase == "completed"
    assert building.final_result == "fail"
    assert stage.result == ResultType.RECALCULATE
    assert stage.review_opinion == "변경된 예비 검토의견"
    assert stage.stage_remarks == "판정의견: 재계산\n관리원 입력 예비판정 결과: 보완"

    actions = [row.action for row in db_session.query(AuditLog).order_by(AuditLog.id).all()]
    assert "ledger_preliminary_result_mismatch" in actions
    assert "ledger_final_result_update" in actions
    phase_log = db_session.query(PhaseTransitionLog).filter_by(mgmt_no="2026-9001").one()
    assert phase_log.from_phase == "preliminary"
    assert phase_log.to_phase == "completed"
    assert phase_log.trigger == "import"


def test_import_ledger_unified_dry_run_does_not_change_db(
    db_session,
    tmp_path,
):
    building = Building(
        mgmt_no="2026-9101",
        building_name="기존 건물",
        current_phase="preliminary",
        final_result="pass",
    )
    db_session.add(building)
    db_session.flush()
    stage = ReviewStage(
        building_id=building.id,
        phase=PhaseType.PRELIMINARY,
        phase_order=0,
        result=ResultType.PASS,
        review_opinion="기존 의견",
    )
    db_session.add(stage)
    db_session.commit()

    wb = Workbook()
    supp_ws = wb.active
    supp_ws.title = "통합 보완대장"
    supp_ws["A4"] = "모니터링\n관리번호"

    ws = wb.create_sheet("통합 관리대장")
    ws["A4"] = "모니터링\n관리번호"
    ws["C4"] = "건축구분"
    ws["BQ3"] = "예비판정"
    ws["BQ4"] = "검토자"
    ws["BR4"] = "1차검토의견\n(기술사회)"
    ws["BV4"] = "예비판정 결과   (관리원 입력)"
    ws["BW4"] = "예비 검토의견"
    ws["CW3"] = "결과보고"
    ws["CW4"] = "최종\n판정결과"

    ws["A5"] = "2026-9101"
    ws["C5"] = "신축"
    ws["BQ5"] = "홍길동"
    ws["BR5"] = "재계산"
    ws["BV5"] = "보완"
    ws["BW5"] = "변경 예정 의견"
    ws["CW5"] = "부적합"

    path = tmp_path / "unified_dry_run.xlsx"
    wb.save(path)

    result = import_ledger_unified(path, db_session, dry_run=True)

    assert result["mode"] == "validate"
    assert result["updated"] == 1
    assert result["final_result_updated"] == 1
    assert result["warning_count"] == 2
    assert any("반영 예정" in warning for warning in result["warnings"])

    db_session.refresh(building)
    db_session.refresh(stage)
    assert building.current_phase == "preliminary"
    assert building.final_result == "pass"
    assert stage.result == ResultType.PASS
    assert stage.review_opinion == "기존 의견"
    assert db_session.query(AuditLog).count() == 0
    assert db_session.query(PhaseTransitionLog).count() == 0


def test_import_ledger_unified_checks_only_updates_checked_fields(
    db_session,
    tmp_path,
):
    building = Building(
        mgmt_no="2026-9103",
        building_name="기존 건물",
        building_type="기존유형",
        current_phase="supplement_1",
        final_result=None,
    )
    db_session.add(building)
    db_session.flush()
    preliminary = ReviewStage(
        building_id=building.id,
        phase=PhaseType.PRELIMINARY,
        phase_order=0,
        result=ResultType.PASS,
        review_opinion="기존 예비 의견",
    )
    supplement = ReviewStage(
        building_id=building.id,
        phase=PhaseType.SUPPLEMENT_1,
        phase_order=1,
        result=ResultType.PASS,
        review_opinion="기존 보완 의견",
    )
    db_session.add_all([preliminary, supplement])
    db_session.commit()

    wb = Workbook()
    supp_ws = wb.active
    supp_ws.title = "통합 보완대장"
    supp_ws["A4"] = "모니터링\n관리번호"
    supp_ws["AO3"] = "1차"
    supp_ws["AP4"] = "판정 결과\n(이의신청반영)"
    supp_ws["A5"] = "2026-9103"
    supp_ws["AP5"] = "단순오류"

    ws = wb.create_sheet("통합 관리대장")
    ws["A4"] = "모니터링\n관리번호"
    ws["C4"] = "건축구분"
    ws["BQ3"] = "예비판정"
    ws["BQ4"] = "검토자"
    ws["BR4"] = "1차검토의견\n(기술사회)"
    ws["BV4"] = "예비판정 결과   (관리원 입력)"
    ws["BW4"] = "예비 검토의견"
    ws["CW3"] = "결과보고"
    ws["CW4"] = "최종\n판정결과"

    ws["A5"] = "2026-9103"
    ws["C5"] = "변경유형"
    ws["BQ5"] = "홍길동"
    ws["BR5"] = "재계산"
    ws["BV5"] = "보완"
    ws["BW5"] = "변경하면 안 되는 예비 의견"
    ws["CW5"] = "부적합"

    path = tmp_path / "unified_checks_only.xlsx"
    wb.save(path)

    result = import_ledger_unified(path, db_session, checks_only=True)

    assert result["mode"] == "checks"
    assert result["check_updated"] == 2
    assert result["final_result_updated"] == 1
    db_session.refresh(building)
    db_session.refresh(preliminary)
    db_session.refresh(supplement)
    assert building.building_type == "기존유형"
    assert building.current_phase == "completed"
    assert building.final_result == "fail"
    assert preliminary.result == ResultType.RECALCULATE
    assert preliminary.review_opinion == "기존 예비 의견"
    assert supplement.result == ResultType.SIMPLE_ERROR
    assert supplement.review_opinion == "기존 보완 의견"
    actions = [row.action for row in db_session.query(AuditLog).order_by(AuditLog.id).all()]
    assert actions.count("ledger_check_result_update") == 2
    assert "ledger_final_result_update" in actions


def test_validate_ledger_endpoint_does_not_apply_changes(
    client,
    db_session,
    make_user,
):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    building = Building(
        mgmt_no="2026-9102",
        current_phase="preliminary",
        final_result="pass",
    )
    db_session.add(building)
    db_session.flush()
    db_session.add(ReviewStage(
        building_id=building.id,
        phase=PhaseType.PRELIMINARY,
        phase_order=0,
        result=ResultType.PASS,
    ))
    db_session.commit()

    wb = Workbook()
    supp_ws = wb.active
    supp_ws.title = "통합 보완대장"
    supp_ws["A4"] = "모니터링\n관리번호"
    ws = wb.create_sheet("통합 관리대장")
    ws["A4"] = "모니터링\n관리번호"
    ws["C4"] = "건축구분"
    ws["BQ3"] = "예비판정"
    ws["BR4"] = "1차검토의견\n(기술사회)"
    ws["BV4"] = "예비판정 결과   (관리원 입력)"
    ws["BW4"] = "예비 검토의견"
    ws["CW3"] = "결과보고"
    ws["CW4"] = "최종\n판정결과"
    ws["A5"] = "2026-9102"
    ws["C5"] = "신축"
    ws["BR5"] = "재계산"
    ws["BV5"] = "보완"
    ws["CW5"] = "부적합"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    files = {
        "file": (
            "unified.xlsx",
            buf,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    res = client.post("/api/ledger/validate", headers=headers, files=files)

    assert res.status_code == 200
    body = res.json()
    assert body["mode"] == "validate"
    assert body["updated"] == 1
    assert body["final_result_updated"] == 1
    db_session.refresh(building)
    assert building.current_phase == "preliminary"
    assert building.final_result == "pass"


def test_apply_checks_endpoint_updates_only_checked_fields(
    client,
    db_session,
    make_user,
):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    building = Building(
        mgmt_no="2026-9104",
        building_type="기존유형",
        current_phase="preliminary",
    )
    db_session.add(building)
    db_session.flush()
    stage = ReviewStage(
        building_id=building.id,
        phase=PhaseType.PRELIMINARY,
        phase_order=0,
        result=ResultType.PASS,
        review_opinion="기존 의견",
    )
    db_session.add(stage)
    db_session.commit()

    wb = Workbook()
    supp_ws = wb.active
    supp_ws.title = "통합 보완대장"
    supp_ws["A4"] = "모니터링\n관리번호"
    ws = wb.create_sheet("통합 관리대장")
    ws["A4"] = "모니터링\n관리번호"
    ws["C4"] = "건축구분"
    ws["BQ3"] = "예비판정"
    ws["BR4"] = "1차검토의견\n(기술사회)"
    ws["BV4"] = "예비판정 결과   (관리원 입력)"
    ws["BW4"] = "예비 검토의견"
    ws["CW3"] = "결과보고"
    ws["CW4"] = "최종\n판정결과"
    ws["A5"] = "2026-9104"
    ws["C5"] = "변경유형"
    ws["BR5"] = "재계산"
    ws["BV5"] = "보완"
    ws["BW5"] = "변경하면 안 되는 의견"
    ws["CW5"] = "부적합"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    files = {
        "file": (
            "unified.xlsx",
            buf,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    }

    res = client.post("/api/ledger/apply-checks", headers=headers, files=files)

    assert res.status_code == 200
    body = res.json()
    assert body["mode"] == "checks"
    assert body["check_updated"] == 1
    assert body["final_result_updated"] == 1
    db_session.refresh(building)
    db_session.refresh(stage)
    assert building.building_type == "기존유형"
    assert building.current_phase == "completed"
    assert building.final_result == "fail"
    assert stage.result == ResultType.RECALCULATE
    assert stage.review_opinion == "기존 의견"


def test_import_ledger_unified_checks_supplement_sheet_results_against_db(
    db_session,
    tmp_path,
):
    building = Building(mgmt_no="2026-9002", current_phase="supplement_2")
    db_session.add(building)
    db_session.flush()
    db_session.add_all([
        ReviewStage(
            building_id=building.id,
            phase=PhaseType.SUPPLEMENT_1,
            phase_order=1,
            result=ResultType.PASS,
        ),
        ReviewStage(
            building_id=building.id,
            phase=PhaseType.SUPPLEMENT_2,
            phase_order=2,
            result=ResultType.RECALCULATE,
        ),
    ])
    db_session.commit()

    wb = Workbook()
    supp_ws = wb.active
    supp_ws.title = "통합 보완대장"
    supp_ws["A4"] = "모니터링\n관리번호"
    supp_ws["AO3"] = "1차"
    supp_ws["AP4"] = "판정 결과\n(이의신청반영)"
    supp_ws["AW3"] = "2차"
    supp_ws["AX4"] = "판정 결과\n(이의신청반영)"
    supp_ws["A5"] = "2026-9002"
    supp_ws["AP5"] = "단순오류"
    supp_ws["AX5"] = "보완"

    ws = wb.create_sheet("통합 관리대장")
    ws["A4"] = "모니터링\n관리번호"
    ws["A5"] = "2026-9002"

    path = tmp_path / "unified_supplement.xlsx"
    wb.save(path)

    result = import_ledger_unified(path, db_session)

    assert result["imported"] == 0
    assert result["updated"] == 1
    assert result["warning_count"] == 1
    assert "통합 보완대장 1차 판정결과 불일치" in result["warnings"][0]
    log = db_session.query(AuditLog).filter_by(action="ledger_supplement_result_mismatch").one()
    assert log.after_data["mgmt_no"] == "2026-9002"
    assert log.after_data["phase"] == "supplement_1"


def test_import_ledger_unified_excludes_supplement_appeal_from_comparison(
    db_session,
    tmp_path,
):
    building = Building(mgmt_no="2026-9003", current_phase="supplement_1")
    db_session.add(building)
    db_session.flush()
    db_session.add(ReviewStage(
        building_id=building.id,
        phase=PhaseType.SUPPLEMENT_1,
        phase_order=1,
        result=ResultType.PASS,
    ))
    db_session.commit()

    wb = Workbook()
    supp_ws = wb.active
    supp_ws.title = "통합 보완대장"
    supp_ws["A4"] = "모니터링\n관리번호"
    supp_ws["AO3"] = "1차"
    supp_ws["AP4"] = "판정 결과\n(이의신청반영)"
    supp_ws["A5"] = "2026-9003"
    supp_ws["AP5"] = "이의신청"

    ws = wb.create_sheet("통합 관리대장")
    ws["A4"] = "모니터링\n관리번호"
    ws["A5"] = "2026-9003"

    path = tmp_path / "unified_supplement_appeal.xlsx"
    wb.save(path)

    result = import_ledger_unified(path, db_session)

    assert result["updated"] == 1
    assert result["warning_count"] == 0
    assert (
        db_session.query(AuditLog)
        .filter_by(action="ledger_supplement_result_mismatch")
        .count()
        == 0
    )


def test_import_ledger_unified_checks_only_excludes_supplement_appeal(
    db_session,
    tmp_path,
):
    building = Building(mgmt_no="2026-9004", current_phase="supplement_1")
    db_session.add(building)
    db_session.flush()
    stage = ReviewStage(
        building_id=building.id,
        phase=PhaseType.SUPPLEMENT_1,
        phase_order=1,
        result=ResultType.PASS,
    )
    db_session.add(stage)
    db_session.commit()

    wb = Workbook()
    supp_ws = wb.active
    supp_ws.title = "통합 보완대장"
    supp_ws["A4"] = "모니터링\n관리번호"
    supp_ws["AO3"] = "1차"
    supp_ws["AP4"] = "판정 결과\n(이의신청반영)"
    supp_ws["A5"] = "2026-9004"
    supp_ws["AP5"] = "이의신청"

    ws = wb.create_sheet("통합 관리대장")
    ws["A4"] = "모니터링\n관리번호"
    ws["A5"] = "2026-9004"

    path = tmp_path / "unified_checks_only_supplement_appeal.xlsx"
    wb.save(path)

    result = import_ledger_unified(path, db_session, checks_only=True)

    db_session.refresh(stage)
    assert result["mode"] == "checks"
    assert result["check_updated"] == 0
    assert result["warning_count"] == 0
    assert stage.result == ResultType.PASS
    assert (
        db_session.query(AuditLog)
        .filter_by(action="ledger_check_result_update")
        .count()
        == 0
    )
