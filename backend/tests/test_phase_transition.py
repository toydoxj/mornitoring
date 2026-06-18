"""building.current_phase 전환 가드 + 영구 로그 회귀.

- transition_phase 헬퍼 매트릭스 (통과/거부) 단위 테스트
- distribution.receive 가 RECEIVE 트리거 + 로그를 남기는지
- reviews.upload 가 출발 _received 일 때만 phase 전환 / 그 외엔 phase 불변 + 로그 X
- POST /buildings/{id}/phase 가 매트릭스 외 변경을 400 으로 거부
- PATCH /buildings/{id} 에서 current_phase 필드는 더 이상 받지 않음 (서버 사이드에서 무시)
- assignments.assign 신규 배정 시 INITIAL 로그
- IMPORT 트리거: 엑셀 초기 적재의 전진 전환 허용 / 역행·미지 단계 거부
- import_ledger_unified 가 직접 대입 없이 trigger=import 로그를 남기는지 (통합)
"""

import pytest

from models.phase_transition_log import PhaseTransitionLog
from models.user import UserRole
from services.phase_transition import (
    InvalidPhaseTransition,
    next_phase_for,
    transition_phase,
)


# ===== 헬퍼 단위 매트릭스 =====

@pytest.mark.parametrize("trigger,from_phase,expected", [
    ("initial", None, "assigned"),
    ("initial", "", "assigned"),
    ("receive", "assigned", "doc_received"),
    ("receive", "preliminary", "supplement_1_received"),
    ("receive", "supplement_4", "supplement_5_received"),
    ("upload", "doc_received", "preliminary"),
    ("upload", "supplement_5_received", "supplement_5"),
    # _received가 아닌 출발에서 upload 트리거 → None (no-op)
    ("upload", "preliminary", None),
    ("upload", "assigned", None),
    # receive 트리거가 _received 출발로 들어오면 매트릭스 외 → None
    ("receive", "doc_received", None),
])
def test_next_phase_for_matrix(trigger, from_phase, expected):
    assert next_phase_for(trigger, from_phase) == expected


def test_transition_phase_initial(db_session, make_building):
    b = make_building(mgmt_no="PT-INIT")
    b.current_phase = None
    db_session.commit()

    log = transition_phase(db_session, b, to_phase="assigned", trigger="initial")
    db_session.commit()

    assert b.current_phase == "assigned"
    assert log is not None
    assert log.from_phase is None and log.to_phase == "assigned"
    assert log.mgmt_no == "PT-INIT"
    assert log.trigger == "initial"


def test_transition_phase_no_op_when_same(db_session, make_building):
    b = make_building(mgmt_no="PT-NOOP")
    b.current_phase = "preliminary"
    db_session.commit()

    log = transition_phase(db_session, b, to_phase="preliminary", trigger="manual")
    db_session.commit()

    assert log is None
    assert (
        db_session.query(PhaseTransitionLog)
        .filter(PhaseTransitionLog.mgmt_no == "PT-NOOP")
        .count()
        == 0
    )


def test_transition_phase_manual_allows_adjacent_previous(db_session, make_building):
    b = make_building(mgmt_no="PT-PREV")
    b.current_phase = "preliminary"
    db_session.commit()

    log = transition_phase(
        db_session,
        b,
        to_phase="doc_received",
        trigger="manual",
    )
    db_session.commit()

    assert b.current_phase == "doc_received"
    assert log is not None
    assert log.from_phase == "preliminary"
    assert log.to_phase == "doc_received"


def test_transition_phase_rejects_off_matrix(db_session, make_building):
    b = make_building(mgmt_no="PT-BAD")
    b.current_phase = "preliminary"
    db_session.commit()

    # 임의 점프 (preliminary → supplement_3) — 매트릭스 외
    with pytest.raises(InvalidPhaseTransition):
        transition_phase(db_session, b, to_phase="supplement_3", trigger="manual")


# ===== 통합: distribution.receive =====

def test_distribution_receive_logs_transition(client, db_session, make_user, make_building):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    b = make_building(mgmt_no="PT-RC1")
    b.current_phase = "assigned"
    db_session.commit()

    res = client.post(
        "/api/distribution/receive", headers=headers,
        json={"mgmt_nos": ["PT-RC1"]},
    )
    assert res.status_code == 200, res.text
    db_session.expire_all()

    logs = (
        db_session.query(PhaseTransitionLog)
        .filter(PhaseTransitionLog.mgmt_no == "PT-RC1")
        .order_by(PhaseTransitionLog.created_at)
        .all()
    )
    # assigned → doc_received 1건
    assert len(logs) == 1
    assert logs[0].trigger == "receive"
    assert logs[0].from_phase == "assigned"
    assert logs[0].to_phase == "doc_received"


def test_distribution_receive_initial_then_receive_when_phase_missing(
    client, db_session, make_user, make_building
):
    """phase 없는 building에 도서접수 시 INITIAL → RECEIVE 두 단계 모두 기록."""
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    b = make_building(mgmt_no="PT-INIT-RC")
    b.current_phase = None
    db_session.commit()

    res = client.post(
        "/api/distribution/receive", headers=headers,
        json={"mgmt_nos": ["PT-INIT-RC"]},
    )
    assert res.status_code == 200, res.text
    db_session.expire_all()

    logs = (
        db_session.query(PhaseTransitionLog)
        .filter(PhaseTransitionLog.mgmt_no == "PT-INIT-RC")
        .order_by(PhaseTransitionLog.created_at)
        .all()
    )
    triggers = [l.trigger for l in logs]
    assert triggers == ["initial", "receive"]


# ===== 통합: PATCH /buildings — current_phase 필드는 무시 =====

def test_patch_building_rejects_current_phase(
    client, db_session, make_user, make_building
):
    """PATCH 본문에 current_phase 가 들어오면 명시적으로 422 거부 (조용한 무시 X)."""
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    b = make_building(mgmt_no="PT-PATCH")
    b.current_phase = "assigned"
    db_session.commit()

    res = client.patch(
        f"/api/buildings/{b.id}", headers=headers,
        json={"current_phase": "supplement_3", "building_name": "이름변경"},
    )
    assert res.status_code == 422

    # phase 와 이름 둘 다 변경되지 않아야 한다 (트랜잭션 안 일어남)
    db_session.refresh(b)
    assert b.current_phase == "assigned"


def test_patch_building_normal_fields_still_work(
    client, db_session, make_user, make_building
):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    b = make_building(mgmt_no="PT-PATCH-OK")
    db_session.commit()

    res = client.patch(
        f"/api/buildings/{b.id}", headers=headers,
        json={"building_name": "이름변경"},
    )
    assert res.status_code == 200, res.text
    db_session.refresh(b)
    assert b.building_name == "이름변경"


# ===== 통합: POST /buildings/{id}/phase =====

def test_phase_change_endpoint_accepts_matrix(
    client, db_session, make_user, make_building
):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    b = make_building(mgmt_no="PT-MAN-OK")
    b.current_phase = "doc_received"
    db_session.commit()

    res = client.post(
        f"/api/buildings/{b.id}/phase", headers=headers,
        json={"to_phase": "preliminary", "reason": "데이터 복구"},
    )
    assert res.status_code == 200, res.text
    db_session.refresh(b)
    assert b.current_phase == "preliminary"

    log = (
        db_session.query(PhaseTransitionLog)
        .filter(PhaseTransitionLog.mgmt_no == "PT-MAN-OK")
        .one()
    )
    assert log.trigger == "manual"
    assert log.reason == "데이터 복구"


def test_phase_change_endpoint_allows_doc_received_to_assigned_correction(
    client, db_session, make_user, make_building
):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    b = make_building(mgmt_no="PT-MAN-CORRECT")
    b.current_phase = "doc_received"
    db_session.commit()

    res = client.post(
        f"/api/buildings/{b.id}/phase", headers=headers,
        json={"to_phase": "assigned", "reason": "잘못 접수된 도서 보정"},
    )
    assert res.status_code == 200, res.text
    db_session.refresh(b)
    assert b.current_phase == "assigned"

    log = (
        db_session.query(PhaseTransitionLog)
        .filter(PhaseTransitionLog.mgmt_no == "PT-MAN-CORRECT")
        .one()
    )
    assert log.trigger == "manual"
    assert log.from_phase == "doc_received"
    assert log.to_phase == "assigned"
    assert log.reason == "잘못 접수된 도서 보정"


def test_phase_change_endpoint_rejects_off_matrix(
    client, db_session, make_user, make_building
):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    b = make_building(mgmt_no="PT-MAN-BAD")
    b.current_phase = "preliminary"
    db_session.commit()

    res = client.post(
        f"/api/buildings/{b.id}/phase", headers=headers,
        json={"to_phase": "supplement_4"},
    )
    assert res.status_code == 400


def test_phase_change_endpoint_requires_admin(client, make_user, make_building):
    _, headers = make_user(UserRole.SECRETARY)  # 간사는 PATCH는 가능하나 phase 변경은 admin
    b = make_building(mgmt_no="PT-MAN-PERM")
    res = client.post(
        f"/api/buildings/{b.id}/phase", headers=headers,
        json={"to_phase": "doc_received"},
    )
    assert res.status_code == 403


# ===== 통합: assignments.assign INITIAL =====

def test_assignments_assign_logs_initial(
    client, db_session, make_user, make_reviewer, make_building
):
    _, headers = make_user(UserRole.CHIEF_SECRETARY)
    _, reviewer, _ = make_reviewer()
    b = make_building(mgmt_no="PT-ASSIGN")
    b.current_phase = None
    db_session.commit()

    res = client.post(
        "/api/assignments/assign", headers=headers,
        json={"building_id": b.id, "reviewer_id": reviewer.id},
    )
    assert res.status_code == 200, res.text
    db_session.refresh(b)
    assert b.current_phase == "assigned"

    log = (
        db_session.query(PhaseTransitionLog)
        .filter(PhaseTransitionLog.mgmt_no == "PT-ASSIGN")
        .one()
    )
    assert log.trigger == "initial"
    assert log.from_phase is None
    assert log.to_phase == "assigned"


# ===== IMPORT 트리거 (엑셀 초기 적재) =====

def test_transition_phase_import_from_empty(db_session, make_building):
    """신규 적재: 빈 phase → 임의의 알려진 단계 허용 + 로그 trigger=import."""
    b = make_building(mgmt_no="PT-IMP-NEW")
    b.current_phase = None
    db_session.commit()

    log = transition_phase(
        db_session, b, to_phase="supplement_1", trigger="import",
        actor_user_id=None, reason="ledger_import_unified",
    )
    db_session.commit()

    assert b.current_phase == "supplement_1"
    assert log is not None
    assert log.trigger == "import"
    assert log.from_phase is None and log.to_phase == "supplement_1"
    assert log.reason == "ledger_import_unified"


def test_transition_phase_import_progressive_forward(db_session, make_building):
    """같은 import 패스 안에서 preliminary → supplement_1 누진 설정 허용."""
    b = make_building(mgmt_no="PT-IMP-FWD")
    b.current_phase = None
    db_session.commit()

    transition_phase(db_session, b, to_phase="preliminary", trigger="import")
    transition_phase(db_session, b, to_phase="supplement_1", trigger="import")
    db_session.commit()

    assert b.current_phase == "supplement_1"
    logs = (
        db_session.query(PhaseTransitionLog)
        .filter(PhaseTransitionLog.mgmt_no == "PT-IMP-FWD")
        .order_by(PhaseTransitionLog.id)
        .all()
    )
    assert [(l.from_phase, l.to_phase) for l in logs] == [
        (None, "preliminary"),
        ("preliminary", "supplement_1"),
    ]


def test_transition_phase_import_rejects_backward(db_session, make_building):
    """import 로 기존 단계를 되돌리는 역행은 거부."""
    b = make_building(mgmt_no="PT-IMP-BACK")
    b.current_phase = "supplement_1"
    db_session.commit()

    with pytest.raises(InvalidPhaseTransition):
        transition_phase(db_session, b, to_phase="preliminary", trigger="import")
    assert b.current_phase == "supplement_1"


def test_transition_phase_import_keeps_completed(db_session, make_building):
    """최종완료 건은 관리대장 재업로드 시 데이터만 갱신하고 단계는 유지."""
    b = make_building(mgmt_no="PT-IMP-COMPLETE")
    b.current_phase = "completed"
    b.final_result = "pass"
    db_session.commit()

    log = transition_phase(db_session, b, to_phase="preliminary", trigger="import")
    db_session.commit()

    assert log is None
    assert b.current_phase == "completed"
    assert b.final_result == "pass"
    assert db_session.query(PhaseTransitionLog).filter_by(mgmt_no=b.mgmt_no).count() == 0


def test_transition_phase_import_rejects_unknown_phase(db_session, make_building):
    """알 수 없는 단계 문자열로의 import 는 거부 (엑셀 오염 방어)."""
    b = make_building(mgmt_no="PT-IMP-UNKNOWN")
    b.current_phase = None
    db_session.commit()

    with pytest.raises(InvalidPhaseTransition):
        transition_phase(db_session, b, to_phase="totally_bogus", trigger="import")
    assert b.current_phase is None


def test_import_ledger_unified_guards_and_logs(db_session, tmp_path):
    """importer 가 직접 대입으로 회귀하면 이 테스트가 잡는다.

    소형 통합 관리대장 xlsx 를 만들어 import_ledger_unified 실행:
    - 예비+1차 데이터 행 → preliminary → supplement_1 누진 전환 + 로그 2건
    - 예비만 있는 행 → preliminary + 로그 1건
    - 단계 데이터 없는 행 → phase None + 로그 0건
    """
    from openpyxl import Workbook

    from engines.ledger_import_unified import import_ledger_unified

    wb = Workbook()
    ws = wb.active
    ws.title = "통합 관리대장"
    # 관리번호는 9자 이상 + '-' 포함이어야 파싱된다 (DATA_START_ROW=5)
    ws["A5"] = "2026-0001"
    ws["BQ5"] = "홍길동"     # 예비 검토자
    ws["BV5"] = "적합"       # 예비판정
    ws["CI5"] = "홍길동"     # 1차 보완 검토자
    ws["CJ5"] = "재계산"     # 1차 보완 판정
    ws["A6"] = "2026-0002"
    ws["BV6"] = "적합"
    ws["A7"] = "2026-0003"
    path = tmp_path / "ledger.xlsx"
    wb.save(path)

    result = import_ledger_unified(path, db_session, actor_user_id=None)
    assert result["imported"] == 3 and not result["errors"]

    from models.building import Building

    phases = {
        b.mgmt_no: b.current_phase
        for b in db_session.query(Building).filter(
            Building.mgmt_no.in_(["2026-0001", "2026-0002", "2026-0003"])
        )
    }
    assert phases == {
        "2026-0001": "supplement_1",
        "2026-0002": "preliminary",
        "2026-0003": None,
    }

    logs = (
        db_session.query(PhaseTransitionLog)
        .filter(PhaseTransitionLog.mgmt_no.like("2026-000%"))
        .order_by(PhaseTransitionLog.id)
        .all()
    )
    assert [(l.mgmt_no, l.from_phase, l.to_phase) for l in logs] == [
        ("2026-0001", None, "preliminary"),
        ("2026-0001", "preliminary", "supplement_1"),
        ("2026-0002", None, "preliminary"),
    ]
    assert all(l.trigger == "import" for l in logs)
    assert all(l.reason == "ledger_import_unified" for l in logs)
