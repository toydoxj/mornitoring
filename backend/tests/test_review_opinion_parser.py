"""검토서 상세의견/심각도 파싱 회귀 테스트."""

from openpyxl import Workbook, load_workbook

from engines.review_extractor import extract_review_data
from engines.review_validator import validate_review_file


def _make_workbook(path, *, missing_severity: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "검토서 (1차)"
    ws["C4"] = "2026-0001"
    ws["F4"] = "이공우"
    ws["H4"] = "재계산"
    ws["C5"] = "1차 적정성 검토"
    ws["D17"] = "상세의견"
    ws["H17"] = "심각도"

    ws["B33"] = "2. 부재 설계 적정성"
    ws["C33"] = "구조설계 요소"
    ws["D33"] = "전이보 스트럽 간격 보완할 것."
    ws["H33"] = "2"
    ws["D34"] = "동결심도 깊이에 대한 단면 상세 자료 추가 보완할 것"
    if not missing_severity:
        ws["H34"] = "L3"

    ws["B78"] = "기타의견"
    ws["D78"] = "지반조사서 누락되었으니 보완 바람."
    ws["H78"] = "0"

    ws["B81"] = "적정성 검토 결과"
    ws["D81"] = "기존 수식 요약 텍스트"
    ws["F83"] = "판정결과 부적합 유형"
    ws["G83"] = "기타오류"
    wb.save(path)


def _clear_detail_opinions(path) -> None:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    for coord in ("B33", "C33", "D33", "H33", "D34", "H34", "B78", "D78", "H78"):
        ws[coord] = None
    wb.save(path)
    wb.close()


def test_validate_review_file_builds_opinion_from_detail_rows(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path)

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="preliminary",
    )

    assert result.is_valid is True
    assert result.extracted_data["review_opinion"] == (
        "[부재설계의 적정성 - 구조설계 요소]\n"
        "{L2} 전이보 스트럽 간격 보완할 것.\n"
        "{L3} 동결심도 깊이에 대한 단면 상세 자료 추가 보완할 것\n\n"
        "[기타의견]\n"
        "{L0} 지반조사서 누락되었으니 보완 바람."
    )
    assert result.extracted_data["severity_counts"] == {
        "L0": 1,
        "L1": 0,
        "L2": 1,
        "L3": 1,
        "L4": 0,
    }
    assert result.extracted_data["category_severity_counts"] == [
        {
            "category": "부재설계의 적정성 - 구조설계 요소",
            "severity": "L2",
            "count": 1,
        },
        {
            "category": "부재설계의 적정성 - 구조설계 요소",
            "severity": "L3",
            "count": 1,
        },
        {"category": "기타의견", "severity": "L0", "count": 1},
    ]
    assert result.extracted_data["opinion_entries"] == [
        {
            "row": 33,
            "category": "부재설계의 적정성 - 구조설계 요소",
            "severity": "L2",
            "content": "전이보 스트럽 간격 보완할 것.",
        },
        {
            "row": 34,
            "category": "부재설계의 적정성 - 구조설계 요소",
            "severity": "L3",
            "content": "동결심도 깊이에 대한 단면 상세 자료 추가 보완할 것",
        },
        {
            "row": 78,
            "category": "기타의견",
            "severity": "L0",
            "content": "지반조사서 누락되었으니 보완 바람.",
        },
    ]

    extracted = extract_review_data(path)
    assert extracted["review_opinion"] == result.extracted_data["review_opinion"]
    assert extracted["severity_counts"]["L3"] == 1
    assert extracted["category_severity_counts"] == result.extracted_data[
        "category_severity_counts"
    ]
    assert extracted["opinion_entries"] == result.extracted_data["opinion_entries"]


def test_validate_review_file_accepts_supplement_procedure_suffix(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path)

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws.title = "검토서 (2차)"
    ws["C5"] = "2차 적정성 검토(2)"
    wb.save(path)
    wb.close()

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="supplement_2",
    )

    assert result.is_valid is True


def test_validate_review_file_accepts_supplement_common_procedure(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path)

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws.title = "검토서 (2차)"
    ws["C5"] = "2차 적정성 검토"
    wb.save(path)
    wb.close()

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="supplement_4",
    )

    assert result.is_valid is True


def test_validate_review_file_rejects_wrong_supplement_procedure_suffix(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path)

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws.title = "검토서 (2차)"
    ws["C5"] = "2차 적정성 검토(2)"
    wb.save(path)
    wb.close()

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="supplement_3",
    )

    assert result.is_valid is False
    assert any(
        "'2차 적정성 검토(3)' 또는 '2차 적정성 검토'여야 합니다" in error
        for error in result.errors
    )


def test_validate_review_file_rejects_detail_without_severity(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path, missing_severity=True)

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="preliminary",
    )

    assert result.is_valid is False
    assert any("34행의 심각도" in error for error in result.errors)


def test_validate_review_file_accepts_pass_when_detail_is_empty(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path)
    _clear_detail_opinions(path)

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws["H4"] = "적합"
    ws["D81"] = "적합"
    ws["G83"] = "적합"
    wb.save(path)
    wb.close()

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="preliminary",
    )

    assert result.is_valid is True
    assert result.extracted_data["review_opinion"] is None
    assert result.extracted_data["severity_counts"] == {
        "L0": 0,
        "L1": 0,
        "L2": 0,
        "L3": 0,
        "L4": 0,
    }


def test_validate_review_file_rejects_non_pass_when_detail_is_empty(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path)
    _clear_detail_opinions(path)

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws["H4"] = "단순오류"
    ws["D81"] = None
    ws["G83"] = None
    wb.save(path)
    wb.close()

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="preliminary",
    )

    assert result.is_valid is False
    assert any("상세의견/심각도 기준 '적합'" in error for error in result.errors)


def test_validate_review_file_rejects_non_recalculate_with_l3_or_l4(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path)

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws["H4"] = "단순오류"
    wb.save(path)
    wb.close()

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="preliminary",
    )

    assert result.is_valid is False
    assert any("상세의견/심각도 기준 '재계산'" in error for error in result.errors)


def test_validate_review_file_rejects_non_simple_error_without_l3_l4(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path)

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws["H4"] = "재계산"
    ws["H34"] = "L1"
    wb.save(path)
    wb.close()

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="preliminary",
    )

    assert result.is_valid is False
    assert any("상세의견/심각도 기준 '단순오류'" in error for error in result.errors)


def test_validate_review_file_handles_inserted_detail_rows(tmp_path):
    path = tmp_path / "2026-0001.xlsm"
    _make_workbook(path)

    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    # 상세의견 구간에 행을 대량 삽입하면 기존 결과/부적합 유형 행도 아래로 밀린다.
    ws.insert_rows(35, amount=70)
    ws["D35"] = "삽입된 추가 상세의견입니다."
    ws["H35"] = "L4"
    wb.save(path)
    wb.close()

    result = validate_review_file(
        path,
        filename="2026-0001.xlsm",
        expected_mgmt_no="2026-0001",
        submitter_name="이공우",
        expected_phase="preliminary",
    )

    assert result.is_valid is True
    assert "{L4} 삽입된 추가 상세의견입니다." in result.extracted_data["review_opinion"]
    assert result.extracted_data["severity_counts"] == {
        "L0": 1,
        "L1": 0,
        "L2": 1,
        "L3": 1,
        "L4": 1,
    }
    assert result.extracted_data["defect_type_1"] == "기타오류"

    extracted = extract_review_data(path)
    assert extracted["defect_type_1"] == "기타오류"
    assert extracted["severity_counts"]["L4"] == 1
