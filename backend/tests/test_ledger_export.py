from openpyxl import load_workbook

from engines import ledger_export
from engines.ledger_export import export_ledger


def test_export_ledger_marks_related_tech_target_and_cooperation(
    db_session, make_building, tmp_path
):
    target_missing = make_building(mgmt_no="EXPORT-001")
    target_missing.floors_above = 6

    target_done = make_building(mgmt_no="EXPORT-002")
    target_done.is_special_structure = True
    target_done.struct_eng_name = "홍길동"

    not_target = make_building(mgmt_no="EXPORT-003")
    not_target.floors_above = 2

    db_session.commit()

    output = export_ledger(db_session, tmp_path / "ledger.xlsx")
    wb = load_workbook(output, data_only=True)
    ws = wb["통합 관리대장"]

    assert ws["AC2"].value == "협력대상"
    assert ws["AD2"].value == "협력여부"
    assert ws["AC3"].value == "Y"
    assert ws["AD3"].value == "N"
    assert ws["AC4"].value == "Y"
    assert ws["AD4"].value == "Y"
    assert ws["AC5"].value == "N"
    assert ws["AD5"].value == "N"

    wb.close()


def test_export_ledger_keeps_headers_and_column_widths(db_session, make_building, tmp_path):
    """write_only 스트리밍으로 바꾼 뒤에도 대분류/상세 헤더와 열 너비가 유지된다."""
    make_building(mgmt_no="EXPORT-010")
    db_session.commit()

    output = export_ledger(db_session, tmp_path / "ledger.xlsx")
    wb = load_workbook(output, data_only=True)
    ws = wb["통합 관리대장"]

    assert ws["C1"].value == "대상 건축물 개요(허가대장 DB)"
    assert ws["CJ1"].value == "결과보고"
    assert ws["A2"].value == "모니터링\n관리번호"
    assert ws["B2"].value == "검토\n위원"
    assert ws["CJ2"].value == "최종\n판정결과"
    assert ws["A3"].value == "EXPORT-010"
    assert ws.column_dimensions["A"].width == 15
    assert ws.column_dimensions["B"].width == 10
    assert ws.column_dimensions["K"].width == 25

    wb.close()


def test_export_ledger_streams_every_building_across_chunks(
    db_session, make_building, tmp_path, monkeypatch
):
    """건물을 청크로 나눠 읽어도 한 건도 빠뜨리지 않는다."""
    monkeypatch.setattr(ledger_export, "EXPORT_CHUNK_SIZE", 2)
    expected = [f"EXPORT-2{idx:02d}" for idx in range(5)]
    for mgmt_no in expected:
        make_building(mgmt_no=mgmt_no)
    db_session.commit()

    output = export_ledger(db_session, tmp_path / "ledger.xlsx")
    wb = load_workbook(output, data_only=True)
    ws = wb["통합 관리대장"]

    assert [ws.cell(row=row, column=1).value for row in range(3, 8)] == expected

    wb.close()
